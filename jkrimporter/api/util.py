import csv
import logging
from enum import Enum
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl.reader.excel import load_workbook
from pydantic import BaseModel

from jkrimporter.datasheets import (
    get_ilmoitustiedosto_headers,
    get_kaivotiedosto_headers,
    get_liete_ilmoitustiedosto_headers,
    get_liete_kuljetustiedosto_headers,
    get_lopetustiedosto_headers,
    get_paatostiedosto_headers,
    get_viemari_ilmoitustiedosto_headers,
    get_viemari_lopetustiedosto_headers,
    get_siirtotiedosto_headers,
    get_hapa_kohteet_headers,
    get_dvv_rakennus_headers,
    get_dvv_osoite_headers,
    get_dvv_omistaja_headers,
    get_dvv_asukas_headers,
    get_perusmaksu_headers,
    get_tiedontuottajat_headers,
    get_huoneistomaara_headers,
)

logger = logging.getLogger("jkr-sharepoint")


class FileType(str, Enum):
    PERUSMAKSUAINEISTO = "Perusmaksuaineisto"
    ILMOITUSTIEDOSTO = "Kompostointi"
    TAAJAMAT = "asukkaan taajamat"
    KOMPOSTOINNIN_LOPETUS = "Kompostoinnin_lopettami"
    PAATOSTIEDOSTO = "Paatokset"
    HAPATIEDOSTO = "Hapa-kohteet"
    HUONEISTOMAARAT = "Huoneistomäärät"
    TIEDONTUOTTAJAT = "Tiedontuottajat"
    DVVTIEDOSTO = "DVV-aineisto"
    KAIVOTIEDOT_ALKU = "Kaivotiedot_aloitus"
    KAIVOTIEDOT_LOPPU = "Kaivotiedot_lopetus"
    KULJETUSTIETO_LIETE = "Liete_kuljetustiedot"
    KULJETUSTIETO = "Kiintea_kuljetustiedot"
    LIETE_KOMPOSTOINTI = "Lietteen_kompostointi"
    LIETE_PELTOLEVITYS = "Lietteenpeltolevitys"
    VIEMARIVERKOSTO_ALKU = "Viemariverkosto"
    VIEMARIVERKOSTO_LOPPU = "Viemäriverkosto_lopetus"
    POSTINUMEROT = "PCF"
    UNKNOWN = ""


class FileInfo(BaseModel):
    class Config:
        extra = "allow"

    filename: str
    target_path: str
    size: Optional[int] = None
    rows: Optional[int] = None
    sharepoint_path: str = ""
    type: str = ""
    fileType: Optional[FileType] = None
    runnable: Optional[bool] = None


# Map FileType to expected Excel headers.
# Types without an entry here are accepted as-is (no header schema defined yet).
_HEADERS_BY_TYPE: Dict[FileType, List[str]] = {
    FileType.ILMOITUSTIEDOSTO: get_ilmoitustiedosto_headers(),
    FileType.KOMPOSTOINNIN_LOPETUS: get_lopetustiedosto_headers(),
    FileType.PAATOSTIEDOSTO: get_paatostiedosto_headers(),
    FileType.KAIVOTIEDOT_ALKU: get_kaivotiedosto_headers(),
    FileType.KAIVOTIEDOT_LOPPU: get_kaivotiedosto_headers(),
    FileType.KULJETUSTIETO_LIETE: get_liete_kuljetustiedosto_headers(),
    FileType.LIETE_KOMPOSTOINTI: get_liete_ilmoitustiedosto_headers(),
    FileType.VIEMARIVERKOSTO_ALKU: get_viemari_ilmoitustiedosto_headers(),
    FileType.VIEMARIVERKOSTO_LOPPU: get_viemari_lopetustiedosto_headers(),
    FileType.KULJETUSTIETO: get_siirtotiedosto_headers(),
    FileType.HAPATIEDOSTO: get_hapa_kohteet_headers(),
    FileType.PERUSMAKSUAINEISTO: get_perusmaksu_headers(),
    FileType.TIEDONTUOTTAJAT: get_tiedontuottajat_headers(),
    FileType.HUONEISTOMAARAT: get_huoneistomaara_headers(),
}

# DVV files contain multiple named sheets, each with its own expected headers.
_DVV_SHEETS: Dict[str, List[str]] = {
    "R1 rakennus": get_dvv_rakennus_headers(),
    "R3 osoite": get_dvv_osoite_headers(),
    "R4 omistaja": get_dvv_omistaja_headers(),
    "R9 huon asukk": get_dvv_asukas_headers(),
}

# Sorted longest-first so more specific prefixes are tried before shorter ones.
_FILE_TYPES_BY_PREFIX = sorted(FileType, key=lambda ft: len(ft.value), reverse=True)


def _detect_file_type(filename: str) -> Optional[FileType]:
    name_lower = filename.lower()
    for ft in _FILE_TYPES_BY_PREFIX:
        if name_lower.startswith(ft.value.lower()):
            return ft
    return None


def _verify_excel_headers(target_path: str, expected_headers: List[str]):
    """Return (missing_headers, data_row_count) for an Excel file."""
    try:
        workbook = load_workbook(filename=target_path, data_only=True, read_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        actual_headers = [cell.value for cell in sheet[1]]
        rows = max((sheet.max_row or 1) - 1, 0)
        workbook.close()
        return [h for h in expected_headers if h not in actual_headers], rows
    except Exception as e:
        logger.error("Tiedoston otsakkeiden lukeminen epäonnistui: %s – %s", target_path, e)
        return expected_headers, None


def _verify_csv_headers(target_path: str, expected_headers: List[str]):
    """Return (missing_headers, data_row_count) for a CSV file."""
    try:
        with open(target_path, mode="rb") as f:
            content = f.read()
        if content.startswith(b"\xef\xbb\xbf"):
            content = content[3:]
        decoded = content.decode("cp1252")
        reader = csv.DictReader(decoded.splitlines(), delimiter=";", quotechar='"', skipinitialspace=True)
        actual_lower = {h.lower() for h in (reader.fieldnames or [])}
        rows = sum(1 for _ in reader)
        return [h for h in expected_headers if h.lower() not in actual_lower], rows
    except Exception as e:
        logger.error("Tiedoston otsakkeiden lukeminen epäonnistui: %s – %s", target_path, e)
        return expected_headers, None


def _verify_dvv_sheets(target_path: str) -> bool:
    """Verify that all expected DVV sheets exist and have the correct headers."""
    try:
        workbook = load_workbook(filename=target_path, data_only=True, read_only=True)
        all_ok = True
        for sheet_name, expected_headers in _DVV_SHEETS.items():
            if sheet_name not in workbook.sheetnames:
                logger.error("DVV-tiedostosta puuttuu välilehti: %s", sheet_name)
                all_ok = False
                continue
            sheet = workbook[sheet_name]
            actual_headers = [cell.value for cell in sheet[1]]
            missing = [h for h in expected_headers if h not in actual_headers]
            if missing:
                logger.error(
                    "DVV-tiedosto: välilehti '%s', puuttuvat sarakeotsikot: %s",
                    sheet_name, missing,
                )
                all_ok = False
        workbook.close()
        return all_ok
    except Exception as e:
        logger.error("DVV-tiedoston lukeminen epäonnistui: %s – %s", target_path, e)
        return False


def _verify_csv_folder(folder_path: str, expected_headers: List[str]):
    """Return (all_ok, total_data_rows) for all .csv files inside a folder."""
    csv_files = list(Path(folder_path).glob("*.csv"))
    if not csv_files:
        logger.error("Kuljetustieto-kansiossa ei ole CSV-tiedostoja: %s", folder_path)
        return False, 0

    all_ok = True
    total_rows = 0
    for csv_path in csv_files:
        missing, rows = _verify_csv_headers(str(csv_path), expected_headers)
        if missing:
            logger.error(
                "Tiedosto: %s, puuttuvat sarakeotsikot: %s", csv_path.name, missing
            )
            all_ok = False
        total_rows += rows or 0

    return all_ok, total_rows


def _verify_dat_readable(target_path: str) -> bool:
    """Check that a fixed-width .dat file can be opened and is non-empty."""
    try:
        with open(target_path, mode="rb") as f:
            return len(f.read(1)) > 0
    except Exception as e:
        logger.error("Tiedoston lukeminen epäonnistui: %s – %s", target_path, e)
        return False


def verify_contents(raw: Dict[str, Any]) -> FileInfo:
    file = FileInfo(**raw)
    filename = file.filename
    target_path = file.target_path
    suffix = Path(filename).suffix.lower()

    detected_type = _detect_file_type(filename)

    if detected_type is None:
        logger.warning("Tuntematon tiedostotyyppi: %s", filename)
        file.fileType = None
        file.runnable = False
        file.rows = None
        return file

    file.fileType = detected_type
    logger.info("Tiedostotyyppi tunnistettu: %s → %s", filename, detected_type.name)

    if suffix == ".dat":
        file.runnable = _verify_dat_readable(target_path)
        file.rows = None
        return file

    if detected_type == FileType.DVVTIEDOSTO:
        file.runnable = _verify_dvv_sheets(target_path)
        file.rows = file.size
        return file
    
    expected_headers = _HEADERS_BY_TYPE.get(detected_type)
    
    # Kuljetustieto target_path is a folder containing multiple CSV files.
    if detected_type == FileType.KULJETUSTIETO:
        folder_headers = _HEADERS_BY_TYPE.get(FileType.KULJETUSTIETO, [])
        all_ok, total_rows = _verify_csv_folder(target_path, folder_headers)
        file.runnable = all_ok
        file.rows = total_rows
        return file

    if expected_headers:
        if suffix == ".csv":
            missing, rows = _verify_csv_headers(target_path, expected_headers)
        else:
            missing, rows = _verify_excel_headers(target_path, expected_headers)
        file.rows = rows
        if missing:
            logger.error(
                "Tiedosto: %s, puuttuvat sarakeotsikot: %s", filename, missing
            )
            file.runnable = False
        else:
            file.runnable = True
    else:
        # No header schema defined for this type, deny to inform user of possible misnaming.
        file.runnable = False
        file.rows = None

    return file
