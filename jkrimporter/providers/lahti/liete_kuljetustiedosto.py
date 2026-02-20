"""
LIETE-kuljetustiedoston lukija.

Lukee LIETE-rekisterin kuljetustiedot Excel-tiedostoista.
"""

import logging
from pathlib import Path
from typing import List

from openpyxl.reader.excel import load_workbook
from pydantic import ValidationError

from jkrimporter.datasheets import get_liete_kuljetustiedosto_headers
from jkrimporter.providers.lahti.liete_models import LieteKuljetusRow

logger = logging.getLogger(__name__)


class LieteKuljetustiedosto:
    """
    Lukee LIETE-kuljetustiedot Excel-tiedostoista.
    
    LIETE-kuljetustiedostot ovat muotoa:
    - Liete_kuljetustiedot_2024Q1.xlsx
    - Liete_kuljetustiedot_2024Q2.xlsx
    jne.
    
    Tiedostot sisältävät yhden välilehden jossa on kuljetustiedot.
    """
    
    def __init__(self, file_path: Path):
        """
        Alustaa LIETE-kuljetustiedoston lukijan.
        
        Args:
            file_path: Polku LIETE-kuljetustiedostoon (Excel)
        """
        self._file_path = Path(file_path)
        self._failed_rows: List[dict] = []
        
        if not self._file_path.exists():
            raise FileNotFoundError(f"Tiedostoa ei löydy: {self._file_path}")
        
        if self._file_path.suffix not in ['.xlsx', '.xls']:
            raise ValueError(f"Tiedoston pitää olla Excel-tiedosto (.xlsx tai .xls): {self._file_path}")
    
    @property
    def kuljetustiedot(self):
        """
        Iteroi kuljetustiedot tiedostosta.
        
        Yields:
            LieteKuljetusRow: Yksittäinen kuljetusrivi
        """
        logger.info(f"Luetaan LIETE-kuljetustiedot: {self._file_path}")
        
        workbook = load_workbook(filename=self._file_path, data_only=True, read_only=True)
        sheet = workbook.active
        
        # Lue otsikkorivi
        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value).strip())

        logger.info(f"Löydettiin {len(headers)} saraketta")

        # Validoi otsikot (case-insensitive, koska osa tiedostoista käyttää ISOJA KIRJAIMIA)
        expected_headers = get_liete_kuljetustiedosto_headers()
        actual_lower = {h.lower() for h in headers}
        missing_headers = [h for h in expected_headers if h.lower() not in actual_lower]
        if missing_headers:
            workbook.close()
            print(f"Tiedosto: {self._file_path}, puuttuvat sarakeotsikot: {missing_headers}")
            raise RuntimeError(
                f"LIETE-kuljetustiedostosta puuttuu oletettuja sarakeotsikoita: {missing_headers}"
            )

        # Luo case-insensitive mapping: tiedoston header -> odotettu alias (Pydantic)
        expected_lower_map = {h.lower(): h for h in expected_headers}
        header_normalize_map = {}
        for actual in headers:
            expected = expected_lower_map.get(actual.lower())
            if expected:
                header_normalize_map[actual] = expected

        # Lue datarivit
        row_count = 0
        success_count = 0
        error_count = 0

        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_count += 1

            # Luo dictionary riveistä, normalisoi avaimet Pydantic-aliaksiin
            row_dict = {}
            for i, value in enumerate(row):
                if i < len(headers):
                    key = header_normalize_map.get(headers[i], headers[i])
                    row_dict[key] = value
            
            # Ohita tyhjät rivit
            if all(v is None or str(v).strip() == "" for v in row_dict.values()):
                continue
            
            try:
                kuljetus_row = LieteKuljetusRow.parse_obj(row_dict)
                success_count += 1
                yield kuljetus_row
                
            except ValidationError as e:
                error_count += 1
                error_msg = "; ".join(
                    f"{''.join(error['loc'])}: {error['msg']}" 
                    for error in e.errors()
                )
                logger.warning(
                    f"Rivin {row_count} validointi epäonnistui: {error_msg}\n"
                    f"Data: {row_dict}"
                )
                self._failed_rows.append({
                    'row_number': row_count,
                    'error': error_msg,
                    'data': row_dict
                })
        
        workbook.close()
        
        logger.info(
            f"Luettiin {row_count} riviä: "
            f"{success_count} onnistui, {error_count} epäonnistui"
        )
    
    def get_failed_rows(self) -> List[dict]:
        """
        Palauttaa epäonnistuneet rivit.
        
        Returns:
            Lista epäonnistuneista riveistä
        """
        return self._failed_rows
