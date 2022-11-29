import logging
from pathlib import Path

from openpyxl.reader.excel import load_workbook

from jkrimporter.datasheets import ExcelCombinedFileSheetCollection, SiirtotiedostoSheet
from jkrimporter.providers.lahti.models import Asiakas


logger = logging.getLogger(__name__)


class AsiakastiedotSheet(SiirtotiedostoSheet[Asiakas]):
    @staticmethod
    def _obj_from_dict(data):
        return Asiakas.parse_obj(data)


class LahtiSiirtotiedosto:
    # Lahti has no set sheet names. It has a directory with different sheets
    # for different providers, all having identical format.

    def __init__(self, path):
        self._sheet_collection = ExcelCombinedFileSheetCollection(path)

    @classmethod
    def readable_by_me(cls, path):
        p = Path(path)
        for f in p.iterdir():
            if f.is_file() and f.suffix == ".xlsx":
                try:
                    workbook = load_workbook(filename=f, data_only=True, read_only=True)
                    sheets = workbook.sheetnames
                    if "in" in sheets:
                        return True
                except Exception:
                    pass
        return False

    @property
    def asiakastiedot(self):
        return AsiakastiedotSheet(self._sheet_collection)
