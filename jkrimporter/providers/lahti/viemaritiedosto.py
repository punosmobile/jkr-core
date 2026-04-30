import asyncio
import logging
import os
from pathlib import Path
from typing import Dict, List

import openpyxl

from openpyxl.reader.excel import load_workbook
from pydantic import ValidationError


from jkrimporter.conf import (
    get_kohdentumattomat_viemari_ilmoitus_filename,
    get_kohdentumattomat_viemarin_lopetus_filename
)
from jkrimporter.datasheets import (
    get_viemari_lopetustiedosto_headers,
    get_viemari_ilmoitustiedosto_headers
)
from jkrimporter.providers.lahti.models import ViemariIlmoitus, ViemariLopetusIlmoitus
from jkrimporter.api import sharepoint as sp

logger = logging.getLogger(__name__)

class ViemariIlmoitustiedosto:
    def __init__(self, path):
        self._path = path

    @classmethod
    def readable_by_me(cls, path):
        file = Path(path)
        if file.is_file() and file.suffix == ".xlsx":
            try:
                workbook = load_workbook(filename=file, data_only=True, read_only=True)
                if len(workbook.sheetnames):
                    return True
            except Exception:
                pass
        return False

    @property
    def viemariilmoitukset(self) -> List[ViemariIlmoitus]:
        lopetus_list = []
        missing_headers_list = []
        failed_validations = []

        workbook = load_workbook(self._path)
        sheet = workbook[workbook.sheetnames[0]]

        # In Excel files, the row indices start from 1.
        headers = [cell.value for cell in sheet[1]]
        for header in get_viemari_ilmoitustiedosto_headers():
            if header not in headers:
                missing_headers_list.append(header)

        if missing_headers_list:
            print(
                f"Tiedosto: {self._path}, puuttuvat sarakeotsikot: {missing_headers_list}"
            )
            raise RuntimeError("Viemäri-ilmoitustiedostosta puuttuu oletettuja sarakeotsikoita.")

        for row in sheet.iter_rows(min_row=2, values_only=True):
            try:
                data = dict(zip(headers, row))
                lopetus_obj = ViemariIlmoitus.parse_obj(data)
                lopetus_obj.rawdata = data
                lopetus_list.append(lopetus_obj)
            except ValidationError as e:
                logger.warning(
                    f"Viemäri-olion luonti epäonnistui datalla: {row}. Virhe: {e}"
                )
                failed_validations.append(data)

        export_kohdentumattomat_viemariilmoitukset(
            Path(self._path).parent, failed_validations, self._path
        )

        return lopetus_list

class ViemariLopetustiedosto:
    def __init__(self, path):
        self._path = path

    @classmethod
    def readable_by_me(cls, path):
        file = Path(path)
        if file.is_file() and file.suffix == ".xlsx":
            try:
                workbook = load_workbook(filename=file, data_only=True, read_only=True)
                if len(workbook.sheetnames):
                    return True
            except Exception:
                pass
        return False

    @property
    def lopetusilmoitukset(self)-> List[ViemariLopetusIlmoitus]:
        lopetus_list = []
        missing_headers_list = []
        failed_validations = []

        workbook = load_workbook(self._path)
        sheet = workbook[workbook.sheetnames[0]]

        # In Excel files, the row indices start from 1.
        headers = [cell.value for cell in sheet[1]]
        for header in get_viemari_lopetustiedosto_headers():
            if header not in headers:
                missing_headers_list.append(header)

        if missing_headers_list:
            print(
                f"Tiedosto: {self._path}, puuttuvat sarakeotsikot: {missing_headers_list}"
            )
            raise RuntimeError("Viemärin lopetusilmoitustiedostosta puuttuu oletettuja sarakeotsikoita.")

        for row in sheet.iter_rows(min_row=2, values_only=True):
            try:
                data = dict(zip(headers, row))
                lopetus_obj = ViemariLopetusIlmoitus.parse_obj(data)
                lopetus_obj.rawdata = data
                lopetus_list.append(lopetus_obj)
            except ValidationError as e:
                logger.warning(
                    f"Viemärin lopetus epäonnistui datalla: {row}. Virhe: {e}"
                )
                failed_validations.append(data)

        export_kohdentumattomat_viemarilopetusilmoitukset(
            Path(self._path).parent, failed_validations
        )

        return lopetus_list



def export_kohdentumattomat_viemariilmoitukset(
        folder: Path,
        kohdentumattomat: List[Dict[str, str]],
        fullpath: str
):
    print(folder)
    expected_headers = get_viemari_ilmoitustiedosto_headers()
    if '_' in str(fullpath):
        lahettaja = str(fullpath).split('_')[1].split('.')[0]
    else:
        lahettaja = str(fullpath).split('.',maxsplit=1)[0]

    print(lahettaja)

    output_file_path_failed = folder / get_kohdentumattomat_viemari_ilmoitus_filename(lahettaja)

    if output_file_path_failed.exists():
        workbook_failed = openpyxl.load_workbook(output_file_path_failed)
        sheet_failed = workbook_failed[workbook_failed.sheetnames[0]]
    else:
        workbook_failed = openpyxl.Workbook()
        sheet_failed = workbook_failed[workbook_failed.sheetnames[0]]
        sheet_failed.append(expected_headers)

    filtered_kohdentumattomat = [
        {key: value for key, value in data.items() if key in expected_headers}
        for data in kohdentumattomat
    ]
    for row in filtered_kohdentumattomat:
        sheet_failed.append([row.get(header, "") for header in expected_headers])

    workbook_failed.save(output_file_path_failed)

    file_content = output_file_path_failed.read_bytes()
    asyncio.run(sp.upload_file(file_content=file_content, filename=output_file_path_failed.name, user_name="jkr-core"))


def export_kohdentumattomat_viemarilopetusilmoitukset(
        folder: Path,
        kohdentumattomat: List[Dict[str, str]]
):
    expected_headers = get_viemari_lopetustiedosto_headers()

    output_file_path_failed = folder / get_kohdentumattomat_viemarin_lopetus_filename()
    

    if output_file_path_failed.exists():
        workbook_failed = openpyxl.load_workbook(output_file_path_failed)
        sheet_failed = workbook_failed[workbook_failed.sheetnames[0]]
    else:
        workbook_failed = openpyxl.Workbook()
        sheet_failed = workbook_failed[workbook_failed.sheetnames[0]]
        sheet_failed.append(expected_headers)

    filtered_kohdentumattomat = [
        {key: value for key, value in data.items() if key in expected_headers}
        for data in kohdentumattomat
    ]
    for row in filtered_kohdentumattomat:
        sheet_failed.append([row.get(header, "") for header in expected_headers])

    workbook_failed.save(output_file_path_failed)
    
    file_content = output_file_path_failed.read_bytes()
    asyncio.run(sp.upload_file(file_content=file_content, filename=output_file_path_failed.name, user_name="jkr-core"))