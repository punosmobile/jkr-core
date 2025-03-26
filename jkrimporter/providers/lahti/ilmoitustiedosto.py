import logging
import os
from pathlib import Path

from openpyxl.reader.excel import load_workbook
from pydantic import ValidationError

from jkrimporter.datasheets import (
    get_ilmoitustiedosto_headers,
    get_lopetustiedosto_headers
)
from jkrimporter.providers.lahti.models import Ilmoitus, LopetusIlmoitus
from jkrimporter.utils.ilmoitus import (
    export_kohdentumattomat_ilmoitukset,
    export_kohdentumattomat_lopetusilmoitukset,
)

logger = logging.getLogger(__name__)


class Ilmoitustiedosto:
    def __init__(self, path):
        self._path = path

    @classmethod
    def readable_by_me(cls, path):
        file = Path(path)
        if file.is_file() and file.suffix == ".xlsx":
            try:
                workbook = load_workbook(filename=file, data_only=True, read_only=True)
                if len(workbook.sheetnames):
                    return True
            except Exception:
                pass
        return False

    @property
    def ilmoitukset(self):
        ilmoitus_list = []
        missing_headers_list = []
        failed_validations = []

        workbook = load_workbook(self._path)
        sheet = workbook[workbook.sheetnames[0]]

        # In Excel files, the row indices start from 1.
        headers = [cell.value for cell in sheet[1]]
        for header in get_ilmoitustiedosto_headers():
            if header not in headers:
                missing_headers_list.append(header)

        if missing_headers_list:
            print(
                f"Tiedosto: {self._path}, puuttuvat sarakeotsikot: {missing_headers_list}"
            )
            raise RuntimeError("Ilmoitustiedostosta puuttuu oletettuja sarakeotsikoita.")

        for row in sheet.iter_rows(min_row=2, values_only=True):
            try:
                headerstrings = [str(header) for header in headers]
                data = dict(zip(headerstrings, row))
                ilmoitus_obj = Ilmoitus.parse_obj(data)
                ilmoitus_obj.rawdata = data
                ilmoitus_list.append(ilmoitus_obj)
            except ValidationError as e:
                logger.warning(
                    f"Ilmoitus-olion luonti epäonnistui datalla: {row}. Virhe: ", e
                )
                failed_validations.append(data)

        export_kohdentumattomat_ilmoitukset(
            os.path.dirname(self._path), failed_validations
        )

        return ilmoitus_list


class LopetusIlmoitustiedosto:
    def __init__(self, path):
        self._path = path

    @classmethod
    def readable_by_me(cls, path):
        file = Path(path)
        if file.is_file() and file.suffix == ".xlsx":
            try:
                workbook = load_workbook(filename=file, data_only=True, read_only=True)
                if len(workbook.sheetnames):
                    return True
            except Exception:
                pass
        return False

    @property
    def lopetusilmoitukset(self):
        lopetus_list = []
        missing_headers_list = []
        failed_validations = []

        workbook = load_workbook(self._path)
        sheet = workbook[workbook.sheetnames[0]]

        # In Excel files, the row indices start from 1.
        headers = [cell.value for cell in sheet[1]]
        for header in get_lopetustiedosto_headers():
            if header not in headers:
                missing_headers_list.append(header)

        if missing_headers_list:
            print(
                f"Tiedosto: {self._path}, puuttuvat sarakeotsikot: {missing_headers_list}"
            )
            raise RuntimeError("Lopetus lopetusilmoitustiedostosta puuttuu oletettuja sarakeotsikoita.")

        for row in sheet.iter_rows(min_row=2, values_only=True):
            try:
                data = dict(zip(headers, row))
                lopetus_obj = LopetusIlmoitus.parse_obj(data)
                lopetus_obj.rawdata = data
                lopetus_list.append(lopetus_obj)
            except ValidationError as e:
                logger.warning(
                    f"LopetusIlmoitus-olion luonti epäonnistui datalla: {row}. Virhe: {e}"
                )
                failed_validations.append(data)

        export_kohdentumattomat_lopetusilmoitukset(
            os.path.dirname(self._path), failed_validations
        )

        return lopetus_list
