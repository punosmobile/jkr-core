import warnings
from pathlib import Path

from openpyxl.reader.excel import load_workbook

from jkrimporter.datasheets import ExcelSheetCollection, SiirtotiedostoSheet
from jkrimporter.providers.nokia.models import Asiakas, Tyhjennystapahtuma


class AsiakastiedotSheet(SiirtotiedostoSheet[Asiakas]):
    @staticmethod
    def _obj_from_dict(data):
        return Asiakas.parse_obj(data)


class TyhjennystapahtumatSheet(SiirtotiedostoSheet[Tyhjennystapahtuma]):
    @staticmethod
    def _obj_from_dict(data):
        return Tyhjennystapahtuma.parse_obj(data)


class NokiaSiirtotiedosto:
    class SheetNames:
        ASIAKKAAT = "1. Asiakkaat"
        TYHJENNYKSET = "2. Kuljetukset"

    def __init__(self, path):
        self._sheet_collection = ExcelSheetCollection(path)

    @classmethod
    def readable_by_me(cls, path):
        p = Path(path)

        if p.is_file() and p.suffix == ".xlsx":
            workbook = load_workbook(filename=path, data_only=True, read_only=True)
            sheets = workbook.sheetnames
            if (
                NokiaSiirtotiedosto.SheetNames.ASIAKKAAT in sheets
                and NokiaSiirtotiedosto.SheetNames.TYHJENNYKSET in sheets
            ):
                return True
        else:
            return False

    @property
    def asiakastiedot(self):
        return AsiakastiedotSheet(
            self._sheet_collection, NokiaSiirtotiedosto.SheetNames.ASIAKKAAT
        )

    @property
    def kuljetustiedot(self):
        return TyhjennystapahtumatSheet(
            self._sheet_collection, NokiaSiirtotiedosto.SheetNames.TYHJENNYKSET
        )


if __name__ == "__main__":
    s = NokiaSiirtotiedosto("data/nokia/jkr_Nokia_lietekuskit_v009.xlsx")

    for a in s.asiakastiedot:
        print(a)

    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

        for k in s.kuljetustiedot:
            print(k)
