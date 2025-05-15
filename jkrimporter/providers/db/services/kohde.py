from dataclasses import dataclass
import datetime
import re
import logging
from collections import defaultdict
from datetime import date, datetime as dt
from datetime import timedelta
from functools import lru_cache
from typing import TypeVar, DefaultDict, Set, List, Dict,TYPE_CHECKING,NamedTuple, FrozenSet, Optional, Generic, Iterable, Callable, Union
from ..models import Kohde  # Lisätään puuttuvat importit

from openpyxl import load_workbook
from psycopg2.extras import DateRange
from sqlalchemy import and_, exists, or_, select, update, case, delete, text, func
from sqlalchemy.exc import NoResultFound
from sqlalchemy.orm.decl_api import DeclarativeMeta
from sqlalchemy.orm import Session
from pathlib import Path

from jkrimporter.model import Asiakas, JkrIlmoitukset, LopetusIlmoitus, Yhteystieto

from .. import codes
from ..codes import KohdeTyyppi, OsapuolenrooliTyyppi, RakennuksenKayttotarkoitusTyyppi, RakennuksenOlotilaTyyppi
from ..models import (
    Katu,
    Kohde,
    KohteenOsapuolet,
    KohteenRakennukset,
    Kompostori,
    KompostorinKohteet,
    Kuljetus,
    Osapuoli,
    Osoite,
    RakennuksenOmistajat,
    RakennuksenVanhimmat,
    Rakennus,
    Sopimus,
    UlkoinenAsiakastieto,
    Viranomaispaatokset,
    HapaAineisto,
)
from ..utils import clean_asoy_name, form_display_name, is_asoy, is_company, is_yhteiso
from .buildings import DISTANCE_LIMIT, create_nearby_buildings_lookup, maximum_distance_of_buildings

T = TypeVar('T')

if TYPE_CHECKING:
    from pathlib import Path
    from typing import (
        Callable,
        DefaultDict,
        Dict,
        FrozenSet,
        Hashable,
        List,
        NamedTuple,
        Optional,
        Set,
        Tuple,
        Union,
    )

    from sqlalchemy.orm import Session
    from sqlalchemy.sql.selectable import Select

    from jkrimporter.model import Asiakas, JkrIlmoitukset, Tunnus

    class Kohdetiedot(NamedTuple):
        kohde: Kohde
        rakennukset: FrozenSet[KohteenRakennukset]
        lisarakennukset: FrozenSet[KohteenRakennukset]
        asukkaat: FrozenSet[KohteenOsapuolet]
        omistajat: FrozenSet[KohteenOsapuolet]

@dataclass
class BuildingInfo:
    """Apurakennuksen tiedot"""
    id: int
    prt: str
    owner_ids: Set[int]
    address_ids: Set[tuple]  # (katu_id, osoitenumero)
    is_auxiliary: bool


class BatchProgressTracker(Generic[T]):
    """
    Seuraa eräajon edistymistä ja tulostaa säännöllisiä tilannepäivityksiä.
    """
    def __init__(self, total: int, operation_name: str, update_interval: int = 1):
        self.total = total
        self.current = 0
        self.start_time = dt.now()
        self.last_update = self.start_time
        self.operation_name = operation_name
        self.update_interval = update_interval
        self.logger = logging.getLogger(__name__)
        
        # Tulosta aloitusviesti
        self.logger.info(f"\nAloitetaan {operation_name}")
        self.logger.info(f"Käsitellään yhteensä {total:,} kohdetta")

    def update(self, increment: int = 1) -> None:
        """Päivitä laskuri ja tulosta edistyminen jos on kulunut tarpeeksi aikaa."""
        self.current += increment
        now = dt.now()
        
        if (now - self.last_update).seconds >= self.update_interval:
            self._print_progress(now)
            self.last_update = now

    def _print_progress(self, now: dt) -> None:
        elapsed = now - self.start_time
        if self.current > 0:
            items_per_second = self.current / elapsed.total_seconds()
            remaining_items = self.total - self.current
            eta = timedelta(seconds=int(remaining_items / items_per_second)) if items_per_second > 0 else "---"
        else:
            items_per_second = 0
            eta = "---"

        self.logger.info(
            f"{self.operation_name}: "
            f"({self.current:,}/{self.total:,}) "
            f"nopeudella {items_per_second:.1f}/s, "
            f"aikaa jäljellä {eta}"
        )

    def done(self) -> None:
        """Tulosta yhteenveto kun prosessi on valmis."""
        elapsed = dt.now() - self.start_time
        items_per_second = self.current / elapsed.total_seconds()
        
        self.logger.info(
            f"\n{self.operation_name} valmis!\n"
            f"Käsitelty {self.current:,} kohdetta, "
            f"aikaa kului {elapsed}, "
            f"keskimääräinen nopeus {items_per_second:.1f}/s"
        )


def process_with_progress(
    items: Iterable[T],
    operation_name: str,
    process_func: Callable[[T], None],
    total: Optional[int] = None
) -> None:
    """
    Kääre-funktio, joka prosessoi kokoelman ja näyttää edistymisen.
    
    Args:
        items: Käsiteltävä kokoelma
        operation_name: Operaation nimi logeissa
        process_func: Funktio joka käsittelee yhden kohteen
        total: Kohteiden kokonaismäärä (jos ei ole len(items))
    """
    if total is None:
        try:
            total = len(items)
        except TypeError:
            total = 0  # Jos kokoelman kokoa ei voi määrittää
            
    progress = BatchProgressTracker(total, operation_name)
    
    try:
        for item in items:
            process_func(item)
            progress.update()
    finally:
        progress.done()

class Rakennustiedot(NamedTuple):
    rakennus: 'Rakennus'
    vanhimmat: FrozenSet['RakennuksenVanhimmat']
    omistajat: FrozenSet['RakennuksenOmistajat'] 
    osoitteet: FrozenSet['Osoite']


separators = [r"\s\&\s", r"\sja\s"]
separator_regex = r"|".join(separators)


def match_name(first: str, second: str) -> bool:
    """
    Returns true if one name is subset of the other, i.e. the same name without extra
    parts. Consider typical separators combining names.
    """
    first = re.split(separator_regex, first)
    second = re.split(separator_regex, second)
    if len(first) > 1 or len(second) > 1:
        # TODO: if names have common surname, paste it to the name that is missing
        # surname. Will take some detective work tho.
        return any(
            match_name(first_name, second_name)
            for first_name in first
            for second_name in second
        )
    first_parts = set(first[0].lower().split())
    second_parts = set(second[0].lower().split())
    return first_parts.issubset(second_parts) or second_parts.issubset(first_parts)


def get_ulkoinen_asiakastieto(
    session: "Session", ulkoinen_tunnus: "Tunnus"
) -> "Union[UlkoinenAsiakastieto, None]":
    query = select(UlkoinenAsiakastieto).where(
        UlkoinenAsiakastieto.tiedontuottaja_tunnus == ulkoinen_tunnus.jarjestelma,
        UlkoinenAsiakastieto.ulkoinen_id == ulkoinen_tunnus.tunnus,
    )
    try:
        return session.execute(query).scalar_one()
    except NoResultFound:
        return None


def update_ulkoinen_asiakastieto(ulkoinen_asiakastieto, asiakas: "Asiakas"):
    if ulkoinen_asiakastieto.ulkoinen_asiakastieto != asiakas.ulkoinen_asiakastieto:
        ulkoinen_asiakastieto.ulkoinen_asiakastieto = asiakas.ulkoinen_asiakastieto


def add_ulkoinen_asiakastieto_for_kohde(
    session: Session, kohde: Kohde, asiakas: Asiakas
):
    """
    Lisää ulkoisen asiakastiedon kohteelle.

    Args:
        session: Tietokantaistunto
        kohde: Kohde jolle asiakastieto lisätään
        asiakas: Asiakas jonka tiedot lisätään

    Returns:
        UlkoinenAsiakastieto: Luotu asiakastieto-objekti
    """
    asiakastieto = UlkoinenAsiakastieto(
        tiedontuottaja_tunnus=asiakas.asiakasnumero.jarjestelma,
        ulkoinen_id=asiakas.asiakasnumero.tunnus,
        ulkoinen_asiakastieto=asiakas.ulkoinen_asiakastieto,
        kohde=kohde,
    )

    session.add(asiakastieto)
    return asiakastieto


def find_kohde_by_prt(
    session: "Session", 
    asiakas: "Union[Asiakas, JkrIlmoitukset, LopetusIlmoitus]"
) -> "Union[Kohde, None]":
    """
    Finds a kohde based on PRT identifiers from different source types.
    
    Args:
        session: Database session
        asiakas: The source object, which can be:
            - Asiakas (customer data)  
            - JkrIlmoitukset (composting notifications)
            - LopetusIlmoitus (composting termination notices)
            
    Returns:
        Kohde object if found, None otherwise
    """
    if isinstance(asiakas, JkrIlmoitukset):
        return _find_kohde_by_asiakastiedot(
            session, Rakennus.prt.in_(asiakas.sijainti_prt), asiakas
        )
    elif isinstance(asiakas, LopetusIlmoitus):
        # Lopetusilmoitukselle käytetään erikoiskäsittelyä
        return _find_kohde_by_ilmoitustiedot(
            session,
            Rakennus.prt.in_(asiakas.prt),
            asiakas
        )
    elif isinstance(asiakas, Asiakas):
        return _find_kohde_by_asiakastiedot(
            session, Rakennus.prt.in_(asiakas.rakennukset), asiakas
        )
    else:
        raise ValueError(f"Invalid asiakas type: {type(asiakas)}")


def find_kohteet_by_prt(
        session: "Session",
        asiakas: "JkrIlmoitukset"
) -> "Tuple[List[Kohde], List[str]]":
    found_kohteet = []
    not_found_prts = []

    for kompostoija in asiakas.kompostoijat:
        kompostoija_info = f"Nimi: {kompostoija.nimi}, Rakennus: {kompostoija.rakennus}"
        print(f"Kompostoija: {kompostoija_info}")
        query = (
            select(Kohde.id, Osapuoli.nimi)
            .join(KohteenRakennukset, KohteenRakennukset.kohde_id == Kohde.id)
            .join(Rakennus, Rakennus.id == KohteenRakennukset.rakennus_id)
            .join(KohteenOsapuolet, KohteenOsapuolet.kohde_id == Kohde.id, isouter=True)
            .join(Osapuoli, Osapuoli.id == KohteenOsapuolet.osapuoli_id, isouter=True)
            .join(Osoite, Osoite.rakennus_id == Rakennus.id, isouter=True)
            .where(
                Kohde.voimassaolo.overlaps(
                    DateRange(
                        asiakas.voimassa.lower or datetime.date.min,
                        asiakas.voimassa.upper or datetime.date.max,
                    )
                ),
                Rakennus.prt.in_(kompostoija.rakennus)
            )
            .distinct()
        )

        try:
            kohteet = session.execute(query).all()
        except NoResultFound:
            print(f"Ei löytynyt kohdetta virheen takia, prt: {kompostoija.rakennus}")
            not_found_prts.append(kompostoija.rakennus)
            continue

        kompostoija_nimi = clean_asoy_name(kompostoija.nimi)
        if len(kohteet) > 1:
            names_by_kohde_id = defaultdict(set)
            for kohde_id, db_osapuoli_name in kohteet:
                names_by_kohde_id[kohde_id].add(db_osapuoli_name)
            for kohde_id, db_osapuoli_names in names_by_kohde_id.items():
                for db_osapuoli_name in db_osapuoli_names:
                    if db_osapuoli_name is not None:
                        db_osapuoli_name = clean_asoy_name(db_osapuoli_name)
                        print(kompostoija_nimi)
                        print(db_osapuoli_name)
                        if (
                            match_name(kompostoija_nimi, db_osapuoli_name)
                        ):
                            print(f"{db_osapuoli_name} match")
                            kohde = session.get(Kohde, kohde_id)
                            print("Adding kohde to list")
                            found_kohteet.append(kohde)
        elif len(kohteet) == 1:
            kohde_id = kohteet[0][0]
            kohde = session.get(Kohde, kohde_id)
            found_kohteet.append(kohde)
        else:
            print(f"Ei löytynyt kohdetta, prt: {kompostoija.rakennus}")
            not_found_prts.append(kompostoija.rakennus)

    return found_kohteet, not_found_prts


def find_kohde_by_address(
    session: "Session", asiakas: "Asiakas"
) -> "Union[Kohde, None]":
    print(f"matching by address at {datetime.datetime.now()}:")
    # The osoitenumero may contain dash. In that case, the buildings may be
    # listed as separate in DVV data.
    if (
        asiakas.haltija.osoite.osoitenumero
        and "-" in asiakas.haltija.osoite.osoitenumero
    ):
        osoitenumerot = asiakas.haltija.osoite.osoitenumero.split("-", maxsplit=1)
    else:
        osoitenumerot = [asiakas.haltija.osoite.osoitenumero]

    print(osoitenumerot)
    # The address parser parses Metsätie 33 A so that 33 is osoitenumero and A is
    # huoneistotunnus. While the parsing is correct, it may very well also mean (and
    # in many cases it means) osoitenumero 33a and empty huoneistonumero.
    potential_osoitenumero_suffix = (
        asiakas.haltija.osoite.huoneistotunnus.lower()
        if asiakas.haltija.osoite.huoneistotunnus
        else ""
    )
    if potential_osoitenumero_suffix:
        osoitenumerot_with_suffix = [
            osoitenumero + potential_osoitenumero_suffix
            for osoitenumero in osoitenumerot
        ]
    else:
        osoitenumerot_with_suffix = []
    print(osoitenumerot_with_suffix)

    # - Do *NOT* find Sokeritopankatu 18 *AND* Sokeritopankatu 18a by
    # Sokeritopankatu 18 A. Looks like Sokeritopankatu 18, 18 A and 18 B are *all*
    # separate.
    # - Do *NOT* find Mukkulankatu 51 *AND* Mukkulankatu 51b by Mukkulankatu 51.
    # - Find Mukkulankatu 51 by Mukkulankatu 51 B.
    # - Only find Mukkulankatu 51 by Mukkulankatu 51.
    # - Only find Mukkulankatu 51b by Mukkulankatu 51b.
    if osoitenumerot_with_suffix:
        osoitenumero_filter = or_(
            # first, check if the letter is actually part of osoitenumero
            Osoite.osoitenumero.in_(osoitenumerot_with_suffix),
            # if the letter is not found in osoitenumero, it is huoneistotunnus
            and_(
                ~exists(Osoite.osoitenumero.in_(osoitenumerot_with_suffix)),
                Osoite.osoitenumero.in_(osoitenumerot),
            ),
        )
    else:
        osoitenumero_filter = Osoite.osoitenumero.in_(osoitenumerot)

    # TODO Using this filter WILL cause a seemingly infinite during kuljetustiedot import loop. Function should be fixed at a later time 15.5.2025 EK
    # If katunimi is also present, try to match it and return the result. Otherwise do not try to find it
    # if asiakas.haltija.osoite.katunimi:
    #     kohde_filter = and_(
    #         Osoite.posti_numero == asiakas.haltija.osoite.postinumero,
    #         or_(
    #             Katu.katunimi_fi == asiakas.haltija.osoite.katunimi,
    #             Katu.katunimi_sv == asiakas.haltija.osoite.katunimi,
    #         ),
    #         osoitenumero_filter,
    #     )

    #     return _find_kohde_by_asiakastiedot(session, kohde_filter, asiakas)

    return None

def _find_kohde_by_ilmoitustiedot(
    session: "Session",
    filter,
    ilmoitus: "LopetusIlmoitus" 
) -> "Optional[Kohde]":
    """
    Optimoitu versio lopetusilmoitusten käsittelyyn.
    """
    # 1. Hae kohteiden ID:t ensin
    kohde_ids_query = (
        select(Kohde.id)
        .join(KohteenRakennukset)
        .join(Rakennus)
        .where(
            filter,
            Kohde.voimassaolo.overlaps(DateRange(ilmoitus.Vastausaika))
        )
        .distinct()
    )

    try:
        kohde_ids = session.execute(kohde_ids_query).scalars().all()
    except Exception as e:
        print(f"Virhe kohde-ID:iden haussa: {e}")
        return None

    if not kohde_ids:
        return None

    # 2. Standardoi vastuuhenkilön nimi kerran
    vastuuhenkilo = ilmoitus.nimi.upper()

    # 3. Käy läpi kohteet kunnes löytyy täsmäys
    for kohde_id in kohde_ids:
        osapuolet_query = (
            select(Osapuoli.nimi)
            .join(KohteenOsapuolet)
            .where(
                KohteenOsapuolet.kohde_id == kohde_id,
                KohteenOsapuolet.osapuolenrooli_id.in_([1, 2, 311])
            )
        )
        
        osapuoli_nimet = session.execute(osapuolet_query).scalars().all()
        
        for db_nimi in osapuoli_nimet:
            if not db_nimi:
                continue
                
            if db_nimi.upper() == vastuuhenkilo:
                return session.get(Kohde, kohde_id)

    return None



def _find_kohde_by_asiakastiedot(
    session: "Session", 
    filter,  
    asiakas: "Union[Asiakas, JkrIlmoitukset]"
) -> "Optional[Kohde]":
    """
    Optimoitu versio kohteen haulle. Minimoi tietokantakyselyt ja muistin käytön.
    Säilyttää luotettavan kohdennuksen.
    """
    # 1. Hae ensin pelkät kohteiden ID:t rakennusten perusteella
    kohde_ids_query = (
        select(Kohde.id)
        .join(KohteenRakennukset)
        .join(Rakennus)
        .where(
            filter,
            Kohde.voimassaolo.overlaps(
                DateRange(
                    asiakas.voimassa.lower or datetime.date.min,
                    asiakas.voimassa.upper or datetime.date.max,
                )
            )
        )
        .distinct()
    )

    try:
        kohde_ids = session.execute(kohde_ids_query).scalars().all()
    except Exception as e:
        print(f"Virhe kohde-ID:iden haussa: {e}")
        return None

    if not kohde_ids:
        return None

    # 2. Standardoi asiakkaan nimi kerran
    if isinstance(asiakas, JkrIlmoitukset):
        asiakas_nimi = asiakas.vastuuhenkilo.nimi.upper()
    else:
        asiakas_nimi = asiakas.haltija.nimi.upper()

    # 3. Hae ja validoi kohteet yksi kerrallaan kunnes löytyy täsmäys
    for kohde_id in kohde_ids:
        # Hae vain oleelliset osapuolet kohteelle
        osapuolet_query = (
            select(Osapuoli.nimi)
            .join(KohteenOsapuolet)
            .where(
                KohteenOsapuolet.kohde_id == kohde_id,
                KohteenOsapuolet.osapuolenrooli_id.in_([1, 2, 311])  # Vain oleelliset roolit
            )
        )
        
        osapuoli_nimet = session.execute(osapuolet_query).scalars().all()
        
        # Tarkista täsmääkö jokin osapuoli
        for db_nimi in osapuoli_nimet:
            if not db_nimi:
                continue
                
            db_nimi = db_nimi.upper()

            # Jos kyseessä asoy, vaadi tarkka täsmäys
            if 'ASOY' in db_nimi or 'AS OY' in db_nimi:
                if db_nimi == asiakas_nimi:
                    return session.get(Kohde, kohde_id)
            # Muuten salli osittainen täsmäys
            elif (asiakas_nimi in db_nimi or db_nimi in asiakas_nimi):
                return session.get(Kohde, kohde_id)

    return None



def update_kohde(kohde: Kohde, asiakas: "Asiakas"):
    if kohde.alkupvm != asiakas.voimassa.lower:
        kohde.alkupvm = asiakas.voimassa.lower
    if kohde.loppupvm != asiakas.voimassa.upper:
        kohde.loppupvm = asiakas.voimassa.upper


def get_kohde_by_asiakasnumero(
    session: "Session", tunnus: "Tunnus"
) -> "Union[Kohde, None]":
    query = (
        select(Kohde)
        .join(UlkoinenAsiakastieto)
        .where(
            UlkoinenAsiakastieto.tiedontuottaja_tunnus == tunnus.jarjestelma,
            UlkoinenAsiakastieto.ulkoinen_id == tunnus.tunnus,
        )
    )
    try:
        kohde = session.execute(query).scalar_one()
    except NoResultFound:
        kohde = None

    return kohde


@lru_cache(maxsize=100)
def get_or_create_pseudokohde(session: Session, nimi: str, kohdetyyppi) -> Kohde:
    kohdetyyppi = codes.kohdetyypit[kohdetyyppi]
    query = select(Kohde).where(Kohde.nimi == nimi, Kohde.kohdetyyppi == kohdetyyppi)
    try:
        kohde = session.execute(query).scalar_one()
    except NoResultFound:
        kohde = Kohde(nimi=nimi, kohdetyyppi=kohdetyyppi)
        session.add(kohde)

    return kohde


def get_dvv_rakennustiedot_without_kohde(
   session: Session,
   poimintapvm: Optional[date],
   loppupvm: Optional[date]
) -> Dict[int, "Rakennustiedot"]:
    """
    Hakee DVV:n rakennustiedot rakennuksille, joilla ei ole voimassaolevaa kohdetta.
    
    Rakennustiedot haetaan vain rakennuksille jotka:
    1. Eivät kuulu millekään voimassaolevalle kohteelle annetulla aikavälillä
    2. Eivät ole poistettu käytöstä (kaytostapoisto_pvm) ennen poimintapäivää

    Args:
        session: Tietokantaistunto
        poimintapvm: Uusien kohteiden alkupäivämäärä
        loppupvm: Uusien kohteiden loppupäivämäärä. Jos None, ei aikarajausta.

    Returns:
        Dict[int, Rakennustiedot]: Sanakirja jossa avaimena rakennuksen id ja
        arvona tuple (rakennus, vanhimmat set, omistajat set, osoitteet set)
    """
    # 1. Hae ensin rakennukset joilla ON voimassaoleva kohde
    if loppupvm is None:
        rakennus_id_with_current_kohde = (
            select(Rakennus.id)
            .join(KohteenRakennukset)
            .join(Kohde)
            .filter(
                or_(
                    Kohde.loppupvm.is_(None),
                    poimintapvm < Kohde.loppupvm
                )
            )
        )
    else:
        rakennus_id_with_current_kohde = (
            select(Rakennus.id)
            .join(KohteenRakennukset)
            .join(Kohde)
            .filter(
                Kohde.voimassaolo.overlaps(DateRange(poimintapvm, loppupvm))
            )
        )

    # 2. Hae kaikki rakennukset jotka:
    # - Eivät ole yllä haetussa kohdelistassa
    # - Eivät ole poistuneet käytöstä
    query = (
        select(Rakennus, RakennuksenVanhimmat, RakennuksenOmistajat, Osoite)
        .select_from(Rakennus)
        .outerjoin(RakennuksenVanhimmat)
        .outerjoin(RakennuksenOmistajat)
        .outerjoin(Osoite)
        .filter(
            ~Rakennus.id.in_(rakennus_id_with_current_kohde)
        )
        .filter(
            or_(
                Rakennus.kaytostapoisto_pvm.is_(None),
                Rakennus.kaytostapoisto_pvm > poimintapvm
            )
        )
    )

    rows = session.execute(query).all()

    # 3. Ryhmittele rakennustiedot rakennuksen id:n mukaan
    rakennustiedot_by_id = {}
    for row in rows:
        rakennus_id = row[0].id
        if rakennus_id not in rakennustiedot_by_id:
            # Luo uusi tuple (rakennus, vanhimmat set, omistajat set, osoitteet set)
            rakennustiedot_by_id[rakennus_id] = (row[0], set(), set(), set())
            
        # Lisää vanhimmat, omistajat ja osoitteet setteihin jos niitä on
        if row[1]:  # Vanhimmat
            rakennustiedot_by_id[rakennus_id][1].add(row[1])
        if row[2]:  # Omistajat
            rakennustiedot_by_id[rakennus_id][2].add(row[2])
        if row[3]:  # Osoitteet 
            rakennustiedot_by_id[rakennus_id][3].add(row[3])

    # 4. Muunna setit frozenset:eiksi että niitä voi käyttää esim. avaimina
    return {
        id: (
            rakennus,
            frozenset(vanhimmat),
            frozenset(omistajat),
            frozenset(osoitteet)
        )
        for id, (rakennus, vanhimmat, omistajat, osoitteet) 
        in rakennustiedot_by_id.items()
    }


def _get_identifiers(osapuolet: Union[Set[Osapuoli], FrozenSet[RakennuksenOmistajat], FrozenSet[RakennuksenVanhimmat]]) -> Set[str]:
    """Kerää osapuolten tunnisteet (id:t, y-tunnukset ja henkilötunnukset)"""
    identifiers = set()
    for osapuoli in osapuolet:
        # Käsittele RakennuksenOmistajat ja RakennuksenVanhimmat
        if hasattr(osapuoli, 'osapuoli'):
            osapuoli = osapuoli.osapuoli

        # Kerää kaikki mahdolliset tunnisteet
        if hasattr(osapuoli, 'osapuoli_id'):
            identifiers.add(str(osapuoli.osapuoli_id))
        if hasattr(osapuoli, 'id'):
            identifiers.add(str(osapuoli.id))
        if hasattr(osapuoli, 'ytunnus') and osapuoli.ytunnus:
            identifiers.add(osapuoli.ytunnus)
        if hasattr(osapuoli, 'henkilotunnus') and osapuoli.henkilotunnus:
            identifiers.add(osapuoli.henkilotunnus)
    return identifiers

def _match_ownership_or_residents(
    cluster_buildings: Set["Rakennustiedot"],
    building2: "Rakennustiedot" 
) -> bool:
    """
    Tarkistaa, onko rakennusryhmällä ja rakennuksella sama omistaja JA asukas.
    Vertailee sekä ID:n että tunnisteen (y-tunnus/henkilötunnus) perusteella.
    """

    yhteensopivat: Set[bool] = set()
   
    for building in cluster_buildings:
         # Kerää omistajien tunnisteet
        owners1 = _get_identifiers(building[2])
        owners2 = _get_identifiers(building2[2])

        # Kerää asukkaiden tunnisteet
        residents1 = _get_identifiers(building[1])
        residents2 = _get_identifiers(building2[1])

        print(f"asukkaat 1: {residents1} {len(residents1)}, 2: {residents2} {len(residents2)}")
        if len(residents1) > 0 and len(residents2) > 0:
            print(f"yhteiset asukkaat: {residents1 and residents2 and (residents1 & residents2)}")
            print(f"yhteiset omistajat: {owners1 and owners2 and (owners1 & owners2)}")
            yhteensopivat.add(bool(
                (owners1 and owners2 and (owners1 & owners2)) and
                (residents1 and residents2 and (residents1 & residents2))
            ))

        # Palauta True jos joko omistajissa JA asukkaissa on yhteinen tunniste
        # if bool(
        #     (owners1 and owners2 and (owners1 & owners2)) or
        #     (residents1 and residents2 and (residents1 & residents2))
        # ):
        #     print(f"{building1[0].prt} {building2[0].prt} Match by: {owners1} {owners2} {residents1} {residents2}")
        # else:
        #     print(f"{building1[0].prt} {building2[0].prt} No match by: {owners1} {owners2} {residents1} {residents2}")
        print(f"omistajavertaus 1. {owners1} 2. {owners2}")
        print(f"omistajavertaus osumat: {(owners1 and owners2 and (owners1 & owners2))}")
        yhteensopivat.add(bool(
            (owners1 and owners2 and (owners1 & owners2))
        ))

    if False in yhteensopivat:
        return False

    return True


def _normalize_address(address: "Osoite") -> str:
    """
    Normalisoi osoitteen vertailua varten.
    Palauttaa osoitteen muodossa 'kadunnimi numero' ilman kirjaimia tai muita lisäosia.
    Esim. 'Leninpolku 1 a 5' -> 'leninpolku 1'

    Args:
        address: Osoite-objekti

    Returns:
        str: Normalisoitu osoite
    """
    # Tarkista että osoitteella on katu ja numero
    if not address.katu or not address.osoitenumero:
        return ""

    # Poista välilyönnit alusta ja lopusta, muuta pieniksi kirjaimiksi
    katunimi = address.katu.katunimi_fi.strip().lower() if address.katu.katunimi_fi else ""
    
    # Ota numerosta vain ensimmäinen numero (ennen kirjainta tai välilyöntiä)
    numero_match = re.match(r'^\d+', address.osoitenumero.strip())
    numero = numero_match.group(0) if numero_match else ""

    return f"{katunimi} {numero}"


def _match_addresses(addresses1: "FrozenSet[Osoite]", addresses2: "FrozenSet[Osoite]") -> bool:
    """
    Tarkistaa onko osoitteilla yhtäläisyyksiä.
    Vertailee osoitteita vain kadunnimen ja numeron perusteella, jättäen huomiotta
    kirjaimet ja muut lisäosat.

    Esimerkiksi seuraavat osoitteet tulkitaan samoiksi:
    - Leninpolku 1
    - Leninpolku 1 a
    - Leninpolku 1a
    - Leninpolku 1b
    - Leninpolku 1 c 5

    Args:
        addresses1: Ensimmäisen rakennuksen osoitteet
        addresses2: Toisen rakennuksen osoitteet

    Returns:
        bool: True jos osoitteilla on yhtäläisyyksiä, muuten False
    """
    # Jos jommalla kummalla ei ole osoitteita, ei voida verrata
    if not addresses1 or not addresses2:
        return False

    # Normalisoi osoitteet vertailua varten
    normalized1 = {_normalize_address(addr) for addr in addresses1}
    normalized2 = {_normalize_address(addr) for addr in addresses2}

    # Poista tyhjät osoitteet
    normalized1 = {addr for addr in normalized1 if addr}
    normalized2 = {addr for addr in normalized2 if addr}

    # Jos jommalla kummalla ei ole valideja osoitteita, ei voida verrata
    if not normalized1 or not normalized2:
        return False

    # Tarkista onko yhteisiä normalisoituja osoitteita
    return bool(normalized1.intersection(normalized2))


def update_old_kohde_data(
    session: Session, 
    old_kohde_id: int, 
    new_kohde_id: int, 
    new_kohde_alkupvm: datetime.date
) -> None:
    """
    Päivittää vanhan kohteen tiedot ja siirtää tarvittavat tiedot uudelle kohteelle.
    
    Prosessi suoritetaan seuraavasti:
    1. Päivitetään vanhan kohteen loppupvm (uuden alkupvm - 1 päivä)
    2. Siirretään sopimukset ja kuljetukset joiden loppupvm >= uuden kohteen alkupvm
    3. Käsitellään viranomaispäätökset:
       - Päätöksille joiden alkupvm <= vanhan kohteen loppupvm asetetaan loppupvm
       - Muut päätökset siirretään uudelle kohteelle
    4. Käsitellään kompostorit:
       - Kompostoreille joiden alkupvm <= vanhan kohteen loppupvm asetetaan loppupvm
       - Muiden kompostorien osalta:
         * Siirretään osapuolet uudelle kohteelle
         * Päivitetään kompostorin kohdeviittaus

    Args:
        session: Tietokantaistunto
        old_kohde_id: Vanhan kohteen ID
        new_kohde_id: Uuden kohteen ID
        new_kohde_alkupvm: Uuden kohteen alkupäivämäärä

    Raises:
        SQLAlchemyError: Jos tietokantaoperaatioissa tapahtuu virhe
    """
    print(f"Päivitetään vanhan kohteen {old_kohde_id} tiedot uudelle kohteelle {new_kohde_id}")

    try:
        # Määritä vanhan kohteen loppupvm
        loppupvm = new_kohde_alkupvm - timedelta(days=1)
        
        with session.begin_nested():
            # 1. Päivitä kohteen loppupvm
            stmt = (
                update(Kohde)
                .where(Kohde.id == old_kohde_id)
                .values(loppupvm=loppupvm)
                .execution_options(synchronize_session=False)
            )
            session.execute(stmt)
            print(f"Kohteen {old_kohde_id} loppupvm påivitetty: {loppupvm}")

            # 2. Siirrä sopimukset ja kuljetukset
            for model in [Sopimus, Kuljetus]:
                stmt = (
                    update(model)
                    .where(
                        model.kohde_id == old_kohde_id,
                        model.loppupvm >= new_kohde_alkupvm
                    )
                    .values(kohde_id=new_kohde_id)
                    .execution_options(synchronize_session=False)
                )
                result = session.execute(stmt)
                print(f"Siirretty {result.rowcount} {model.__name__.lower()}ta uudelle kohteelle")

            # 3. Käsittele viranomaispäätökset
            # Hae vanhan kohteen rakennusten id:t
            rakennus_ids = select(KohteenRakennukset.rakennus_id).where(
                KohteenRakennukset.kohde_id == old_kohde_id
            )

            # 3.1 Aseta loppupvm vanhoille päätöksille
            stmt = (
                update(Viranomaispaatokset)
                .where(
                    Viranomaispaatokset.rakennus_id.in_(rakennus_ids.scalar_subquery()),
                    Viranomaispaatokset.alkupvm <= loppupvm
                )
                .values(loppupvm=loppupvm)
                .execution_options(synchronize_session=False)
            )
            result = session.execute(stmt)
            print(f"Päivitetty {result.rowcount} viranomaispäätöksen loppupvm")

            # 3.2 Siirrä voimassa olevat päätökset uudelle kohteelle
            stmt = (
                update(Viranomaispaatokset)
                .where(
                    Viranomaispaatokset.rakennus_id.in_(rakennus_ids.scalar_subquery()),
                    Viranomaispaatokset.alkupvm > loppupvm
                )
                .values(loppupvm=loppupvm)
                .execution_options(synchronize_session=False)
            )
            result = session.execute(stmt)
            print(f"Siirretty {result.rowcount} viranomaispäätöstä uudelle kohteelle")

            # 4. Käsittele kompostorit
            # Hae vanhan kohteen kompostorien id:t
            kompostori_ids = select(KompostorinKohteet.kompostori_id).where(
                KompostorinKohteet.kohde_id == old_kohde_id
            )

            # 4.1 Aseta loppupvm vanhoille kompostoreille
            stmt = (
                update(Kompostori)
                .where(
                    Kompostori.id.in_(kompostori_ids.scalar_subquery()),
                    Kompostori.alkupvm <= loppupvm
                )
                .values(loppupvm=loppupvm)
                .execution_options(synchronize_session=False)
            )
            result = session.execute(stmt)
            print(f"Päivitetty {result.rowcount} kompostorin loppupvm")

            # 4.2 Hae jatkuvien kompostorien id:t
            jatkuvat_kompostorit = select(Kompostori.id).where(
                and_(
                    Kompostori.id.in_(kompostori_ids.scalar_subquery()),
                    Kompostori.alkupvm > loppupvm
                )
            )

            # 4.3 Siirrä kompostorien osapuolet
            stmt = (
                update(KohteenOsapuolet)
                .where(
                    and_(
                        KohteenOsapuolet.osapuoli_id.in_(
                            select(Kompostori.osapuoli_id).where(
                                Kompostori.id.in_(jatkuvat_kompostorit.scalar_subquery())
                            )
                        ),
                        KohteenOsapuolet.kohde_id == old_kohde_id,
                        KohteenOsapuolet.osapuolenrooli_id == 311  # Kompostin yhteyshenkilö
                    )
                )
                .values(kohde_id=new_kohde_id)
                .execution_options(synchronize_session=False)
            )
            result = session.execute(stmt)
            print(f"Siirretty {result.rowcount} kompostorin osapuolta")

            # 4.4 Päivitä kompostorien kohdeviittaukset
            stmt = (
                update(KompostorinKohteet)
                .where(
                    and_(
                        KompostorinKohteet.kompostori_id.in_(jatkuvat_kompostorit.scalar_subquery()),
                        KompostorinKohteet.kohde_id == old_kohde_id
                    )
                )
                .values(kohde_id=new_kohde_id)
                .execution_options(synchronize_session=False)
            )
            result = session.execute(stmt)
            print(f"Päivitetty {result.rowcount} kompostorin kohdeviittausta")

        # Commit ulompi transaktio
        session.commit()
        print(f"Kohteen {old_kohde_id} tiedot påivitetty onnistuneesti")

    except SQLAlchemyError as e:
        session.rollback()
        print(
            f"Virhe kohteen {old_kohde_id} tietojen påivityksessä: {str(e)}"
        )
        raise


def determine_kohdetyyppi(session: "Session", rakennus: "Rakennus", asukkaat: "Optional[Set[Osapuoli]]" = None) -> KohdeTyyppi:
    """
    Määrittää kohteen tyypin rakennuksen tietojen perusteella.
    Logiikka:
    1. Tarkista onko asuinkiinteistö seuraavassa järjestyksessä:
       1. Rakennusluokka 2018 (0110-0211)
       2. Vanha rakennusluokka (011-041)
       3. Käyttötarkoitus (011-041)
       4. Huoneistomaara > 0
       5. Rakennuksenolotila (VAKINAINEN_ASUMINEN)
       6. Vähintään yksi asukas
    2. Jos mikään ehto ei täyty -> MUU

    Args:
        session: Tietokantaistunto
        rakennus: Rakennus-objekti
        asukkaat: Lista asukkaista (optional)

    Huom: HAPA/BIOHAPA tyypit määritellään myös tietokannassa triggerillä
    (katso V2.48.0__Add_hapa_trigger_and_functions.sql)
    """

    # 1. Tarkista rakennusluokka 2018
    if hasattr(rakennus, 'rakennusluokka_2018'):
        if rakennus.rakennusluokka_2018 is not None:
            try:
                luokka = int(rakennus.rakennusluokka_2018)
                if 110 <= luokka <= 211:
                    print(f"-> ASUINKIINTEISTO (rakennusluokka_2018: {luokka})")
                    return KohdeTyyppi.ASUINKIINTEISTO
            except (ValueError, TypeError):
                print(f"- rakennusluokka_2018 ei ole validi numero: {rakennus.rakennusluokka_2018}")
                pass
        else:
            print("- rakennusluokka_2018 ei ole annettu")

    # 2. Jos ei rakennusluokkaa 2018, tarkista käyttötarkoitus
    try:
        if hasattr(rakennus, 'rakennuksenkayttotarkoitus'):
            if rakennus.rakennuksenkayttotarkoitus is not None:
                kayttotarkoitus = int(rakennus.rakennuksenkayttotarkoitus.koodi if rakennus.rakennuksenkayttotarkoitus else None)       
                if 11 <= kayttotarkoitus <= 41:
                    print(f"-> ASUINKIINTEISTO (käyttötarkoitus): {kayttotarkoitus} {rakennus.rakennuksenkayttotarkoitus.koodi}")
                    return KohdeTyyppi.ASUINKIINTEISTO
        else:
            print("- rakennuksenkayttotarkoitus ei ole annettu")
    except (ValueError, TypeError):
        print("- kayttotarkoitus ei ole validi numero")
        pass

    # 3. Tarkista huoneistomäärä
    if hasattr(rakennus, 'huoneistomaara'):
        if rakennus.huoneistomaara is not None and rakennus.huoneistomaara > 0:
            print(f"-> ASUINKIINTEISTO (huoneistomaara: {rakennus.huoneistomaara})")
            return KohdeTyyppi.ASUINKIINTEISTO
    else:
        print("- huoneistomaara ei ole annettu")

    # 4. Tarkista rakennuksenolotila
    if hasattr(rakennus, 'rakennuksenolotila') or hasattr(rakennus, 'rakennuksenolotila_koodi'):
        if hasattr(rakennus, 'rakennuksenolotila') and rakennus.rakennuksenolotila is not None and rakennus.rakennuksenolotila.koodi in [
            RakennuksenOlotilaTyyppi.VAKINAINEN_ASUMINEN.value
        ] or hasattr(rakennus, 'rakennuksenolotila_koodi') and rakennus.rakennuksenolotila_koodi is not None and rakennus.rakennuksenolotila_koodi in [
            RakennuksenOlotilaTyyppi.VAKINAINEN_ASUMINEN.value
        ]:
            if hasattr(rakennus, 'rakennuksenolotila'):
                print(f"-> ASUINKIINTEISTO (rakennuksenolotila: {rakennus.rakennuksenolotila.koodi})")
            elif hasattr(rakennus, 'rakennuksenolotila_koodi'):
                print(f"-> ASUINKIINTEISTO (rakennuksenolotila: {rakennus.rakennuksenolotila_koodi})")
            return KohdeTyyppi.ASUINKIINTEISTO
    else:
        print("- rakennuksenolotila ei ole annettu")

    # 5. Tarkista asukkaat
    if asukkaat and len(asukkaat) > 0:
        print(f"-> ASUINKIINTEISTO (asukkaat) {len(asukkaat)}")
        return KohdeTyyppi.ASUINKIINTEISTO

    # 6. Jos mikään ehto ei täyttynyt, kyseessä on muu kohde
    if hasattr(rakennus, 'prt'):
        print(f"-> MUU (Asuinrakennuksen ehdot ei täyttynyt) prt: {rakennus.prt}")
    else:
        print("- prt ei ole annettu")
    if hasattr(rakennus, 'rakennusluokka_2018'):
        print(f"- rakennusluokka_2018: {rakennus.rakennusluokka_2018}")
    else:
        print("- rakennusluokka_2018 ei ole annettu")
    if hasattr(rakennus, 'rakennuksenkayttotarkoitus'):
        print(f"- rakennuksenkayttotarkoitus: {rakennus.rakennuksenkayttotarkoitus.koodi if rakennus.rakennuksenkayttotarkoitus else None}")
    else:
        print("- rakennuksenkayttotarkoitus ei ole annettu")
    if hasattr(rakennus, 'huoneistomaara'):
        print(f"- huoneistomaara: {rakennus.huoneistomaara}")
    else:
        print("- huoneistomaara ei ole annettu")
    if hasattr(rakennus, 'rakennuksenolotila'):
        print(f"- rakennuksenolotila: {rakennus.rakennuksenolotila.koodi if rakennus.rakennuksenolotila else None}")
    else:
        print("- rakennuksenolotila ei ole annettu")
    if asukkaat:
        print(f"- asukkaat: {len(asukkaat) if asukkaat else 0}")
    else:
        print("- ei asukkaita")
    print("")
    return KohdeTyyppi.MUU


def get_hapa_aineisto(session: "Session") -> Dict[str, str]:
    """
    Hakee HAPA-aineiston tiedot muistiin.
    
    Args:
        session: Tietokantaistunto
    
    Returns:
        Dict[str, str]: Sanakirja, jossa avaimena rakennus_id_tunnus ja arvona kohdetyyppi
    """
    # Haetaan koko aineisto muistiin
    query = select(HapaAineisto.rakennus_id_tunnus, HapaAineisto.kohdetyyppi)
    
    try:
        results = session.execute(query).all()
        return {row[0]: row[1] for row in results if row[0] is not None}
    except Exception as e:
        print(f"Virhe HAPA-aineiston haussa: {str(e)}")
        return {}


def create_new_kohde(session: Session, asiakas: Asiakas, keraysalueet=None) -> Kohde:
    """
    Luo uusi kohde asiakkaan tietojen perusteella.
    
    Args:
        session: Tietokantaistunto
        asiakas: Asiakas jonka tiedoista kohde luodaan
        keraysalueet: Valinnainen dictionary keräysaluetiedoilla
            {
                'biojate': bool,
                'hyotyjate': bool 
            }
        asukkaat: Valinnainen set asukkaista kohdetyypin määritystä varten
    Returns:
        Kohde: Luotu kohdeobjekti
    """

    # Tarkista rakennusten perusteella
    kohdetyyppi = KohdeTyyppi.MUU
    try:
        for prt in asiakas.rakennukset:
            rakennus = session.query(Rakennus).filter(Rakennus.prt == prt).first()
            if rakennus:           
                # Hae rakennuksen asukkaat suoraan RakennuksenVanhimmat-taulusta
                asukkaat = set(session.query(RakennuksenVanhimmat)
                    .filter(RakennuksenVanhimmat.rakennus_id == rakennus.id)
                    .all())
                
                building_type = determine_kohdetyyppi(session, rakennus, asukkaat)
                if building_type in (KohdeTyyppi.ASUINKIINTEISTO):
                    kohdetyyppi = building_type
                    break
    
    except Exception as e:
        logger = logging.getLogger(__name__)
        logger.error(f"Virhe rakennusten tyypin määrityksessä: {str(e)}")
        # Jatketaan oletustyypillä MUU

    kohde = Kohde(
        nimi=form_display_name(asiakas.haltija),
        kohdetyyppi=codes.kohdetyypit[kohdetyyppi],
        alkupvm=asiakas.voimassa.lower,
        loppupvm=asiakas.voimassa.upper,
    )

    return kohde

def parse_alkupvm_for_kohde(
    session: "Session",
    rakennus_ids: List[int],
    old_kohde_alkupvm: datetime.date,
    poimintapvm: Optional[datetime.date],
) -> datetime.date:
    """
    Määrittää uuden kohteen alkupäivämäärän määrityksen mukaisesti.
    
    Alkupäivä määräytyy seuraavien sääntöjen mukaan:
    - Jos rakennusten omistajissa/asukkaissa ei ole muutoksia -> poimintapvm
    - Muuten valitaan uusin seuraavista:
        1. Uusin omistajan muutospäivä
        2. Uusin asukkaan muutospäivä 
        3. Vanhan kohteen alkupvm

    Args:
        session: Tietokantaistunto
        rakennus_ids: Lista käsiteltävien rakennusten ID:istä
        old_kohde_alkupvm: Vanhan kohteen alkupäivämäärä 
        poimintapvm: Uusi poimintapäivämäärä

    Returns:
        datetime.date: Määritetty alkupäivämäärä

    Raises:
        SQLAlchemyError: Jos tietokantakyselyssä tapahtuu virhe
    """
    try:
        # Hae uusin omistajan muutospäivä
        latest_omistaja_change = (
            session.query(func.max(RakennuksenOmistajat.omistuksen_alkupvm))
            .filter(RakennuksenOmistajat.rakennus_id.in_(rakennus_ids))
            .scalar()
        )

        # Hae uusin asukkaan muutospäivä
        latest_vanhin_change = (
            session.query(func.max(RakennuksenVanhimmat.alkupvm))
            .filter(RakennuksenVanhimmat.rakennus_id.in_(rakennus_ids))
            .scalar()
        )

        # Jos ei muutoksia kummassakaan, palauta poimintapvm
        if latest_omistaja_change is None and latest_vanhin_change is None:
            return poimintapvm

        # Valitse uusin päivämäärä muutoksista
        latest_change = old_kohde_alkupvm
        if latest_omistaja_change is None and latest_vanhin_change is not None:
            latest_change = latest_vanhin_change
        elif latest_vanhin_change is None and latest_omistaja_change is not None:
            latest_change = latest_omistaja_change
        else:
            latest_change = max(latest_omistaja_change, latest_vanhin_change)

        # Jos uusin muutos on vanhan kohteen alkupvm:n jälkeen,
        # käytetään muutospäivää, muuten poimintapvm:ää
        if latest_change > old_kohde_alkupvm:
            return latest_change
        else:
            return poimintapvm

    except SQLAlchemyError as e:
        print(
            f"Virhe alkupäivämäärän määrityksessä rakennuksille {rakennus_ids}: {str(e)}"
        )
        raise


def create_new_kohde_from_buildings(
    session: Session,
    rakennus_ids: List[int],
    asukkaat: Set[Osapuoli],
    omistajat: Set[Osapuoli],
    poimintapvm: Optional[datetime.date],
    loppupvm: Optional[datetime.date],
    old_kohde: Optional[Kohde],
    lukittu: bool = False
):
    """
    Luo uuden kohteen annettujen rakennusten perusteella ja yhdistää niihin liittyvät tiedot.
    
    Kohteen nimeäminen priorisoidaan seuraavassa järjestyksessä:
    1. Asunto-osakeyhtiön nimi
    2. Yrityksen nimi 
    3. Yhteisön nimi
    4. Vanhimman asukkaan nimi
    5. Omistajan nimi
    Jos mitään edellä mainituista ei löydy, kohde nimetään "Tuntematon".

    Alkupäivämäärä määräytyy seuraavasti:
    - Jos kyseessä on vanhan kohteen päivitys, käytetään parse_alkupvm_for_kohde -funktion määrittämää päivää
    - Muussa tapauksessa käytetään annettua poimintapäivämäärää
    
    Loppupäivämääräksi asetetaan:
    - Annettu loppupvm, jos se on määritelty
    - Muussa tapauksessa järjestelmän oletusarvo (31.12.2100)

    Args:
        session: Tietokantaistunto
        rakennus_ids: Lista rakennusten ID-tunnisteista
        asukkaat: Rakennusten asukkaat 
        omistajat: Rakennusten omistajat
        poimintapvm: Uuden kohteen alkupäivämäärä
        loppupvm: Uuden kohteen loppupäivämäärä
        old_kohde: Vanha kohde, jos kyseessä on päivitys (vapaaehtoinen)

    Returns:
        Kohde: Luotu kohdeobjekti kaikkine riippuvuuksineen

    Huomautukset:
        - Funktio lisää kaikki asukkaat VANHIN_ASUKAS -roolilla
        - Omistajat lisätään OMISTAJA-roolilla, vaikka olisivat myös asukkaita
        - Omistajatiedot tallennetaan aina, jotta kohde voidaan tunnistaa myöhemmissä tuonneissa
        - Muutokset tallennetaan session.flush()-komennolla, mutta ei commitoida
    """
    # Alustetaan asiakas None:ksi
    asiakas = None
    
    if omistajat:
        # prefer companies over private owners when naming combined objects
        asoy_asiakkaat = {osapuoli for osapuoli in omistajat if is_asoy(osapuoli.nimi)}
        company_asiakkaat = {
            osapuoli for osapuoli in omistajat if is_company(osapuoli.nimi)
        }
        yhteiso_asiakkaat = {
            osapuoli for osapuoli in omistajat if is_yhteiso(osapuoli.nimi)
        }
        if asoy_asiakkaat:
            asiakas = min(asoy_asiakkaat, key=lambda x: x.nimi)
        elif company_asiakkaat:
            asiakas = min(company_asiakkaat, key=lambda x: x.nimi)
        elif yhteiso_asiakkaat:
            asiakas = min(yhteiso_asiakkaat, key=lambda x: x.nimi)
        else:
            # Jos ei löydy yritystä/yhteisöä, käytetään ensimmäistä omistajaa
            asiakas = min(omistajat, key=lambda x: x.nimi)
            
    # Jos ei omistajaa, käytetään asukasta
    if not asiakas and asukkaat:
        asiakas = min(asukkaat, key=lambda x: x.nimi)

    if asiakas:
        kohde_display_name = form_display_name(
            Yhteystieto(
                asiakas.nimi,
                asiakas.katuosoite,
                asiakas.ytunnus,
                asiakas.henkilotunnus,
            )
        )
    else:
        kohde_display_name = "Tuntematon"
        
    alkupvm = poimintapvm
    if old_kohde:
        alkupvm = parse_alkupvm_for_kohde(
            session, rakennus_ids, old_kohde.alkupvm, poimintapvm
        )

        
    # Määritä kohdetyyppi rakennusten perusteella
    kohdetyyppi = KohdeTyyppi.MUU  # Oletuksena MUU
    found_asuinkiinteisto = False
    for rakennus_id in rakennus_ids:
        rakennus = session.query(Rakennus).filter(Rakennus.id == rakennus_id).first()
        if rakennus:
            building_type = determine_kohdetyyppi(session, rakennus, asukkaat)
            if building_type == KohdeTyyppi.ASUINKIINTEISTO:
                found_asuinkiinteisto = True
                break

    if not found_asuinkiinteisto:
        kohdetyyppi = KohdeTyyppi.MUU
    else:
        kohdetyyppi = KohdeTyyppi.ASUINKIINTEISTO

    kohde = Kohde(
        nimi=kohde_display_name,
        kohdetyyppi=codes.kohdetyypit[kohdetyyppi],
        alkupvm=alkupvm,
        loppupvm=loppupvm,
        lukittu=lukittu
    )
    session.add(kohde)
    # we need to get the id for the kohde from db
    session.flush()

    # create dependent objects
    for rakennus_id in rakennus_ids:
        kohteen_rakennus = KohteenRakennukset(
            rakennus_id=rakennus_id, kohde_id=kohde.id
        )
        session.add(kohteen_rakennus)
    # save all asukkaat as asiakas
    for osapuoli in asukkaat:
        asiakas = KohteenOsapuolet(
            osapuoli_id=osapuoli.id,
            kohde_id=kohde.id,
            osapuolenrooli=codes.osapuolenroolit[OsapuolenrooliTyyppi.VANHIN_ASUKAS]
        )
        session.add(asiakas)
    # save all omistajat as yhteystiedot, even if they also live in the building.
    # This is needed so the kohde can be identified by owner at later import.
    for osapuoli in omistajat:
        yhteystieto = KohteenOsapuolet(
            osapuoli_id=osapuoli.id,
            kohde_id=kohde.id,
            osapuolenrooli=codes.osapuolenroolit[
                OsapuolenrooliTyyppi.OMISTAJA
            ],
        )
        session.add(yhteystieto)
    return kohde


def old_kohde_for_buildings(
    session: Session, rakennus_ids: List[int], poimintapvm: datetime.date
):
    kohde_query = (
        select(Kohde)
        .join(KohteenRakennukset)
        .filter(
            KohteenRakennukset.rakennus_id.in_(rakennus_ids),
            Kohde.loppupvm == poimintapvm - timedelta(days=1),
        )
    )
    old_kohde_id = session.execute(kohde_query).scalar()
    return old_kohde_id


def set_old_kohde_loppupvm(session: Session, kohde_id: int, loppupvm: datetime.date):
    session.execute(
        update(Kohde)
        .where(Kohde.id == kohde_id)
        .values(loppupvm=loppupvm)
    )
    session.commit()


def set_paatos_loppupvm_for_old_kohde(
    session: Session, kohde_id: int, loppupvm: datetime.date
):
    """
    Asettaa päätösten loppupäivämäärän vanhan kohteen päätöksille.
    """
    # Muunnetaan subquery explisiittiseksi select-lauseeksi
    rakennus_ids = select(KohteenRakennukset.rakennus_id).where(
        KohteenRakennukset.kohde_id == kohde_id
    )

    session.query(Viranomaispaatokset).filter(
        and_(
            Viranomaispaatokset.rakennus_id.in_(rakennus_ids.scalar_subquery()),
            Viranomaispaatokset.alkupvm <= loppupvm
        )
    ).update({Viranomaispaatokset.loppupvm: loppupvm}, synchronize_session=False)
    session.commit()


def update_kompostori(
    session: Session, old_kohde_id: int, loppupvm: datetime.date, new_kohde_id: int
):
    """
    Päivittää kompostorien tiedot kohteen vaihtuessa.
    """
    # Haetaan vanhan kohteen kompostorit
    kompostori_ids = select(KompostorinKohteet.kompostori_id).where(
        KompostorinKohteet.kohde_id == old_kohde_id
    )

    # Asetetaan loppupvm
    session.query(Kompostori).filter(
        and_(
            Kompostori.id.in_(kompostori_ids.scalar_subquery()),
            Kompostori.alkupvm <= loppupvm
        )
    ).update({Kompostori.loppupvm: loppupvm}, synchronize_session=False)

    # Haetaan kompostorit jotka jatkuvat loppupvm:n jälkeen
    kompostori_id_by_date = select(Kompostori.id).where(
        and_(
            Kompostori.id.in_(kompostori_ids.scalar_subquery()),
            Kompostori.alkupvm > loppupvm
        )
    )

    # Haetaan kompostorien osapuolet
    kompostori_osapuoli_ids = select(Kompostori.osapuoli_id).where(
        Kompostori.id.in_(kompostori_id_by_date.scalar_subquery())
    )

    # Päivitetään osapuolten kohde
    session.query(KohteenOsapuolet).filter(
        and_(
            KohteenOsapuolet.osapuoli_id.in_(kompostori_osapuoli_ids.scalar_subquery()),
            KohteenOsapuolet.kohde_id == old_kohde_id,
            KohteenOsapuolet.osapuolenrooli_id == 311
        )
    ).update({KohteenOsapuolet.kohde_id: new_kohde_id}, synchronize_session=False)

    # Päivitetään KompostorinKohteet
    session.query(KompostorinKohteet).filter(
        and_(
            KompostorinKohteet.kompostori_id.in_(kompostori_id_by_date.scalar_subquery()),
            KompostorinKohteet.kohde_id == old_kohde_id,
        )
    ).update({KompostorinKohteet.kohde_id: new_kohde_id}, synchronize_session=False)
    
    session.commit()


def move_sopimukset_and_kuljetukset_to_new_kohde(
    session: Session, alkupvm: datetime.date, old_kohde_id: int, new_kohde_id: int
):
    session.execute(
        update(Sopimus)
        .where(Sopimus.kohde_id == old_kohde_id)
        .where(Sopimus.loppupvm >= alkupvm)
        .values(kohde_id=new_kohde_id)
    )
    session.execute(
        update(Kuljetus)
        .where(Kuljetus.kohde_id == old_kohde_id)
        .where(Kuljetus.loppupvm >= alkupvm)
        .values(kohde_id=new_kohde_id)
    )
    session.commit()

def check_and_update_old_other_building_kohde_kohdetyyppi(
    session: Session,
    poimintapvm: Optional[datetime.date]
) -> List[int]:
    """
    Etsii vanhat kohteet joilla on kohdetyyppi MUU

    Toiminta:
    1. Hakee kannasta ennen poimintavuotta tuodut kohteet joilla on kohdetyyppinä MUU
    2. Tarkistaa kohteisiin kohdistettujen rakennusten avulla kohdetyypin
    3. Päivittää kohdetyypin vastaamaan oikeaa tilannetta
    
    Args:
        session: Tietokantaistunto
        poimintapvm: Tarkasteltavien kohteiden alkupäivämäärä on ennen tämän päivän vuotta

    Returns:
        Kohteet: päivitetyt kohteet
    """

    maxYear = poimintapvm.year


    print(f"\n\nEtsitään kohteita, joiden alkupvm on < {datetime.date(maxYear, 1, 1)}")

    updated_kohteet = set()

    try:
        # Hae olemassa olevat kohteet ennen poimintapäivän vuotta
        kohteet = (
            session.query(Kohde)
            .join(KohteenRakennukset, KohteenRakennukset.kohde_id == Kohde.id)
            .filter(Kohde.kohdetyyppi_id == 8, Kohde.alkupvm < datetime.date(maxYear, 1, 1))
            .all()
        )
        
        print(f"Löydetty {len(kohteet)} vanhaa muuta kohdetta")
        
        for kohde in kohteet:
            print(f"\n\nKohde ID: {kohde.id}, Tyyppi: {kohde.kohdetyyppi_id}")
            original_kohdetyyppi = kohde.kohdetyyppi_id
            
            # Hae kohteen rakennukset
            rakennus_query = (
                select(
                    Rakennus.id, 
                    Rakennus.prt, 
                    Rakennus.rakennusluokka_2018, 
                    Rakennus.rakennuksenkayttotarkoitus_koodi, 
                    Rakennus.huoneistomaara, 
                    Rakennus.rakennuksenolotila_koodi
                )
                .join(KohteenRakennukset, KohteenRakennukset.rakennus_id == Rakennus.id)
                .where(KohteenRakennukset.kohde_id == kohde.id)
            )
            rakennukset = session.execute(rakennus_query).all()
            print(f"- Rakennukset: {[r[1] for r in rakennukset]}")
            
            found_asuinkiinteisto = False
            
            for rakennus_tiedot in rakennukset:  
                # Hae rakennuksen asukkaat suoraan RakennuksenVanhimmat-taulusta
                asukkaat = set(session.query(RakennuksenVanhimmat)
                    .filter(RakennuksenVanhimmat.rakennus_id == rakennus_tiedot.id)
                    .all())     
                building_type = determine_kohdetyyppi(session, rakennus_tiedot, asukkaat)
                print(f"Tarkastelussa kohdetyypin arvo {original_kohdetyyppi} vs {codes.kohdetyypit[building_type].id} kohteelle {kohde.id} prt:llä {rakennus_tiedot.prt}")
                
                if building_type == KohdeTyyppi.ASUINKIINTEISTO:
                    if building_type != original_kohdetyyppi:
                        print(f"Päivitetään kohdetyypin arvo {original_kohdetyyppi} arvoksi {codes.kohdetyypit[building_type].id} kohteelle {kohde.id} asuinkiinteistö")
                        kohde.kohdetyyppi_id = codes.kohdetyypit[building_type].id
                        updated_kohteet.add(kohde.id)

                # Jos yksikään rakennus ei ole asuinkiinteistö, asetetaan tyypiksi MUU
                if not found_asuinkiinteisto:
                    new_kohdetyyppi = codes.kohdetyypit[KohdeTyyppi.MUU]
                    if new_kohdetyyppi.id != original_kohdetyyppi:
                        print(f"Päivitetään kohdetyypin arvo {original_kohdetyyppi} arvoksi {new_kohdetyyppi.id} kohteelle {kohde.id} ei asuinkiinteistö")
                        setattr(kohde, 'kohdetyyppi_id', new_kohdetyyppi.id)
                        updated_kohteet.add(kohde.id)
        
        if len(updated_kohteet) > 0:
            print(f"Päivitetty {len(updated_kohteet)} kohdetta")
            session.flush()        
    except NoResultFound:
        updated_kohteet = []
        
    return updated_kohteet


def update_or_create_kohde_from_buildings(
    session: Session,
    dvv_rakennustiedot: Dict[int, Rakennustiedot],
    rakennukset: Set[Rakennustiedot],
    asukkaat: Set[Osapuoli],
    omistajat: Set[Osapuoli],
    poimintapvm: Optional[datetime.date],
    loppupvm: Optional[datetime.date],
    lukittu: bool = False
) -> Kohde:
    """
    Optimoitu versio kohteen päivitys/luontifunktiosta.
    Luo uuden kohteen tai päivittää olemassa olevaa annettujen rakennusten perusteella.

    Toiminta:
    1. Hakee rakennusten perusteella mahdollisen olemassa olevan kohteen
    2. Jos kohdetta ei löydy, luo uusi
    3. Jos kohde löytyy, päivittää sen tiedot
    4. Käsittelee vanhaan kohteeseen liittyvät tietojen siirrot ja päivitykset

    Optimoinnit:
    - Käyttää eksplisiittisiä join-määrittelyjä kyselyissä
    - Vähentää tietokantakyselyiden määrää kokoamalla tietoja muistiin
    - Käyttää tehokkaita bulk-operaatioita päivityksiin
    - Minimoi sarakkeiden päivitykset
    
    Args:
        session: Tietokantaistunto
        dvv_rakennustiedot: Rakennustietojen lookup-taulukko
        rakennukset: Käsiteltävät rakennukset
        asukkaat: Rakennusten asukkaat 
        omistajat: Rakennusten omistajat
        poimintapvm: Uuden kohteen alkupäivämäärä
        loppupvm: Uuden kohteen loppupäivämäärä
        lukittu: Valinnainen parametri perusmaksurekisterikohteiden käsittelyyn

    Returns:
        Kohde: Luotu tai päivitetty kohde
    """
    rakennus_ids = set()
    rakennus_prts = set()
    for rakennustiedot in rakennukset:
        if isinstance(rakennustiedot, tuple):
            rakennus_ids.add(rakennustiedot[0].id)
            rakennus_prts.add(rakennustiedot[0].prt)
            print(f"DEBUG: Tuple rakennus {rakennus.id}, {rakennus.prt} data:", rakennus.__dict__)
        else:
            print(f"DEBUG: not Tuple rakennus {rakennustiedot.id}, {rakennustiedot.prt}")
            rakennus_ids.add(rakennustiedot.id)
            rakennus_prts.add(rakennustiedot.prt)

    asukas_ids = {osapuoli.id for osapuoli in asukkaat}
    omistaja_ids = {osapuoli.id for osapuoli in omistajat}
    
    logger = logging.getLogger(__name__)
    print("")
    print(
        f"Etsitään kohdetta: rakennukset={rakennus_ids}, prts={rakennus_prts}, "
        f"asukkaat={asukas_ids}, omistajat={omistaja_ids}"
    )
    logger.debug(
        f"Etsitään kohdetta: rakennukset={rakennus_ids}, prts={rakennus_prts}, "
        f"asukkaat={asukas_ids}, omistajat={omistaja_ids}"
    )

    # Hae olemassa oleva kohde eksplisiittisillä join-määrittelyillä
    kohde_query = (
        select(Kohde.id, Osapuoli.nimi)
        .select_from(Kohde)
        .join(KohteenRakennukset, KohteenRakennukset.kohde_id == Kohde.id)
        .join(KohteenOsapuolet, KohteenOsapuolet.kohde_id == Kohde.id, isouter=True)
        .join(Osapuoli, Osapuoli.id == KohteenOsapuolet.osapuoli_id, isouter=True)
        .where(KohteenRakennukset.rakennus_id.in_(rakennus_ids))
    )
    
    # Jos on aikarajaus, lisää se kyselyyn
    if poimintapvm and loppupvm:
        kohde_query = kohde_query.where(
            Kohde.voimassaolo.overlaps(DateRange(poimintapvm, loppupvm))
        )

    try:
        kohteet = session.execute(kohde_query).all()
        print(f"Löydetty {len(kohteet)} kohdetta")
        for kohde in kohteet:
            print(f"- Kohde ID: {kohde[0]}, Nimi: {kohde[1]}")
            # Hae kohteen rakennukset
            rakennus_query = (
                select(Rakennus.id, Rakennus.prt)
                .join(KohteenRakennukset, KohteenRakennukset.rakennus_id == Rakennus.id)
                .where(KohteenRakennukset.kohde_id == kohde[0])
            )
            rakennukset = session.execute(rakennus_query).all()
            print(f"  Rakennukset: {[r[1] for r in rakennukset]}")
        logger.debug(f"Löydetty {len(kohteet)} kohdetta")
    except NoResultFound:
        kohteet = []

    # Jos kohdetta ei löydy, luo uusi
    if len(kohteet) == 0:
        print("Kohdetta ei löytynyt, luodaan uusi")
        logger.debug("Kohdetta ei löytynyt, luodaan uusi")
        
        # Tarkista mahdollinen vanha kohde
        old_kohde = None
        if poimintapvm:
            old_kohde = old_kohde_for_buildings(session, list(rakennus_ids), poimintapvm)
            
        # Määritä alkupvm
        alkupvm = poimintapvm
        if old_kohde:
            alkupvm = parse_alkupvm_for_kohde(
                session, rakennus_ids, old_kohde.alkupvm, poimintapvm
            )
            
        # Luo uusi kohde
        new_kohde = create_new_kohde_from_buildings(
            session,
            rakennus_ids,
            asukkaat,
            omistajat,
            alkupvm,
            loppupvm,
            old_kohde if 'old_kohde' in locals() else None,
            lukittu=lukittu
        )
        
        # Käsittele vanhan kohteen tiedot
        if new_kohde and poimintapvm and old_kohde:
            logger.debug(f"Päivitetään vanhan kohteen {old_kohde.id} tiedot")
            update_old_kohde_data(
                session,
                old_kohde.id,
                new_kohde.id,
                new_kohde.alkupvm
            )
            
        return new_kohde

    # Jos kohde löytyi, päivitä sen tiedot
    found_kohde = session.get(Kohde, kohteet[0][0])

    # Päivitä kohteen voimassaoloaika
    needs_update = False
    if not poimintapvm or (found_kohde.alkupvm and poimintapvm < found_kohde.alkupvm):
        found_kohde.alkupvm = poimintapvm
        needs_update = True
        
    if poimintapvm and found_kohde.loppupvm == poimintapvm - timedelta(days=1):
        found_kohde.loppupvm = None
        needs_update = True
    
    original_kohdetyyppi = found_kohde.kohdetyyppi
    found_asuinkiinteisto = False

    for rakennus_tiedot in rakennukset:       
        # Yritä hakea täydelliset rakennustiedot dvv_rakennustiedot lookup-taulukosta
        if isinstance(rakennus_tiedot, tuple):
            rakennus = rakennus_tiedot[0]
        else:
            rakennus = rakennus_tiedot
            
        # Käytä dvv_rakennustiedot lookup-taulukkoa jos mahdollista
        if rakennus.id in dvv_rakennustiedot:
            rakennus_tiedot = dvv_rakennustiedot[rakennus.id]
            rakennus = rakennus_tiedot[0]
            print(f"Käytetään DVV rakennustietoja rakennukselle {rakennus.id}")
        
        building_type = determine_kohdetyyppi(session, rakennus, asukkaat)
        if building_type == KohdeTyyppi.ASUINKIINTEISTO:
            found_asuinkiinteisto = True
            new_kohdetyyppi = codes.kohdetyypit[building_type]
            if new_kohdetyyppi != original_kohdetyyppi:
                print(f"Päivitetään kohdetyypin arvo {new_kohdetyyppi} kohdeelle {found_kohde.id}")
                found_kohde.kohdetyyppi = new_kohdetyyppi
                needs_update = True
            break

    # Jos yksikään rakennus ei ole asuinkiinteistö, asetetaan tyypiksi MUU
    if not found_asuinkiinteisto:
        new_kohdetyyppi = codes.kohdetyypit[KohdeTyyppi.MUU]
        if new_kohdetyyppi != original_kohdetyyppi:
            found_kohde.kohdetyyppi = new_kohdetyyppi
            needs_update = True

    if needs_update:
        print(f"Päivitetty kohde: ID={found_kohde.id}, Alkupvm={found_kohde.alkupvm}, Loppupvm={found_kohde.loppupvm}")
        session.flush()
        
    return found_kohde


def get_kohde_by_asiakasnumero(
    session: "Session", tunnus: "Tunnus"
) -> "Union[Kohde, None]":
    query = (
        select(Kohde)
        .join(UlkoinenAsiakastieto)
        .where(
            UlkoinenAsiakastieto.tiedontuottaja_tunnus == tunnus.jarjestelma,
            UlkoinenAsiakastieto.ulkoinen_id == tunnus.tunnus,
        )
    )
    try:
        kohde = session.execute(query).scalar_one()
    except NoResultFound:
        kohde = None

    return kohde


@lru_cache(maxsize=100)
def get_or_create_pseudokohde(session: Session, nimi: str, kohdetyyppi) -> Kohde:
    kohdetyyppi = codes.kohdetyypit[kohdetyyppi]
    query = select(Kohde).where(Kohde.nimi == nimi, Kohde.kohdetyyppi == kohdetyyppi)
    try:
        kohde = session.execute(query).scalar_one()
    except NoResultFound:
        kohde = Kohde(nimi=nimi, kohdetyyppi=kohdetyyppi)
        session.add(kohde)

    return kohde


def get_or_create_kohteet_from_kiinteistot(
    session: Session,
    kiinteistotunnukset: "Select", 
    poimintapvm: "Optional[datetime.date]",
    loppupvm: "Optional[datetime.date]",
) -> "List[Kohde]":
    """
    Luo vähintään yksi kohde jokaisesta kiinteistötunnuksesta, jonka select-kysely palauttaa,
    jos kiinteistotunnuksella on rakennuksia ilman olemassa olevaa kohdetta annetulla aikavälillä.

    Prosessi:
    1. Erotellaan kiinteistöjen rakennukset klustereihin etäisyyden perusteella
       - Jotkin kiinteistöt voivat sisältää kaukana toisistaan olevia rakennuksia
       - Osoitteet voivat olla virheellisiä, joten erottelu tehdään etäisyyden perusteella
    
    2. Käsittele samassa klusterissa olevat rakennukset:
       - Jos kaikki rakennukset ovat saman asunto-osakeyhtiön omistamia, luodaan yksi kohde
       - Muuten käsitellään omistajittain:
         * Ensimmäinen kohde sisältää eniten rakennuksia omistavan omistajan rakennukset
         * Toinen kohde sisältää toiseksi eniten rakennuksia omistavan rakennukset
         * Prosessi jatkuu kunnes kaikilla rakennuksilla on kohde
         * Ilman omistajaa olevat rakennukset saavat oman kohteensa

    3. Lopuksi erotellaan saman omistajan rakennukset osoitteen perusteella,
       paitsi jos kyseessä on asunto-osakeyhtiö.

    Args:
        session: Tietokantaistunto
        kiinteistotunnukset: Select-kysely joka palauttaa kiinteistötunnukset
        poimintapvm: Uusien kohteiden alkupäivämäärä
        loppupvm: Uusien kohteiden loppupäivämäärä

    Returns:
        Lista luoduista kohteista

    Raises:
        SQLAlchemyError: Jos tietokantaoperaatioissa tapahtuu virhe
    """
    logger = logging.getLogger(__name__)

    # Lataa kiinteistötunnukset ja logita määrä
    kiinteistotunnukset = [
        result[0] for result in session.execute(kiinteistotunnukset).all()
    ]
    logger.info(f"{len(kiinteistotunnukset)} tuotavaa kiinteistötunnusta löydetty")
    print(f"{len(kiinteistotunnukset)} tuotavaa kiinteistötunnusta löydetty")
    
    # Hae DVV:n rakennustiedot ilman kohdetta ja logita määrä
    dvv_rakennustiedot = get_dvv_rakennustiedot_without_kohde(
        session, poimintapvm, loppupvm
    )
    logger.info(
        f"Löydetty {len(dvv_rakennustiedot)} DVV-rakennusta ilman voimassaolevaa kohdetta"
    )
    print(f"Löydetty {len(dvv_rakennustiedot)} DVV-rakennusta ilman voimassaolevaa kohdetta")

    # Ryhmittele rakennustiedot kiinteistötunnuksittain
    rakennustiedot_by_kiinteistotunnus = defaultdict(set)
    for rakennus, vanhimmat, omistajat, osoitteet in dvv_rakennustiedot.values():
        rakennustiedot_by_kiinteistotunnus[rakennus.kiinteistotunnus].add(
            (rakennus, vanhimmat, omistajat, osoitteet)
        )

    # Hae omistajatiedot ja muodosta lookup-taulut
    rakennus_owners = session.execute(
        select(RakennuksenOmistajat.rakennus_id, Osapuoli).join(
            Osapuoli, RakennuksenOmistajat.osapuoli_id == Osapuoli.id
        )
    ).all()
    
    owners_by_rakennus_id = defaultdict(set)  # rakennus_id -> {omistajat}
    rakennus_ids_by_owner_id = defaultdict(set)  # omistaja.id -> {rakennus_id:t}
    for (rakennus_id, owner) in rakennus_owners:
        owners_by_rakennus_id[rakennus_id].add(owner)
        rakennus_ids_by_owner_id[owner.id].add(rakennus_id)

    rakennus_inhabitants = session.execute(
        select(RakennuksenVanhimmat.rakennus_id, Osapuoli).join(
            Osapuoli, RakennuksenVanhimmat.osapuoli_id == Osapuoli.id
        )
    ).all()
    
    inhabitants_by_rakennus_id = defaultdict(set)  # rakennus_id -> {asukkaat}
    for (rakennus_id, inhabitant) in rakennus_inhabitants:
        inhabitants_by_rakennus_id[rakennus_id].add(inhabitant)

    rakennus_addresses = session.execute(
        select(Rakennus.id, Osoite)
        .join(Osoite, Rakennus.id == Osoite.rakennus_id)
    ).all()
    
    addresses_by_rakennus_id = defaultdict(set)  # rakennus_id -> {osoitteet}
    for (rakennus_id, address) in rakennus_addresses:
        addresses_by_rakennus_id[rakennus_id].add(address)

    # Lista muodostettavista rakennusryhmistä
    building_sets: List[Set[Rakennustiedot]] = []

    # Käsittele kaikki kiinteistöt
    for kiinteistotunnus in kiinteistotunnukset:

        # Kerää kiinteistön rakennukset
        rakennustiedot_to_add = rakennustiedot_by_kiinteistotunnus[kiinteistotunnus]
        if not rakennustiedot_to_add:
            print(f"Ei rakennuksia kiinteistötunnukselle: {kiinteistotunnus}")
            continue

        # 1. Jaa rakennukset klustereihin etäisyyden perusteella
        clustered_rakennustiedot = _cluster_rakennustiedot(
            rakennustiedot_to_add, DISTANCE_LIMIT
        )

        # Käsittele klusterit yksitellen
        for rakennustiedot_cluster in clustered_rakennustiedot:
            print(f"\nKäsitellään klusteri:")
            for rakennus, _, _, _ in rakennustiedot_cluster:
                print(f"- Rakennus {rakennus.id} (PRT: {rakennus.prt})")
            
            cluster_ids = {rakennus.id for rakennus, _, _, _ in rakennustiedot_cluster}
            
            # Kerää klusterin omistajat ja niiden omistamat rakennukset
            cluster_owners_buildings = defaultdict(set)
            for rakennus_id in cluster_ids:
                for owner in owners_by_rakennus_id[rakennus_id]:
                    cluster_owners_buildings[owner].add(rakennus_id)
            
            # Järjestä omistajat sen mukaan kuinka monta rakennusta omistavat
            sorted_owners = sorted(
                cluster_owners_buildings.items(),
                key=lambda x: len(x[1]),
                reverse=True
            )
            
            print("\nOmistajat:")
            for owner, buildings in sorted_owners:
                print(f"- {owner.nimi} omistaa {len(buildings)} rakennusta: {[session.get(Rakennus, id).prt for id in buildings]}")
            
            # 2. Jaa klusteri omistajittain
            remaining_ids = set(cluster_ids)
            
            while remaining_ids:
                # 2.1 Jaa ensin omistajien mukaan
                if cluster_owners_buildings:
                    # Aloita suurimmasta omistajaryhmästä
                    owner_id, owner_buildings = max(
                        cluster_owners_buildings.items(),
                        key=lambda x: len(x[1])
                    )
                    if owner_buildings:  # Varmista että rakennuksia löytyy
                        buildings_to_process = remaining_ids & owner_buildings
                        
                        # Tarkista omistaja asoy
                        owner = next(
                            (owner for owner in owners_by_rakennus_id[next(iter(owner_buildings))]
                            if owner.id == owner_id),
                            None  # Default arvo jos omistajaa ei löydy
                        )
                        is_owner_asoy = is_asoy(owner.nimi) if owner else False
                    else:
                        owner = None
                        is_owner_asoy = False
                else:
                    # Jos ei omistajia, käsittele kaikki jäljellä olevat
                    buildings_to_process = remaining_ids
                    is_owner_asoy = False

                # 2.2 Jos ei ole asoy, jaa vielä osoitteen mukaan
                if not is_owner_asoy:
                    buildings_by_address = defaultdict(set)
                    for building_id in buildings_to_process:
                        for address in addresses_by_rakennus_id.get(building_id, set()):
                            key = (address.katu_id, address.osoitenumero)
                            buildings_by_address[key].add(building_id)

                    # Käsittele osoiteryhmät suuruusjärjestyksessä
                    while buildings_to_process:
                        if buildings_by_address:
                            # Aloita suurimmasta osoiteryhmästä
                            _, address_group = max(
                                buildings_by_address.items(),
                                key=lambda x: len(x[1])
                            )
                            current_group = buildings_to_process & address_group
                        else:
                            # Jos ei osoitteita, käsittele kaikki kerralla
                            current_group = buildings_to_process

                        # Luo rakennusryhmä
                        rakennustiedot_group = {
                            dvv_rakennustiedot[id] 
                            for id in current_group 
                            if id in dvv_rakennustiedot
                        }
                        
                        if rakennustiedot_group:
                            building_sets.append(rakennustiedot_group)

                        # Päivitä jäljellä olevat rakennukset
                        buildings_to_process -= current_group
                        remaining_ids -= current_group

                        # Päivitä osoiteryhmät
                        if buildings_by_address:
                            buildings_by_address = {
                                addr: buildings - current_group
                                for addr, buildings in buildings_by_address.items()
                                if buildings - current_group
                            }
                else:
                    # Asoy - kaikki omistajan rakennukset samaan ryhmään
                    rakennustiedot_group = {
                        dvv_rakennustiedot[id] 
                        for id in buildings_to_process 
                        if id in dvv_rakennustiedot
                    }
                    if rakennustiedot_group:
                        building_sets.append(rakennustiedot_group)
                    remaining_ids -= buildings_to_process

                # Päivitä omistajaryhmät
                if cluster_owners_buildings:
                    cluster_owners_buildings.pop(owner_id, None)

    # Luo kohteet rakennusryhmien perusteella
    kohteet = get_or_create_kohteet_from_rakennustiedot(
        session,
        dvv_rakennustiedot,
        building_sets,
        owners_by_rakennus_id,
        inhabitants_by_rakennus_id,
        poimintapvm,
        loppupvm,
    )


    return kohteet

def get_or_create_single_asunto_kohteet(
    session: Session,
    poimintapvm: "Optional[datetime.date]",
    loppupvm: "Optional[datetime.date]",
) -> "List[Kohde]":
    """
    Hae tai luo kohteet kaikille yhden asunnon taloille ja paritaloille, joilla ei ole
    kohdetta määritellyllä aikavälillä. Huomioi myös talot, joissa ei ole asukkaita.

    Jos kiinteistöllä on useita asuttuja rakennuksia, sitä ei tuoda tässä.
    """
    logger = logging.getLogger(__name__)
    logger.info("\n----- LUODAAN YHDEN ASUNNON KOHTEET -----")

    # Älä tuo rakennuksia, joilla on jo olemassa olevia kohteita
    if loppupvm is None:
        rakennus_id_with_current_kohde = (
            select(Rakennus.id)
            .join(KohteenRakennukset)
            .join(Kohde)
            .filter(poimintapvm == Kohde.alkupvm)
        )
    else:
        rakennus_id_with_current_kohde = (
            select(Rakennus.id)
            .join(KohteenRakennukset)
            .join(Kohde)
            .filter(Kohde.voimassaolo.overlaps(DateRange(poimintapvm, loppupvm)))
        )

    # Määrittele sallitut rakennustyypit
    ALLOWED_TYPES_2018 = {
        '0111',  # Yhden asunnon talot
        '0110'   # Paritalot
    }
    
    ALLOWED_TYPES_OLD = {
        '011',  # Yhden asunnon talot
        '012'   # Kahden asunnon talot
    }

    # Hae rakennukset joilla ei ole kohdetta
    rakennus_id_without_kohde = (
        select(Rakennus.id)
        .filter(
            or_(
                # Tarkista ensisijaisesti 2018 luokitus
                Rakennus.rakennusluokka_2018.in_(ALLOWED_TYPES_2018),
                # Jos 2018 luokka puuttuu, käytä vanhaa luokitusta
                and_(
                    Rakennus.rakennusluokka_2018.is_(None),
                    Rakennus.rakennuksenkayttotarkoitus_koodi.in_(ALLOWED_TYPES_OLD)
                )
            )
        )
        .filter(~Rakennus.id.in_(rakennus_id_with_current_kohde))
        # Älä tuo rakennuksia, jotka on poistettu DVV:n tiedoista
        .filter(
            or_(
                Rakennus.kaytostapoisto_pvm.is_(None),
                Rakennus.kaytostapoisto_pvm > poimintapvm,
            )
        )
    )

    # Hae kiinteistöt joilla on vain yksi rakennus
    single_asunto_kiinteistotunnus = (
        select(
            Rakennus.kiinteistotunnus,
            func.count(Rakennus.id),
        )
        .filter(Rakennus.id.in_(rakennus_id_without_kohde))
        .group_by(Rakennus.kiinteistotunnus)
        .having(func.count(Rakennus.id) == 1)
    )

    return get_or_create_kohteet_from_kiinteistot(
        session, single_asunto_kiinteistotunnus, poimintapvm, loppupvm
    )


def get_or_create_multiple_and_uninhabited_kohteet(
    session: Session,
    poimintapvm: "Optional[datetime.date]",
    loppupvm: "Optional[datetime.date]",
) -> "List[Kohde]":
    """
    Luo kohteet kaikille kiinteistötunnuksille, joilla on rakennuksia ilman kohdetta
    määritellylle aikavälille. Tämä funktio käsittelee jäljellä olevat rakennukset,
    mukaan lukien:
    - Useita rakennuksia sisältävät kiinteistöt
    - Yhden asunnon talot jotka ovat samalla kiinteistöllä muiden rakennusten kanssa
    - Asumattomat rakennukset
    """
    logger = logging.getLogger(__name__)
    logger.info("\n----- LUODAAN JÄLJELLÄ OLEVAT KOHTEET -----")


    rakennus_id_with_current_kohde = (
        select(Rakennus.id)
        .join(KohteenRakennukset)
        .join(Kohde)
        .filter(
            or_(
                Kohde.voimassaolo.overlaps(DateRange(poimintapvm, loppupvm)),
                loppupvm is None
            )
        )
    )
    
    kiinteistotunnus_without_kohde = (
        select(Rakennus.kiinteistotunnus)
        .filter(~Rakennus.id.in_(rakennus_id_with_current_kohde))
        # Älä tuo rakennuksia, jotka on poistettu DVV:n tiedoista
        .filter(
            or_(
                Rakennus.kaytostapoisto_pvm.is_(None),
                Rakennus.kaytostapoisto_pvm > poimintapvm,
            )
        )
        .group_by(Rakennus.kiinteistotunnus)
    )

    return get_or_create_kohteet_from_kiinteistot(
        session, kiinteistotunnus_without_kohde, poimintapvm, loppupvm
    )


def _should_have_perusmaksu_kohde(rakennus: "Rakennus"):
    """
    Determines which buildings should be combined using perusmaksu data.
    """
    return rakennus.rakennuksenkayttotarkoitus in (
        codes.rakennuksenkayttotarkoitukset[
            RakennuksenKayttotarkoitusTyyppi.KERROSTALO
        ],
        codes.rakennuksenkayttotarkoitukset[RakennuksenKayttotarkoitusTyyppi.RIVITALO],
        codes.rakennuksenkayttotarkoitukset[RakennuksenKayttotarkoitusTyyppi.LUHTITALO],
        codes.rakennuksenkayttotarkoitukset[RakennuksenKayttotarkoitusTyyppi.KETJUTALO],
    )


def batch_update_old_kohde_data(session, updates):
    """
    Optimoitu versio batch-päivityksistä joka säilyttää alkuperäisen logiikan.
    Käyttää synchronize_session=False nopeuttamiseen missä turvallista.
    """
    if not updates:
        return

    # Kerää old_kohde_ids vain niistä updatesista missä old_kohde_id != new_kohde_id
    updates_by_old_id = {}
    for old_id, new_id, new_alkupvm in updates:
        if old_id != new_id:  # Käsitellään vain jos oikeasti siirretään uudelle kohteelle
            updates_by_old_id[old_id] = (new_id, new_alkupvm)
    
    if not updates_by_old_id:
        return
        
    old_kohde_ids = list(updates_by_old_id.keys())
    
    # 1. Päivitä kohteen loppupvm
    loppupvm_values = {
        old_id: (new_alkupvm - timedelta(days=1))
        for old_id, (new_id, new_alkupvm) in updates_by_old_id.items()
    }
    
    stmt = (
        update(Kohde)
        .where(Kohde.id.in_(old_kohde_ids))
        .values({
            Kohde.loppupvm: case(
                whens=loppupvm_values,
                value=Kohde.id
            )
        })
        .execution_options(synchronize_session=False)
    )
    session.execute(stmt)

    # 2. Siirrä sopimukset ja kuljetukset
    for model in [Sopimus, Kuljetus]:
        for old_id, (new_id, new_alkupvm) in updates_by_old_id.items():
            stmt = (
                update(model)
                .where(model.kohde_id == old_id)
                .where(model.loppupvm >= new_alkupvm)
                .values(kohde_id=new_id)
                .execution_options(synchronize_session=False)
            )
            session.execute(stmt)

    # 3. Päivitä viranomaispäätökset
    # Hae kaikki rakennukset vanhoilta kohteilta
    rakennus_ids = session.execute(
        select(KohteenRakennukset.rakennus_id)
        .where(KohteenRakennukset.kohde_id.in_(old_kohde_ids))
    ).scalars().all()

    if rakennus_ids:
        # Päivitä päätökset rakennuksittain
        for old_id, (new_id, new_alkupvm) in updates_by_old_id.items():
            kohteen_rakennus_ids = session.execute(
                select(KohteenRakennukset.rakennus_id)
                .where(KohteenRakennukset.kohde_id == old_id)
            ).scalars().all()
            
            if kohteen_rakennus_ids:
                loppupvm = new_alkupvm - timedelta(days=1)
                stmt = (
                    update(Viranomaispaatokset)
                    .where(
                        Viranomaispaatokset.rakennus_id.in_(kohteen_rakennus_ids),
                        Viranomaispaatokset.alkupvm <= loppupvm
                    )
                    .values(loppupvm=loppupvm)
                    .execution_options(synchronize_session=False)
                )
                session.execute(stmt)

    # 4. Päivitä kompostorit
    for old_id, (new_id, new_alkupvm) in updates_by_old_id.items():
        # Hae kohteen kompostorit
        kompostori_ids = select(KompostorinKohteet.kompostori_id).where(
            KompostorinKohteet.kohde_id == old_id
        )

        # Aseta loppupvm kompostoreille
        stmt = (
            update(Kompostori)
            .where(
                Kompostori.id.in_(kompostori_ids.scalar_subquery()),
                Kompostori.alkupvm <= loppupvm
            )
            .values(loppupvm=loppupvm)
            .execution_options(synchronize_session=False)
        )
        session.execute(stmt)

        # Siirrä jatkuvat kompostorit uudelle kohteelle
        stmt = (
            update(KompostorinKohteet)
            .where(
                KompostorinKohteet.kompostori_id.in_(kompostori_ids.scalar_subquery()),
                KompostorinKohteet.kohde_id == old_id,
                Kompostori.alkupvm > loppupvm
            )
            .values(kohde_id=new_id)
            .execution_options(synchronize_session=False)
        )
        session.execute(stmt)
    
    session.commit()


def create_perusmaksurekisteri_kohteet(
    session: Session,
    perusmaksutiedosto: Path,
    poimintapvm: Optional[datetime.date],
    loppupvm: Optional[datetime.date] = None,
) -> List[Kohde]:
    """
    Luo kohteet perusmaksurekisterin tietojen perusteella.
    
    Args:
        session: Tietokantaistunto
        perusmaksutiedosto: Polku perusmaksurekisterin Excel-tiedostoon
        poimintapvm: Uuden kohteen alkupäivämäärä
        loppupvm: Uuden kohteen loppupäivämäärä
        
    Returns:
        Lista luoduista kohteista

    Raises:
        FileNotFoundError: Jos perusmaksutiedostoa ei löydy
        openpyxl.utils.exceptions.InvalidFileException: Jos tiedosto ei ole validi Excel
    """
    logger = logging.getLogger(__name__)
    logger.info("\nLuodaan perusmaksurekisterin kohteet...")

    # Hae DVV:n rakennustiedot rakennuksista joilla ei vielä ole kohdetta
    dvv_rakennustiedot = get_dvv_rakennustiedot_without_kohde(
        session, poimintapvm, loppupvm
    )
    logger.info(f"Löydetty {len(dvv_rakennustiedot)} DVV-rakennusta ilman kohdetta")

    # Luo lookup PRT -> Rakennustiedot jatkojalostusta varten
    dvv_rakennustiedot_by_prt = {
        tiedot[0].prt: tiedot 
        for tiedot in dvv_rakennustiedot.values()
        if tiedot[0].prt  # Ohita jos PRT puuttuu
    }

    # Ryhmittele rakennukset asiakasnumeron mukaan
    buildings_to_combine = defaultdict(lambda: {"prt": set()})
    
    # Avaa perusmaksurekisteri
    logger.info("Avataan perusmaksurekisteri...")
    perusmaksut = load_workbook(filename=perusmaksutiedosto)
    
    # Käytä ensimmäistä sheetiä jos "Tietopyyntö asiakasrekisteristä" ei löydy
    sheet_name = "Tietopyyntö asiakasrekisteristä"
    if sheet_name not in perusmaksut.sheetnames:
        sheet_name = perusmaksut.sheetnames[0]
        logger.info(f"Käytetään sheet-nimeä: {sheet_name}")
    
    sheet = perusmaksut[sheet_name]

    # Käsittele rivit ja kerää rakennukset asiakasnumeron mukaan
    logger.info("Käsitellään perusmaksurekisterin rivit...")
    rows_processed = 0
    buildings_found = 0

    for index, row in enumerate(sheet.values):
        if index == 0:  # Ohita otsikkorivi
            continue
        
        rows_processed += 1
        asiakasnumero = str(row[2])
        prt = str(row[0])

        # Hae rakennus
        rakennus = session.query(Rakennus).filter(
            Rakennus.prt == prt
        ).first()

        if not rakennus:
            continue

        # Lisää rakennus asiakasnumeron mukaiseen ryhmään
        buildings_to_combine[asiakasnumero]["prt"].add(prt)
        buildings_found += 1

    logger.info(f"Käsitelty {rows_processed:,} riviä")
    logger.info(f"Löydetty {buildings_found:,} yhdistettävää rakennusta")

    # Muodosta rakennusryhmät samalla asiakasnumerolla olevista
    building_sets = []
    valid_sets = 0

    logger.info("\nMuodostetaan rakennusryhmät...")
    for asiakasnumero, data in buildings_to_combine.items():
        building_set = set()
        
        # Kerää rakennukset ryhmään
        for prt in data["prt"]:
            if prt in dvv_rakennustiedot_by_prt:
                rakennustiedot = dvv_rakennustiedot_by_prt[prt]
                building_set.add(rakennustiedot)

        # Lisää vain jos ryhmässä on rakennuksia
        if building_set:
            building_sets.append(building_set)
            valid_sets += 1

    logger.info(f"Muodostettu {valid_sets:,} rakennusryhmää")

    # Luo omistaja- ja asukastiedot lookupeja varten
    owners_by_rakennus_id = defaultdict(set)
    rakennus_owners = session.execute(
        select(RakennuksenOmistajat.rakennus_id, Osapuoli).join(
            Osapuoli, RakennuksenOmistajat.osapuoli_id == Osapuoli.id
        )
    ).all()
    
    owners_by_rakennus_id = defaultdict(set)  # rakennus_id -> {omistajat}
    rakennus_ids_by_owner_id = defaultdict(set)  # omistaja.id -> {rakennus_id:t}
    for (rakennus_id, owner) in rakennus_owners:
        owners_by_rakennus_id[rakennus_id].add(owner)
        rakennus_ids_by_owner_id[owner.id].add(rakennus_id)

    inhabitants_by_rakennus_id = defaultdict(set)
    rakennus_inhabitants = session.execute(
        select(RakennuksenVanhimmat.rakennus_id, Osapuoli).join(
            Osapuoli, RakennuksenVanhimmat.osapuoli_id == Osapuoli.id
        )
    ).all()
    
    inhabitants_by_rakennus_id = defaultdict(set)  # rakennus_id -> {asukkaat}
    for (rakennus_id, inhabitant) in rakennus_inhabitants:
        inhabitants_by_rakennus_id[rakennus_id].add(inhabitant)

    # Luo kohteet
    logger.info("\nLuodaan kohteet...")
    kohteet = get_or_create_kohteet_from_rakennustiedot(
        session,
        dvv_rakennustiedot,
        building_sets,
        owners_by_rakennus_id,
        inhabitants_by_rakennus_id,
        poimintapvm,
        loppupvm,
        lukittu=True
    )

    logger.info(f"\nLuotu {len(kohteet):,} kohdetta perusmaksurekisterin perusteella")
    return kohteet


def get_or_create_kohteet_from_rakennustiedot(
    session: Session,
    dvv_rakennustiedot: Dict[int, Rakennustiedot],
    building_sets: List[Set[Rakennustiedot]],
    owners_by_rakennus_id: Dict[int, Set[Osapuoli]],
    inhabitants_by_rakennus_id: Dict[int, Set[Osapuoli]],
    poimintapvm: Optional[datetime.date],
    loppupvm: Optional[datetime.date],
    lukittu: bool = False
) -> List[Kohde]:
    """
    Luo kohteet rakennusryhmien perusteella.
    """
    logger = logging.getLogger(__name__)
    logger.info(f"\nLuodaan kohteet {len(building_sets)} rakennusryhmälle...")
    
    kohteet = []

    for i, building_set in enumerate(building_sets, 1):
        # Kerää rakennusten ID:t ja rakennustiedot
        rakennus_ids = []
        rakennustiedot_set = set()
        
        for tiedot in building_set:
            if isinstance(tiedot, tuple):
                rakennus_id = tiedot[0].id
                rakennustiedot_set.add(tiedot[0])
            else:
                rakennus_id = tiedot.id
                rakennustiedot_set.add(tiedot)
            rakennus_ids.append(rakennus_id)
        
        # Kerää omistajat ja asukkaat
        asukkaat = set()
        omistajat = set()
        
        for rakennus_id in rakennus_ids:
            if rakennus_id in owners_by_rakennus_id:
                omistajat.update(owners_by_rakennus_id[rakennus_id])
            if rakennus_id in inhabitants_by_rakennus_id:
                asukkaat.update(inhabitants_by_rakennus_id[rakennus_id])
        
        # Yhdistä osapuolet ja asukkaat
       # osapuolet = omistajat.union(asukkaat)
        
        # Luo tai päivitä kohde
        kohde = update_or_create_kohde_from_buildings(
            session,
            dvv_rakennustiedot,
            rakennustiedot_set,  # Käytetään rakennustiedot_set:iä building_set:in sijaan
            asukkaat,
            omistajat,
            poimintapvm,
            loppupvm,
            lukittu=lukittu
        )
        
        if kohde:
            kohteet.append(kohde)
        
        # Näytä edistyminen joka 100. kohteen jälkeen
        if i % 100 == 0:
            logger.info(f"Käsitelty {i}/{len(building_sets)} rakennusryhmää...")

    logger.info(f"Käsitelty kaikki {len(building_sets)} rakennusryhmää.")
    return kohteet


def _cluster_rakennustiedot(
    rakennustiedot_to_cluster: "Set[Rakennustiedot]",
    distance_limit: int,
    existing_cluster: "Optional[Set[Rakennustiedot]]" = None
) -> "List[Set[Rakennustiedot]]":
    """
    Klusteroi rakennukset seuraavien ehtojen perusteella:
    - Etäisyys toisistaan max 300m JA
    - Samat omistajat/asukkaat JA
    - Sama osoite TAI kiinteistötunnus
    """
    clusters: List[set[Rakennustiedot]] = []
    # Aloita klusteri ensimmäisestä rakennuksesta (tai olemassaolevasta klusterista)
    cluster = existing_cluster.copy() if existing_cluster else None
    
    while rakennustiedot_to_cluster:
        if not cluster:
            first_building = rakennustiedot_to_cluster.pop()
            cluster = set([first_building])
            print(f"\nAloitetaan uusi klusteri: {first_building[0].prt} \nRakennuksia jäljellä: {len(rakennustiedot_to_cluster)}")
            if first_building[0].prt is None:
                print("\nRakennuksen tunnus puuttuu, ohitetaan")
                continue

        other_rakennustiedot_to_cluster = rakennustiedot_to_cluster.copy()
        while other_rakennustiedot_to_cluster:
            found_match = False
            for other_rakennustiedot in other_rakennustiedot_to_cluster:
                if not other_rakennustiedot[0].prt:
                    print("\nRakennuksen tunnus puuttuu, ohitetaan")
                    break

                print(f"\nVerrataan rakennuksia {[r[0].prt for r in cluster]} ja {other_rakennustiedot[0].prt}")
                
                # Tarkista etäisyys kaikkiin klusterin rakennuksiin
                all_buildings = [rakennustiedot[0] for rakennustiedot in cluster] + [other_rakennustiedot[0]]
                max_distance = maximum_distance_of_buildings(all_buildings)
                if max_distance >= distance_limit:
                    print(f"- Etäisyys liian suuri: {max_distance}m >= {distance_limit}m")
                    continue

                # Tarkista omistajat/asukkaat
                match_found = False
                if _match_ownership_or_residents(cluster, other_rakennustiedot):
                    match_found = True
                    print(f"- Omistaja tai asukkaat täsmää")

                if not match_found:
                    print("- Ei yhteisiä omistajia/asukkaita")
                    continue

                # Tarkista osoite TAI kiinteistötunnus
                match_found = False
                for cluster_building in cluster:
                    if _match_addresses(cluster_building[3], other_rakennustiedot[3]):
                        print(f"- Osoite täsmää")
                        match_found = True
                        break
                    if cluster_building[0].kiinteistotunnus == other_rakennustiedot[0].kiinteistotunnus:
                        print(f"- Kiinteistötunnus täsmää: {cluster_building[0].kiinteistotunnus}")
                        match_found = True
                        break
                if not match_found:
                    print("- Ei samaa osoitetta/kiinteistötunnusta")
                    continue

                # Kaikki ehdot täyttyvät, lisää rakennus klusteriin
                print(f"=> Lisätään {other_rakennustiedot[0].prt} klusteriin")
                cluster.add(other_rakennustiedot)
                found_match = True
                break

            if not found_match:
                # Klusteri on valmis! Muita sopivia rakennuksia ei löytynyt.
                break

            # Poista löydetty rakennus käsiteltävistä ja jatka silmukkaa
            other_rakennustiedot_to_cluster.remove(other_rakennustiedot)

        # Klusteri on valmis! Poista klusteroidut rakennukset ja aloita silmukka alusta
        clusters.append(cluster)
        rakennustiedot_to_cluster -= cluster
        cluster = None

    return clusters