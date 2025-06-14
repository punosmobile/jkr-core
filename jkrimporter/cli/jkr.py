import subprocess
import csv
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Optional

import typer
import pandas as pd
from sqlalchemy import text
from sqlalchemy.orm import scoped_session
from sqlalchemy.orm.session import sessionmaker
import openpyxl
from openpyxl.utils import get_column_letter

from jkrimporter import __version__
from jkrimporter.providers.db.dbprovider import DbProvider, engine
from jkrimporter.providers.db.services.tiedontuottaja import (
    get_tiedontuottaja,
    insert_tiedontuottaja,
    list_tiedontuottajat,
    remove_tiedontuottaja,
    rename_tiedontuottaja,
)

from jkrimporter.providers.lahti.lahtiprovider import (
    IlmoitusTranslator,
    LahtiTranslator,
    LopetusIlmoitusTranslator,
    PaatosTranslator,
)
from jkrimporter.providers.lahti.ilmoitustiedosto import (
    Ilmoitustiedosto,
    LopetusIlmoitustiedosto,
)
from jkrimporter.providers.lahti.paatostiedosto import Paatostiedosto
from jkrimporter.providers.lahti.siirtotiedosto import LahtiSiirtotiedosto
from jkrimporter.providers.nokia.nokiaprovider import NokiaTranslator
from jkrimporter.providers.nokia.siirtotiedosto import NokiaSiirtotiedosto
from jkrimporter.providers.pjh.pjhprovider import PjhTranslator
from jkrimporter.providers.pjh.siirtotiedosto import PjhSiirtotiedosto
from jkrimporter.utils.date import parse_date_string


@dataclass
class Provider:
    Translator: type
    Siirtotiedosto: type


PROVIDERS = {
    # we may also add other providers using the *same* formats
    "PJH": Provider(Translator=PjhTranslator, Siirtotiedosto=PjhSiirtotiedosto),
    "HKO": Provider(Translator=NokiaTranslator, Siirtotiedosto=NokiaSiirtotiedosto),
    "LSJ": Provider(Translator=LahtiTranslator, Siirtotiedosto=LahtiSiirtotiedosto)
}


def version_callback(value: bool):
    if value:
        print(__version__)
        raise typer.Exit()


def main_callback(
    version: Optional[bool] = typer.Option(
        None,
        "--version",
        callback=version_callback,
        is_eager=True,
        help="Kertoo lataustyökalun versionumeron.",
    ),
):
    ...


app = typer.Typer(callback=main_callback)

provider_app = typer.Typer()
app.add_typer(
    provider_app, name="tiedontuottaja", help="Muokkaa ja tarkastele tiedontuottajia."
)


@app.command("import", help="Import transportation data to JKR.")
def import_data(
    siirtotiedosto: Path = typer.Argument(..., help="Siirtotiedoston kansio"),
    tiedontuottajatunnus: str = typer.Argument(
        ..., help="Tiedon toimittajan tunnus. Esim. 'PJH', 'HKO', 'LSJ'"
    ),
    luo_uudet: bool = typer.Option(
        False,
        "--luo_uudet",
        help="Luo puuttuvat uudet kohteet tästä datasta.",
    ),
    ala_paivita_yhteystietoja: bool = typer.Option(
        False,
        "--ala_paivita_yhteystietoja",
        help="Älä päivitä yhteystietoja tästä datasta.",
    ),
    ala_paivita_kohdetta:  bool = typer.Option(
        True,
        "--ala_paivita_kohdetta",
        help="Älä päivitä kohteen voimassaoloaikaa tästä datasta.",
    ),
    alkupvm: str = typer.Argument(None, help="Importoitavan datan alkupvm"),
    loppupvm: str = typer.Argument(None, help="Importoitavan datan loppupvm"),
):
    tiedontuottaja = get_tiedontuottaja(tiedontuottajatunnus)
    if not tiedontuottaja:
        typer.echo(
            f"Tiedontuottajaa {tiedontuottaja} ei löydy järjestelmästä. Lisää komennolla `jkr tiedontuottaja add`"
        )
        raise typer.Exit()
    provider = PROVIDERS[tiedontuottajatunnus]
    data = provider.Siirtotiedosto(siirtotiedosto)

    if alkupvm:
        alkupvm = parse_date_string(alkupvm)
    if loppupvm:
        loppupvm = parse_date_string(loppupvm)

    translator = provider.Translator(data, tiedontuottajatunnus)
    jkr_data = translator.as_jkr_data(alkupvm, loppupvm)
    print('writing to db...')
    db = DbProvider()
    db.write(jkr_data, tiedontuottajatunnus, not luo_uudet, ala_paivita_yhteystietoja, ala_paivita_kohdetta, siirtotiedosto)

    print("VALMIS!")


@app.command("create_dvv_kohteet", help="Create kohteet from DVV data in database.")
def create_dvv_kohteet(
    poimintapvm: Optional[str] = typer.Argument(None, help="Importoitavan datan poimintapvm"),
    perusmaksutiedosto: Optional[Path] = typer.Argument(
        None, help="Perusmaksurekisteritiedosto"
    ),
):
    db = DbProvider()
    # Currently, typer does not support Union[datetime, None] argument type, so we will
    # have to parse the datetime string ourselves.
    # support all combinations of known and unknown alku- and loppupvm
    if poimintapvm == "None":
        start = None
    else:
        start = datetime.strptime(poimintapvm, "%d.%m.%Y").date()
    end = None

    db.write_dvv_kohteet(start, end, perusmaksutiedosto)

    print("VALMIS!")


@app.command("import_and_create_kohteet",
             help="Imports dvv data and creates dvv kohteet(optionally with perusmaksu). Optionally imports posti data.")
def import_and_create_kohteet(
    poimintapvm: str = typer.Argument(None, help="Dvv-aineiston poimintapäivämäärä, P.K.V muodossa."),
    dvv: Path = typer.Argument(None, help="Dvv-tiedoston sijainti"),
    perusmaksutiedosto: Optional[Path] = typer.Argument(None, help="Perusmaksurekisteri-tiedoston sijainti."),
    posti: Optional[str] = typer.Argument(None, help="Syötä arvoksi 'posti' jos haluat importoida myös posti datan."),
):
    bat_file = ".\\scripts\\import_and_create_kohteet.bat"

    cmd_args = [bat_file, dvv, poimintapvm]

    if posti == "posti":
        cmd_args.append("posti")

    subprocess.call(cmd_args)

    if perusmaksutiedosto is not None:
        create_dvv_kohteet(poimintapvm, perusmaksutiedosto)
    else:
        create_dvv_kohteet(poimintapvm, None)

    print("VALMIS!")


@app.command("update_huoneistomaara",
             help="Imports and updates number of appartments associated with buildings.")
def update_huoneistomaara(
    huoneisto_xlsx: Path = typer.Argument(None, help="Huoniestolukumäärä-tiedoston sijainti")
):
    bat_file = ".\\scripts\\update_huoneistomaara.bat"

    cmd_args = [bat_file, huoneisto_xlsx]

    subprocess.call(cmd_args)

    print("Huoneistolukumäärät päivitetty!")


@app.command("import_paatokset", help="Import decisions to JKR.")
def import_paatokset(
    siirtotiedosto: Path = typer.Argument(..., help="Polku siirtotiedostoon")
):
    translator = PaatosTranslator(Paatostiedosto(siirtotiedosto))
    paatos_data = translator.as_jkr_data()
    db = DbProvider()
    db.write_paatokset(paatos_data, siirtotiedosto)

    print("VALMIS!")


@app.command("import_ilmoitukset", help="Import compost notices to JKR.")
def import_ilmoitukset(
    siirtotiedosto: Path = typer.Argument(..., help="Kompostointi ilmoitus-tiedoston sijainti.")
):
    translator = IlmoitusTranslator(Ilmoitustiedosto(siirtotiedosto))
    ilmoitus_data = translator.as_jkr_data()
    db = DbProvider()
    db.write_ilmoitukset(ilmoitus_data, siirtotiedosto)

    print("VALMIS!")


@app.command("import_lopetusilmoitukset", help="Import compost ending notices to JKR.")
def import_lopetusilmoitukset(
    siirtotiedosto: Path = typer.Argument(..., help="Kompostoinnin lopetusilmoitus-tiedoston sijainti.")
):
    translator = LopetusIlmoitusTranslator(LopetusIlmoitustiedosto(siirtotiedosto))
    lopetusilmoitus_data = translator.as_jkr_data()
    db = DbProvider()
    db.write_lopetusilmoitukset(lopetusilmoitus_data, siirtotiedosto)

    print("VALMIS!")


@app.command("raportti")
def raportti(
    output_path: Path = typer.Argument(..., help="Excel-raportin tallennuspolku (.xlsx)"),
    tarkastelupvm: str = typer.Argument(None, help="Tarkastelupäivämäärä (YYYY-MM-DD tai DD.MM.YYYY)"),
    kunta: str = typer.Argument(None, help="Kunnan nimi (esim. 'Lahti'). Käytä 0 jos ei rajausta."),
    huoneistomaara: int = typer.Argument(0, help="Huoneistomäärä (4 = neljä tai vähemmän, 5 = viisi tai enemmän, 0 = ei rajausta)"),
    taajama: int = typer.Argument(None, help="Taajama (0 = ei rajausta, 1 = yli 10000, 2 = yli 200)"),
    kohde_tyyppi: int = typer.Argument(None, help="Kohdetyyppi 5 = hapa, 6 = biohapa, 7 = asuinkiinteistö, 8 = muu, 0 = ei rajausta"),
):
    """
    Luo Excel-raportin kohteista annetuilla hakuehdoilla.
    """
    try:
        if(tarkastelupvm != '0'):
            tarkastelupvm_date = parse_date_string(tarkastelupvm)
        else:
            tarkastelupvm_date = None
        
        # Convert "0" to None for kunta
        kunta_filter = None if kunta == "0" else kunta

        # Convert "0" to None for taajama
        taajama_filter = None if taajama == "0" else taajama

        # Convert "0" to None for kohdetyyppi
        kohde_filter = None if kohde_tyyppi == 0 else kohde_tyyppi

        taajama_10000_filter = None if taajama_filter == None else (None if taajama_filter not in (2,3, 10000) else (taajama_filter in (2,3, 10000)))
        taajama_200_filter = None if taajama_filter == None else (None if taajama_filter not in (1,3, 200) else (taajama_filter in (1,3, 200)))
        
        print(f"Haetaan raportille ehdoilla: tarkastelupvm={tarkastelupvm_date}, kunta={kunta_filter}, huoneistomaara={huoneistomaara}, taajama_10000={taajama_10000_filter}, taajama_200={taajama_200_filter}, kohde_tyyppi={kohde_filter}")
        # Create SQLAlchemy session
        Session = scoped_session(sessionmaker(bind=engine))
        with Session() as session:
            # Execute report query
            
            result = session.execute(
                text("SELECT * FROM jkr.print_report(:tarkastelupvm, :kunta, :huoneistomaara, :taajama_10000, :taajama_200, :kohde_tyyppi_id)"),
                {
                    "tarkastelupvm": tarkastelupvm_date,
                    "kunta": kunta_filter,
                    "huoneistomaara": huoneistomaara,
                    "taajama_10000": taajama_10000_filter,  # is_taajama_yli_10000
                    "taajama_200": taajama_200_filter,  # is_taajama_yli_200
                    "kohde_tyyppi_id": kohde_filter
                }
            )

            
            # Get column names
            columns = result.keys()
            
            # Fetch all results
            results = result.fetchall()
            
            # Create DataFrame
            df = pd.DataFrame(results, columns=columns)
            
            # Save to Excel
            df.to_excel(output_path, index=False, engine='openpyxl')
            
            # Open the workbook to adjust column widths
            wb = openpyxl.load_workbook(output_path)
            ws = wb.active
            
            # Adjust column widths based on content
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                # Find the maximum length of content in each column
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # Set width with a small padding
                adjusted_width = max_length + 2
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save the workbook with adjusted column widths
            wb.save(output_path)
            
            typer.echo(f"Raportti luotu onnistuneesti: {output_path}")
            
    except ValueError as e:
        typer.echo(f"Virhe raportin luonnissa: {str(e)}", err=True)
        raise typer.Exit(1)
    except Exception as e:
        typer.echo(f"Virhe raportin luonnissa: {str(e)}", err=True)
        raise typer.Exit(1)


@provider_app.command("add", help="Lisää uusi tiedontuottaja järjestelmään.")
def tiedontuottaja_add_new(
    tunnus: str = typer.Argument(..., help="Tiedontuottajan tunnus. Esim. 'PJH'"),
    name: str = typer.Argument(..., help="Tiedontuottajan nimi."),
):
    insert_tiedontuottaja(tunnus.upper(), name)


@provider_app.command("rename", help="Muuta tiedontuottajan nimi järjestelmässä.")
def tiedontuottaja_rename(
    tunnus: str = typer.Argument(..., help="Tiedontuottajan tunnus. Esim. 'PJH'"),
    name: str = typer.Argument(..., help="Tiedontuottajan uusi nimi."),
):
    rename_tiedontuottaja(tunnus.upper(), name)


@provider_app.command("remove", help="Poista tiedontuottaja järjestelmästä.")
def tiedontuottaja_remove(
    tunnus: str = typer.Argument(..., help="Tiedontuottajan tunnus. Esim. 'PJH'")
):
    remove_tiedontuottaja(tunnus.upper())


@provider_app.command("list", help="Listaa järjestelmästä löytyvät tiedontuottajat.")
def tiedontuottaja_list():
    for tiedontuottaja in list_tiedontuottajat():
        print(f"{tiedontuottaja.tunnus}\t{tiedontuottaja.nimi}")


@app.command("import_hapa", help="Import HAPA data from CSV file to JKR database.")
def import_hapa(
    aineistopolku: Path = typer.Argument(..., help="Path to the HAPA CSV file"),
):
    """Import HAPA data from CSV file to JKR database.
    
    The CSV file should have columns like:
    Rakennus-ID, Kohde id, Sijaintikunta, Asiakasnro, Rakennus-ID, Katunimi FI, 
    Talon numero, Postinumero, Postitoimipaikka FI, kohdetyyppi
    
    The file must be in CSV format with semicolon (;) as delimiter, UTF-8 encoding, and include headers.
    """
    try:
        # Check if file exists
        if not aineistopolku.exists():
            typer.echo(f"Error: File {aineistopolku} does not exist", err=True)
            raise typer.Exit(1)
            
        typer.echo(f"Importing HAPA data from {aineistopolku}")
        
        # Create SQLAlchemy session
        Session = scoped_session(sessionmaker(bind=engine))
        with Session() as session:
            # Read CSV file to verify structure before importing
            with open(aineistopolku, 'r', encoding='utf-8') as f:
                reader = csv.reader(f, delimiter=';')
                headers = next(reader)
                
                # Check if all required headers are present
                required_headers = ['Rakennus-ID', 'Kohde id', 'Sijaintikunta', 'Asiakasnro', 
                                   'Katunimi FI', 'Talon numero', 'Postinumero', 
                                   'Postitoimipaikka FI', 'kohdetyyppi']
                
                missing_headers = [h for h in required_headers if h not in headers]
                if missing_headers:
                    typer.echo(f"Error: CSV file is missing required headers: {missing_headers}", err=True)
                    raise typer.Exit(1)
            
            # Create a temporary file with renamed headers to match database columns
            import tempfile
            with tempfile.NamedTemporaryFile(mode='w', encoding='utf-8', delete=False, suffix='.csv') as temp_file:
                temp_path = Path(temp_file.name)
                
                # Write header row with database column names
                db_headers = ['rakennus_id_tunnus', 'kohde_tunnus', 'sijaintikunta', 'asiakasnro', 
                             'rakennus_id_tunnus2', 'katunimi_fi', 'talon_numero', 'postinumero', 
                             'postitoimipaikka_fi', 'kohdetyyppi']
                temp_file.write(';'.join(db_headers) + '\n')
                
                # Read original CSV and write to temp file with correct column order
                with open(aineistopolku, 'r', encoding='utf-8') as f:
                    reader = csv.reader(f, delimiter=';')
                    next(reader)  # Skip header
                    
                    for row in reader:
                        if len(row) >= 10:  # Ensure row has enough columns
                            temp_file.write(';'.join(row) + '\n')
            
            # Use SQL COPY command for efficient bulk loading with copy_expert
            copy_sql = f"""
            COPY jkr.hapa_aineisto(
                rakennus_id_tunnus, kohde_tunnus, sijaintikunta, asiakasnro, 
                rakennus_id_tunnus2, katunimi_fi, talon_numero, postinumero, 
                postitoimipaikka_fi, kohdetyyppi
            ) FROM STDIN WITH (
                FORMAT csv, 
                DELIMITER ';', 
                HEADER true, 
                ENCODING 'UTF8', 
                NULL ''
            );
            """
            
            # Execute the COPY command with the temporary file using copy_expert
            connection = session.connection().connection
            cursor = connection.cursor()
            
            # Open the file and use copy_expert
            with open(temp_path, 'r', encoding='utf-8') as f:
                cursor.copy_expert(copy_sql, f)
                
            connection.commit()
            
            # Clean up the temporary file
            try:
                temp_path.unlink()
            except Exception:
                pass
            
            # Get count of imported rows
            result = session.execute(text("SELECT COUNT(*) FROM jkr.hapa_aineisto WHERE tuonti_pvm >= CURRENT_DATE"))
            count = result.scalar()
            
            typer.echo(f"Successfully imported {count} HAPA records")
            typer.echo("VALMIS!")
            
    except Exception as e:
        typer.echo(f"Error importing HAPA data: {str(e)}", err=True)
        raise typer.Exit(1)


if __name__ == "__main__":
    app()
