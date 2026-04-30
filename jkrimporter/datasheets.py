import csv
from abc import ABC, ABCMeta, abstractmethod
from pathlib import Path
from typing import TYPE_CHECKING, Any, Dict, Generic, Iterator, List, Set, TypeVar

from openpyxl.reader.excel import load_workbook
from pydantic import ValidationError

if TYPE_CHECKING:
    from typing import Iterable, Protocol

    from openpyxl.workbook.workbook import Workbook
    from openpyxl.worksheet.worksheet import Worksheet

    class Sheet(Iterable, Protocol):
        sheet: Any
        headers: List[str]



def get_hapa_kohteet_headers():
    expected_headers = [
        "Rakennus-ID", 
        "Kohde id", 
        "Sijaintikunta", 
        "Asiakasnro",  
        "Katunimi FI", 
        "Talon numero", 
        "Postinumero", 
        "Postitoimipaikka fi", 
        "kohdetyyppi"
    ]
    return expected_headers


def get_sote_kohteet_headers():
    expected_headers = [
        "Rakennus-ID", 
        "Kohde id", 
        "Sijaintikunta", 
        "Asiakasnro",  
        "Katunimi FI", 
        "Talon numero", 
        "Postinumero", 
        "Postitoimipaikka fi", 
        "kohdetyyppi"
    ]
    return expected_headers


def get_tiedontuottajat_headers():
    expected_headers = [
        "tunnus",
        "nimi"
    ]
    return expected_headers

def get_dvv_osoite_headers():
    expected_headers = [
        "Kadunnimi ruotsiksi",
        "Kadunnimi suomeksi",
        "Sijainti-kunta",
        "Posti-numero",
        "Katu-numero",
        "Rakennustunnus",
        "Postitoimipaikan nimi suomeksi"
    ]
    return expected_headers


def get_dvv_rakennus_headers():
    expected_headers = [
        "Rakennustunnus",
        "Sijaintikiinteistön tunnus",
        "Viemäri",
        "Pohjois-koordinaatti",
        "Itä-koordinaatti",
        """Valmis-tumis-
päivä""",
       "Käytössä-olotilanteen muutospäivä",
        "Käyttö-tarkoitus",
        "Käytös-säolo-tilanne",
        """Rakennus-
luokka""",
        "Sijainti-kunta"
    ]
    return expected_headers


def get_dvv_omistaja_headers():
    expected_headers = [
        "Omistajan nimi",
        "Omistajan vakinainen kotimainen asuinosoite",
        "Vakinaisen kotim osoitteen postitoimipaikka",
        "Vak os posti- numero",
        "Omistajan kuolinpäivä",
        "Vakin kotim osoitteen alkupäivä",
        "Postios posti-numero",
        "Postiosoitteen postitoimipaikka",
        "Omistajan postiosoite",
        "Um os valtio-koodi",
        "Omistajan ulkomainen lähiosoite",
        "Ulkomaisen osoitteen paikkakunta",
        "Ulkomaisen osoitteen valtion postinimi",
        "Omist koti-kunta",
        "Henkilötunnus",
        "Y-tunnus",
        "Omistuksen alkupäivä"
    ]
    return expected_headers


def get_dvv_asukas_headers():
    expected_headers = [
        "Huoneiston vanhin asukas (henkilötunnus)",
        "Sukunimi",
        "Etunimet",
        "Vakinainen kotimainen asuinosoite",
        "Vak os posti- numero",
        "Sijainti-kunta",
        "Rakennustunnus",
        "Vakin kotim osoitteen alkupäivä",
        "Vakinaisen kotim osoitteen postitoimipaikka",
        "Huo-neisto-kirjain",
        "Huo-neisto-numero",
        "Jako-kirjain",
    ]
    return expected_headers


def get_perusmaksu_headers():
    expected_headers = [
        "Rakennus-ID",
        "Sijaintikunta",
        "Asiakasnro",
        "Katunimi FI",
        "Talon numero",
        "Postinumero",
        "Postitoimipaikka FI",
    ]
    return expected_headers


def get_siirtotiedosto_headers():
    expected_headers = [
        "UrakoitsijaId",
        "UrakoitsijankohdeId",
        "Kiinteistotunnus",
        "Kiinteistonkatuosoite",
        "Kiinteistonposti",
        "Haltijannimi",
        "Haltijanyhteyshlo",
        "Haltijankatuosoite",
        "Haltijanposti",
        "Haltijanmaakoodi",
        "Haltijanulkomaanpaikkakunta",
        "Pvmalk",
        "Pvmasti",
        "tyyppiIdEWC",
        "COUNT(kaynnit)",
        "SUM(astiamaara)",
        "koko",
        "SUM(paino)",
        "tyhjennysvali",
        "kertaaviikossa",
        "Voimassaoloviikotalkaen",
        "Voimassaoloviikotasti",
        "Voimassaoloviikotalkaen2",
        "Voimassaoloviikotasti2",
        "tyhjennysvali2",
        "kertaaviikossa2",
        "palveluKimppakohdeId",
        "KimpanNimi",
        "Kimpanyhteyshlo",
        "Kimpankatuosoite",
        "Kimpanposti",
        "Kuntatun",
        "Keskeytysalkaen",
        "Keskeytysasti",
    ]
    return expected_headers


def get_paatostiedosto_headers():
    return [
        "Numero",
        "Lähettäjä/vastaanottaja",
        "PRT 1",
        "Päätös 1",
        "Voimassa alkaen 1",
        "Voimassa asti 1",
        "Lisätiedot 1",
        "Lähiosoite",
        "Postinumero",
        "Postitoimipaikka",
    ]

def get_viemari_ilmoitustiedosto_headers():
    return [
        "Viemäriverkosto alkupvm",
        "PRT",
    ]

def get_viemari_lopetustiedosto_headers():
    return [
        "Viemäriverkosto loppupvm",
        "PRT",
    ]

def get_ilmoitustiedosto_headers():
    return [
        "Vastausaika",
        "Kompostoinnin vastuuhenkilön yhteystiedot:Etunimi",
        "Kompostoinnin vastuuhenkilön yhteystiedot:Sukunimi",
        "Kompostoinnin vastuuhenkilön yhteystiedot:Postiosoite",
        "Kompostoinnin vastuuhenkilön yhteystiedot:Postinumero",
        "Kompostoinnin vastuuhenkilön yhteystiedot:Postitoimipaikka",
        "Rakennuksen tiedot, jossa kompostori sijaitsee:Rakennuksen katuosoite",
        "Rakennuksen tiedot, jossa kompostori sijaitsee:Käsittelijän lisäämä tunniste",
        "Kompostoria käyttävien asuinhuoneistojen lukumäärä",
        "1. Kompostoria käyttävän rakennuksen tiedot:Haltijan etunimi",
        "1. Kompostoria käyttävän rakennuksen tiedot:Haltijan sukunimi",
        "1. Kompostoria käyttävän rakennuksen tiedot:Viranomaisen lisäämä tarkenne",
        "1. Kompostoria käyttävän rakennuksen tiedot:Käsittelijän lisäämä tunniste",
        "Voimassaolopäivä",
    ]


def get_liete_kuljetustiedosto_headers():
    return [
        "ID-tunnus",
        "Siirron alkamisaika",
        "Jätteen tuottaja tai muu haltija",
        "Jätteen tuottajan/haltijan osoite",
        "Jätteen tuottajan/haltijan katuosoite",
        "Jätteen tuottajan/haltijan postinumero",
        "Siirron alkamispaikka",
        "Siirron alkamispaikan katuosoite",
        "Siirron alkamispaikan postinumero",
        "Kuljettaja",
        "Vastaanottaja",
        "Siirron päättymispaikka",
        "Siirron päättymispaikan katuosoite",
        "Siirron päättymispaikan postinumero",
        "Siirron päättymisaika",
        "Jäte",
        "Jätteen kuvaus",
        "Jätteen paino (t)",
        "Jätteen tilavuus (m³)",
        "Kiinteistötunnus",
        "Pysyvä rakennustunnus",
    ]


def get_huoneistomaara_headers():
    return [
        "C_VTJ_PRT",
        "I_HUONEISTOJEN_LKM",
        "C_KAYTTARK"
    ]


def get_kaivotiedosto_headers():
    return [
        "Vastausaika",
        "PRT",
        "Etunimi",
        "Sukunimi",
        "Katuosoite",
        "Postinumero",
        "Postitoimipaikka",
        "Kantovesi",
        "Saostussäiliö",
        "Pienpuhdistamo",
        "Umpisäiliö",
        "Vain harmaat vedet",
        "Tietolähde",
    ]


def get_liete_ilmoitustiedosto_headers():
    return [
        "Vastausaika",
        "Lietteen kompostoijan yhteystiedot:Etunimi",
        "Lietteen kompostoijan yhteystiedot:Sukunimi",
        "Lietteen kompostoijan yhteystiedot:Postiosoite",
        "Lietteen kompostoijan yhteystiedot:Postinumero",
        "Tiedot kiinteistöstä, jonka liete kompostoidaan:Käsittelijän lisäämä tunniste",
        ":Voimassa alkaen",
        ":Voimassa asti",
    ]


def get_lopetustiedosto_headers():
    return [
        "Vastausaika",
        "Kompostoinnin vastuuhenkilö:Etunimi",
        "Kompostoinnin vastuuhenkilö:Sukunimi",
        "Rakennuksen tiedot:Käsittelijän lisäämä tunniste",
    ]


class CsvSheet:
    def __init__(self, file_path):
        self._file_path = file_path
        with open(self._file_path, mode="r", encoding="cp1252", newline="") as csv_file:
            csv_reader = csv.reader(csv_file, delimiter=";", quotechar='"')
            self.headers = next(csv_reader)

    def __iter__(self):
        with open(self._file_path, mode="r", encoding="cp1252", newline="") as csv_file:
            csv_reader = csv.DictReader(csv_file, delimiter=";", quotechar='"')
            for row in csv_reader:
                yield row


class ExcelSheet:
    def __init__(self, sheet: "Worksheet"):
        self.sheet = sheet
        self.headers = next(
            self.sheet.iter_rows(min_row=1, max_row=1, values_only=True)
        )

    def __iter__(self) -> Iterator[Dict[str, Any]]:
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            yield dict(zip(self.headers, row))


class CombinedExcelSheet:
    def __init__(self, workbooks: Set["Workbook"], key: str = None):
        if key:
            self.sheets = [ExcelSheet(workbook[key]) for workbook in workbooks]
        else:
            self.sheets = [ExcelSheet(workbook.active) for workbook in workbooks]
        self.headers = self.sheets[0].headers

    def __iter__(self) -> Iterator[Dict[str, Any]]:
        for sheet in self.sheets:
            rows = iter(sheet)
            yield from rows


class SheetCollection(ABC):
    def __init__(self, path):
        self._path = Path(path)
        self._opened_error_files = set()

    def __del__(self):
        for error_sheet in self._opened_error_files:
            error_sheet.close()

    @abstractmethod
    def _open_sheet(self, name) -> "Sheet":
        raise NotImplementedError

    def _open_error_sheet(self, name: str, headers: List[str]):
        log_path = self._path if self._path.is_dir() else self._path.parent
        error_file = open(
            log_path / f"{name}_virheet.csv", "w", newline="", encoding="utf-8"
        )
        writer = csv.DictWriter(error_file, fieldnames=list(headers) + ["virhe"])
        writer.writeheader()

        self._opened_error_files.add(error_file)

        return writer


class CsvSheetCollection(SheetCollection):
    def _open_sheet(self, key):
        return CsvSheet(self._path / f"{key}.csv")


class ExcelFileSheetCollection(SheetCollection):
    """
    Allows reading different worksheets from different files.
    """

    def __init__(self, path):
        super().__init__(path)
        self._opened_workbooks: Set["Workbook"] = set()

    def _open_sheet(self, key):
        workbook = load_workbook(
            filename=self._path / f"{key}.xlsx", data_only=True, read_only=True
        )
        self._opened_workbooks.add(workbook)

        return ExcelSheet(workbook.active)

    def __del__(self):
        for workbook in self._opened_workbooks:
            workbook.close()

        super().__del__()


class ExcelCombinedFileSheetCollection(ExcelFileSheetCollection):
    """
    Allows iterating through a list of xlsx files. Combine active worksheets from all
    files to the same object.
    """

    def __init__(self, path):
        super().__init__(path)
        for f in self._path.iterdir():
            # do not read any extra files, such as any error csvs present
            if f.suffix == ".xlsx":
                workbook = load_workbook(filename=f, data_only=True, read_only=True)
                self._opened_workbooks.add(workbook)

    def _open_sheet(self, key: str = None):
        return CombinedExcelSheet(self._opened_workbooks, key)


class ExcelSheetCollection(SheetCollection):
    def __init__(self, path):
        super().__init__(path)
        self._workbook = load_workbook(
            filename=self._path, data_only=True, read_only=True
        )

    def __del__(self):
        self._workbook.close()

    def _open_sheet(self, key):
        return ExcelSheet(self._workbook[key])


T = TypeVar("T")


class SiirtotiedostoSheet(Generic[T], metaclass=ABCMeta):
    def __init__(self, sheet_collection: SheetCollection, sheet_name: str = None):
        self._sheet = sheet_collection._open_sheet(sheet_name)
        self._error_sheet = sheet_collection._open_error_sheet(
            sheet_name, self._sheet.headers
        )
        self.headers = self._sheet.headers

    @abstractmethod
    def _obj_from_dict(data) -> T:
        raise NotImplementedError

    def __iter__(self) -> Iterator[T]:
        for row in self._sheet:
            try:
                obj = self._obj_from_dict(row)
            except ValidationError as e:
                error = "; ".join(
                    f"{''.join(error['loc'])}: {error['msg']}" for error in e.errors()
                )
                row_with_error = {**row, "virhe": error}
                self._error_sheet.writerow(row_with_error)
                continue

            yield obj
