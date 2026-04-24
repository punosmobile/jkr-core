import os
import asyncio
from pathlib import Path
from typing import Dict, List

import openpyxl

from jkrimporter.conf import (
    get_kohdentumattomat_ilmoitus_filename,
    get_kohdentumattomat_lieteilmoitus_filename,
    get_kohdentumattomat_lopetusilmoitus_filename,
)
from jkrimporter.datasheets import (
    get_ilmoitustiedosto_headers,
    get_liete_ilmoitustiedosto_headers,
    get_lopetustiedosto_headers,
)
from jkrimporter.api import sharepoint as sp


def export_kohdentumattomat_ilmoitukset(
        folder: Path,
        kohdentumattomat: List[Dict[str, str]]
):
    expected_headers = get_ilmoitustiedosto_headers()

    output_file_path_failed = folder / get_kohdentumattomat_ilmoitus_filename()

    if output_file_path_failed.exists():
        workbook_failed = openpyxl.load_workbook(output_file_path_failed)
        sheet_failed = workbook_failed[workbook_failed.sheetnames[0]]
    else:
        workbook_failed = openpyxl.Workbook()
        sheet_failed = workbook_failed[workbook_failed.sheetnames[0]]
        sheet_failed.append(expected_headers)

    filtered_kohdentumattomat = []

    for data in kohdentumattomat:
        if isinstance(data, dict):
            # If the data is a dictionary, append it directly
            filtered_data = {
                key: value for key, value in data.items() if key in expected_headers
            }
            filtered_kohdentumattomat.append(filtered_data)
        elif isinstance(data, list):
            # If the data is a list containing a single dictionary,
            # extract the dictionary and append it
            if len(data) == 1 and isinstance(data[0], dict):
                filtered_data = {
                    key: value for key,
                    value in data[0].items()
                    if key in expected_headers
                }
                filtered_kohdentumattomat.append(filtered_data)
            else:
                print("Unexpected nested structure:", data)
        else:
            print("Unsupported data type:", type(data))

    for row in filtered_kohdentumattomat:
        sheet_failed.append([row.get(header, "") for header in expected_headers])

    workbook_failed.save(output_file_path_failed)

    file_content = output_file_path_failed.read_bytes()
    asyncio.run(sp.upload_file(file_content=file_content, filename=output_file_path_failed.name, user_name="jkr-core"))

def export_kohdentumattomat_lieteIlmoitukset(
        folder: Path,
        kohdentumattomat: List[Dict[str, str]]
):
    expected_headers = get_liete_ilmoitustiedosto_headers()

    output_file_path_failed = folder / get_kohdentumattomat_lieteilmoitus_filename()

    if output_file_path_failed.exists():
        workbook_failed = openpyxl.load_workbook(output_file_path_failed)
        sheet_failed = workbook_failed[workbook_failed.sheetnames[0]]
    else:
        workbook_failed = openpyxl.Workbook()
        sheet_failed = workbook_failed[workbook_failed.sheetnames[0]]
        sheet_failed.append(expected_headers)

    filtered_kohdentumattomat = [
        {key: value for key, value in data.items() if key in expected_headers}
        for data in kohdentumattomat
    ]
    for row in filtered_kohdentumattomat:
        sheet_failed.append([row.get(header, "") for header in expected_headers])

    workbook_failed.save(output_file_path_failed)

    file_content = output_file_path_failed.read_bytes()
    asyncio.run(sp.upload_file(file_content=file_content, filename=output_file_path_failed.name, user_name="jkr-core"))


def export_kohdentumattomat_lopetusilmoitukset(
        folder: Path,
        kohdentumattomat: List[Dict[str, str]]
):
    expected_headers = get_lopetustiedosto_headers()

    output_file_path_failed = folder / get_kohdentumattomat_lopetusilmoitus_filename()
    

    if output_file_path_failed.exists():
        workbook_failed = openpyxl.load_workbook(output_file_path_failed)
        sheet_failed = workbook_failed[workbook_failed.sheetnames[0]]
    else:
        workbook_failed = openpyxl.Workbook()
        sheet_failed = workbook_failed[workbook_failed.sheetnames[0]]
        sheet_failed.append(expected_headers)

    filtered_kohdentumattomat = [
        {key: value for key, value in data.items() if key in expected_headers}
        for data in kohdentumattomat
    ]
    for row in filtered_kohdentumattomat:
        sheet_failed.append([row.get(header, "") for header in expected_headers])

    workbook_failed.save(output_file_path_failed)

    file_content = output_file_path_failed.read_bytes()
    asyncio.run(sp.upload_file(file_content=file_content, filename=output_file_path_failed.name, user_name="jkr-core"))
