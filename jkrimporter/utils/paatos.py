import os
from pathlib import Path
from typing import Dict, List

import openpyxl

from jkrimporter.conf import get_kohdentumattomat_paatos_filename
from jkrimporter.datasheets import get_paatostiedosto_headers


def export_kohdentumattomat_paatokset(folder: Path, kohdentumattomat: List[Dict[str, str]]):
    expected_headers = get_paatostiedosto_headers()

    output_file_path_failed = os.path.join(
        folder, get_kohdentumattomat_paatos_filename()
    )

    if os.path.exists(output_file_path_failed):
        workbook_failed = openpyxl.load_workbook(output_file_path_failed)
        sheet_failed = workbook_failed[workbook_failed.sheetnames[0]]
    else:
        workbook_failed = openpyxl.Workbook()
        sheet_failed = workbook_failed[workbook_failed.sheetnames[0]]
        sheet_failed.append(expected_headers)

    filtered_kohdentumattomat = [
        {key: value for key, value in data.items() if key in expected_headers}
        for data in kohdentumattomat
    ]
    for row in filtered_kohdentumattomat:
        sheet_failed.append([row.get(header, "") for header in expected_headers])

    workbook_failed.save(output_file_path_failed)
