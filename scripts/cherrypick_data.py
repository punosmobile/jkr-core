#!/usr/bin/env python3
import argparse
from pathlib import Path
import sys
import json
import shutil

#   python cherrypick_data.py --rip_path ../data/ --must_contain lahti heinola 171*
#   python cherrypick_data.py --rip_path ../data/ --must_contain 17100 17200

#   Args:
#      rip_path: Path to file or directory to process
#      must_contain: Lista sanoista joita pitää olla vähintään yksi säilytettävällä rivillä.
#                    Villikortit kirjoitetaan 156* ja niitä verrataan sanojen alkuun kullakin rivillä
#      filter_fields: JSON-lista filtteröitävistä kentistä * loppuisille hakusanoille. Voi olla joko suora JSON-merkkijono tai @-alkuinen tiedostopolku'

def clean_old_ripped_files(path: Path) -> None:
    """
    Recursively remove all files with '_ripped' in their name from the given path
    and its subdirectories.
    
    Args:
        path: Path to clean
    """
    try:
        if path.is_file() and '_ripped' in path.name:
            path.unlink()
            print(f"Removed old ripped file: {path}")
        elif path.is_dir():
            for item in path.rglob('*_ripped*'):
                if item.is_file():
                    item.unlink()
                    print(f"Removed old ripped file: {item}")
    except Exception as e:
        print(f"Error while cleaning old ripped files: {str(e)}", file=sys.stderr)

def copy_excel_files(path: Path) -> None:
    """
    Recursively copy all Excel files with _ripped suffix.
    
    Args:
        path: Path to process
    """
    try:
        if path.is_file():
            if path.suffix.lower() in ['.xlsx', '.xls'] and '_ripped' not in path.name:
                ripped_path = path.parent / f"{path.stem}_ripped{path.suffix}"
                shutil.copy2(path, ripped_path)
                print(f"Created ripped copy: {ripped_path}")
        elif path.is_dir():
            for excel_file in path.rglob('*.[xX][lL][sS][xX]'):
                if '_ripped' not in excel_file.name:
                    ripped_path = excel_file.parent / f"{excel_file.stem}_ripped{excel_file.suffix}"
                    shutil.copy2(excel_file, ripped_path)
                    print(f"Created ripped copy: {ripped_path}")
            for excel_file in path.rglob('*.[xX][lL][sS]'):
                if '_ripped' not in excel_file.name:
                    ripped_path = excel_file.parent / f"{excel_file.stem}_ripped{excel_file.suffix}"
                    shutil.copy2(excel_file, ripped_path)
                    print(f"Created ripped copy: {ripped_path}")
    except Exception as e:
        print(f"Error while copying Excel files: {str(e)}", file=sys.stderr)

def process_csv(file_path: Path, must_contain: set[str], must_contain_wild: set[str], filter_fields: set[str]) -> None:
    """
    Process a CSV file, keeping only rows that contain at least one
    of the must_contain words or have a word that starts with given wildcards (case insensitive). Preserves original lines exactly.
    Skip files that don't contain any of the words.
    
    Args:
        file_path: Path to the CSV file
        must_contain: Set of words that a row must contain at least one of (any column)
        must_contain_wild: Prefix strings to match only in specified filter_fields
        filter_fields: Column headers to apply prefix-matching to
    """
    try:
        # Try different encodings in order
        encodings = ['utf-8', 'iso-8859-1', 'cp1252', 'latin1']
        content = None
        used_encoding = None
        
        for encoding in encodings:
            try:
                with open(file_path, 'r', encoding=encoding) as f:
                    content = f.readlines()
                used_encoding = encoding
                break
            except UnicodeDecodeError:
                continue
        
        if content is None:
            raise Exception("Failed to read file with any encoding")

        # Keep header and matching lines
        kept_lines = [content[0]]  # Always keep header
        
        # Store the header row
        header_row = kept_lines[0].strip().split(';')

        # Get header row indexes
        col_index_to_name = {i: name for i, name in enumerate(header_row)}
        wildcard_column_indexes = [i for i, name in col_index_to_name.items() if name in filter_fields]

        if not wildcard_column_indexes:
            print(f'\nEi löytynyt villikorttiotsikoita \nfilter: {filter_fields} \nfile: {file_path}')
        
        for line in content[1:]:  # Skip header when checking
            columns = line.strip().split(';')  # Adjust delimiter if needed
    
            # 1. Match any exact word in any column
            if any(word in val.lower() for val in columns for word in must_contain):
                kept_lines.append(line)
                continue  # Already matched

            # 2. Match wildcard prefixes only in selected columns
            if any(
                not columns[i] is None and columns[i].startswith(prefix)
                for i in wildcard_column_indexes if i < len(columns)
                for prefix in must_contain_wild
            ):
                kept_lines.append(line)

        if len(kept_lines) == 1:
            print(f"Skipping {file_path} - no matching words found")
            return
        
        # Create output filename
        output_path = file_path.parent / f"{file_path.stem}_ripped{file_path.suffix}"
        
        # Write filtered content with same encoding
        with open(output_path, 'w', encoding=used_encoding) as f:
            f.writelines(kept_lines)
            
        print(f"Processed {file_path} using {used_encoding} encoding")
        print(f"Read {len(content)} rows, kept {len(kept_lines)} rows")
        
    except Exception as e:
        print(f"Error processing CSV {file_path}: {str(e)}", file=sys.stderr)

def process_excel(file_path: Path, must_contain: set[str], must_contain_wild: set[str], filter_fields: set[str]) -> None:
    """
    Process an Excel file by copying matching rows to the end and then removing original rows.
    
    Args:
        file_path: Path to the Excel file
        must_contain: Set of words that a row must contain at least one of
    """
    try:
        from openpyxl import load_workbook
        
        print(f"\nAlkaa prosessoida tiedostoa: {file_path}")
        
        # Determine source and target files
        is_ripped = '_ripped' in file_path.name
        source_file = file_path
        if not is_ripped:
            output_path = file_path.parent / f"{file_path.stem}_ripped{file_path.suffix}"
        else:
            output_path = file_path
            
        print("Ladataan työkirjaa...")
        wb = load_workbook(filename=source_file, data_only=False)
        print("Työkirja ladattu!")
        
        file_has_matches = False
        
        # Process each sheet
        for sheet_name in wb.sheetnames:
            print(f"\nKäsitellään välilehteä: {sheet_name}")
            ws = wb[sheet_name]
            
            max_row = ws.max_row
            print(f"Välilehdellä {max_row} riviä")
            
            # Store the header row
            header_row = []
            for cell in ws[1]:
                header_row.append(cell.value)

            # Get header row indexes
            col_index_to_name = {i: name for i, name in enumerate(header_row)}
            wildcard_column_indexes = [i for i, name in col_index_to_name.items() if name in filter_fields]

            if not wildcard_column_indexes:
                print(f'\nEi löytynyt villikorttiotsikoita \nfile: {source_file} sheet: {sheet_name}')

            # Find matching rows and copy them to a list
            matching_rows = []
            current_row = 0
            
            print("Etsitään säilytettäviä rivejä...")
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                
                current_row += 1
                if current_row % 1000 == 0:
                    print(f"Käsitelty {current_row}/{max_row} riviä...")
                
                row_word_list = []
                for cell in row:
                    row_word_list.append(cell.value)

                if any(word in row_word_list for word in must_contain):
                    matching_rows.append(row)
                    file_has_matches = True
                    continue
                    
                # 2. Match wildcard prefixes only in selected columns
                if any(
                    not row_word_list[i] is None and row_word_list[i].startswith(prefix)
                    for i in wildcard_column_indexes if i < len(row_word_list)
                    for prefix in must_contain_wild
                ):
                    matching_rows.append(row)
                    file_has_matches = True
                    continue
            
            if matching_rows:
                print(f"Löydettiin {len(matching_rows)} säilytettävää riviä")
                
                # Delete all rows except header
                if max_row > 1:
                    print("Poistetaan kaikki rivit paitsi otsikko...")
                    ws.delete_rows(2, max_row - 1)
                
                # Append matching rows
                print("Lisätään säilytetyt rivit takaisin...")
                for row in matching_rows:
                    new_row = []
                    for cell in row:
                        new_row.append(cell.value)
                    ws.append(new_row)
                
                print(f"Valmis! Lopullinen rivimäärä: {ws.max_row}")
            else:
                print("Ei löytynyt säilytettäviä rivejä tältä välilehdeltä")
        
        if not file_has_matches:
            print(f"Ohitetaan {source_file} - ei löytynyt haettuja sanoja")
            return
        
        print("\nTallennetaan työkirjaa...")
        wb.save(output_path)
        print("Tallennus valmis!")
        print(f"Käsitelty {source_file} -> {output_path}")
        print(f"Käsitellyt välilehdet: {', '.join(wb.sheetnames)}")
        
    except Exception as e:
        print(f"Virhe Excel-tiedoston {file_path} käsittelyssä: {str(e)}", file=sys.stderr)

def process_file(file_path: Path, must_contain: set[str], must_contain_wild: set[str], filter_fields: set[str]) -> None:
    """
    Process a single file based on its extension.
    
    Args:
        file_path: Path to the file to process
        must_contain: Set of words that a row/line must contain at least one of
    """
    # Skip files with '_ripped' in their name
    if '_ripped' in file_path.name:
        print('skipping already processed file')
        return
        
    suffix = file_path.suffix.lower()
    
    if suffix == '.csv':
        process_csv(file_path, must_contain, must_contain_wild, filter_fields)
    elif suffix in ['.xlsx', '.xls']:
        process_excel(file_path, must_contain, must_contain_wild, filter_fields)
    else:
        print(f"Skipping unsupported file type: {file_path}", file=sys.stderr)

def process_directory(dir_path: Path, must_contain: set[str], must_contain_wild: set[str], filter_fields: set[str]) -> None:
    """
    Recursively process all supported files in directory and its subdirectories.
    
    Args:
        dir_path: Path to directory to process
        must_contain: Set of words that a row/line must contain at least one of
    """
    try:
        # Recursively iterate through all files in directory and subdirectories
        for file_path in dir_path.rglob('*'):
            if file_path.is_file():
                process_file(file_path, must_contain, must_contain_wild, filter_fields)
    except Exception as e:
        print(f"Error processing directory {dir_path}: {str(e)}", file=sys.stderr)

def replace_with_ripped_files(path: Path) -> None:
    """
    Recursively replace original files with their _ripped versions.
    
    Args:
        path: Path to process
    """
    try:
        if path.is_file():
            if '_ripped.' in path.name:
                original_path = path.parent / path.name.replace('_ripped.', '.')
                if original_path.exists():
                    shutil.copy2(path, original_path)
                    path.unlink()
                    print(f"Replaced original file with ripped version: {original_path}")
        elif path.is_dir():
            for ripped_file in path.rglob('*_ripped.*'):
                original_path = ripped_file.parent / ripped_file.name.replace('_ripped.', '.')
                if original_path.exists():
                    shutil.copy2(ripped_file, original_path)
                    ripped_file.unlink()
                    print(f"Replaced original file with ripped version: {original_path}")
    except Exception as e:
        print(f"Error while replacing files: {str(e)}", file=sys.stderr)

def main():
    # Set up argument parser
    parser = argparse.ArgumentParser(
        description='Remove rows from CSV/Excel files that do not contain specified words'
    )
    parser.add_argument(
        '--rip_path',
        type=str,
        required=True,
        help='Path to directory or file to process'
    )
    parser.add_argument(
        '--must_contain',
        type=str,
        nargs='+',
        required=True,
        help='List of words - rows not containing any of these words will be removed'
    )
    parser.add_argument(
        '--filter_fields',
        required=False,
        default='@cherrypickfields.json',
        help='JSON-muotoinen lista filtteröitävistä kentistä * loppuisille hakusanoille. Voi olla joko suora JSON-merkkijono tai @-alkuinen tiedostopolku'
    )

    # Parse arguments
    args = parser.parse_args()

    if args.filter_fields.startswith('@'):
        # Lue kentät JSON-tiedostosta
        with open(args.filter_fields[1:], 'r', encoding='utf-8') as f:
            filter_fields = json.load(f)
    else:
        # Parsitaan suoraan JSON-merkkijonosta
        filter_fields: set[str] = json.loads(args.filter_fields)

    print(f'Etsitään kenttiä: {filter_fields}')

    if len(filter_fields) == 0:
        parser.error("Anna vähintään yksi rajoittava kenttä.")

    # Convert path to Path object
    path = Path(args.rip_path)

    # Convert must_contain to set for faster lookups
    must_contain_raw: set[str] = set(args.must_contain)
    print(must_contain_raw)
    must_contain_wild = {term.removesuffix('*').lower() for term in must_contain_raw if term.endswith('*')}
    print(must_contain_wild)
    must_contain = {term.lower() for term in must_contain_raw if not term.endswith('*')}
    # Clean old ripped files before processing
    clean_old_ripped_files(path)

    # First, create _ripped copies of Excel files
    copy_excel_files(path)

    # Process path based on whether it's a file or directory
    if path.is_file():
        # If it's an Excel file, process the _ripped version instead
        if path.suffix.lower() in ['.xlsx', '.xls']:
            process_file(path, must_contain, must_contain_wild, filter_fields)
        else:
            process_file(path, must_contain, must_contain_wild, filter_fields)
    elif path.is_dir():
        process_directory(path, must_contain, must_contain_wild, filter_fields)
    else:
        print(f"Error: {path} is neither a file nor directory", file=sys.stderr)
        sys.exit(1)
    
    # Ask user if they want to replace original files
    while True:
        response = input("\nKorvataanko alkuperäiset tiedostot _ripped-versioilla? (Y/N): ").strip().upper()
        if response in ['Y', 'N']:
            break
        print("Virheellinen vastaus. Anna Y (Kyllä) tai N (Ei).")
    
    if response == 'Y':
        print("\nKorvataan alkuperäiset tiedostot _ripped-versioilla...")
        replace_with_ripped_files(path)
        print("Korvaus valmis!")

if __name__ == "__main__":
    main()
