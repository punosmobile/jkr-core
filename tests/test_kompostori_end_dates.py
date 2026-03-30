import datetime
import os

import pytest
from openpyxl.reader.excel import load_workbook
from sqlalchemy import create_engine, func, text
from sqlalchemy.orm import Session

from jkrimporter import conf
from jkrimporter.cli.jkr import import_lopetusilmoitukset
from jkrimporter.providers.db.database import json_dumps
from jkrimporter.providers.db.models import Kompostori
from jkrimporter.providers.lahti.ilmoitustiedosto import Ilmoitustiedosto


@pytest.fixture(scope="module", autouse=True)
def engine():
    engine = create_engine(
        "postgresql://{username}:{password}@{host}:{port}/{dbname}".format(
            **conf.dbconf
        ),
        future=True,
        json_serializer=json_dumps,
    )
    # Varmistetaan että kohteet ja kohteen_rakennukset ovat olemassa.
    # Jos import_dvv_kohteet on jo luonut ne, ei luoda uudelleen.
    prts = ['134567890B', '100456789B', '123456789A', '200000000A']
    with engine.connect() as conn:
        for prt in prts:
            exists = conn.execute(text("""
                SELECT kr.kohde_id FROM jkr.kohteen_rakennukset kr
                JOIN jkr.rakennus r ON kr.rakennus_id = r.id
                WHERE r.prt = :prt
                LIMIT 1
            """), {"prt": prt}).fetchone()
            if not exists:
                result = conn.execute(text("""
                    INSERT INTO jkr.kohde (alkupvm, loppupvm, kohdetyyppi_id)
                    VALUES ('2021-01-01', NULL, 1)
                    RETURNING id
                """))
                kohde_id = result.fetchone()[0]
                conn.execute(text("""
                    INSERT INTO jkr.kohteen_rakennukset (kohde_id, rakennus_id)
                    VALUES (:kohde_id, (SELECT id FROM jkr.rakennus WHERE prt = :prt))
                """), {"kohde_id": kohde_id, "prt": prt})
        conn.commit()
    return engine


def test_readable(datadir):
    assert Ilmoitustiedosto.readable_by_me(datadir + "/lopetusilmoitukset.xlsx")


def test_lopetusilmoitus(engine, datadir):
    import_lopetusilmoitukset(datadir + "/lopetusilmoitukset.xlsx")
    session = Session(engine)
    end_date = datetime.date(2022, 8, 18)
    # Kahdelle kompostorille asettuu loppupäivämääräksi 18.8.2022.
    assert session.query(func.count(Kompostori.id)).filter(Kompostori.loppupvm == end_date).scalar() == 2

    # Kohdentumattomat tiedostossa kaksi riviä.
    xlsx_file_path = os.path.join(datadir, "kohdentumattomat_lopetusilmoitukset.xlsx")
    workbook = load_workbook(xlsx_file_path)
    sheet = workbook[workbook.sheetnames[0]]
    assert sheet.max_row == 2
