import json
import os

import pytest
from openpyxl.reader.excel import load_workbook
from sqlalchemy import create_engine, func, text
from sqlalchemy.orm import Session

from jkrimporter import conf
from jkrimporter.cli.jkr import import_ilmoitukset
from jkrimporter.providers.db.database import JSONEncoderWithDateSupport, json_dumps
from jkrimporter.providers.db.models import Kompostori, KompostorinKohteet
from jkrimporter.providers.lahti.ilmoitustiedosto import Ilmoitustiedosto


# Redefine json_dumps with test-specific behavior
def json_dumps(value):
    return json.dumps(value, cls=JSONEncoderWithDateSupport)

@pytest.fixture(scope="module", autouse=True)
def engine():
    engine = create_engine(
        "postgresql://{username}:{password}@{host}:{port}/{dbname}".format(
            **conf.dbconf
        ),
        future=True,
        json_serializer=json_dumps,  # Use the test-specific json_dumps function here
    )
    # Varmistetaan että kompostori-importin tarvitsemat kohteet ja
    # kohteen_rakennukset -linkitykset ovat olemassa testikannassa.
    # Jos import_dvv_kohteet on jo luonut ne, ei luoda uudelleen.
    prts = ['134567890B', '100456789B', '123456789A', '200000000A']
    with engine.connect() as conn:
        for prt in prts:
            exists = conn.execute(text("""
                SELECT kr.kohde_id FROM jkr.kohteen_rakennukset kr
                JOIN jkr.rakennus r ON kr.rakennus_id = r.id
                WHERE r.prt = :prt
                LIMIT 1
            """), {"prt": prt}).fetchone()
            if not exists:
                result = conn.execute(text("""
                    INSERT INTO jkr.kohde (alkupvm, loppupvm, kohdetyyppi_id)
                    VALUES ('2021-01-01', NULL, 1)
                    RETURNING id
                """))
                kohde_id = result.fetchone()[0]
                conn.execute(text("""
                    INSERT INTO jkr.kohteen_rakennukset (kohde_id, rakennus_id)
                    VALUES (:kohde_id, (SELECT id FROM jkr.rakennus WHERE prt = :prt))
                """), {"kohde_id": kohde_id, "prt": prt})
        conn.commit()
    return engine

def test_readable(datadir):
    assert Ilmoitustiedosto.readable_by_me(datadir + "/ilmoitukset.xlsx")

def test_kompostori(engine, datadir):
    import_ilmoitukset(datadir + "/ilmoitukset.xlsx")
    session = Session(engine)
    # Ilmoitus.xlsx sisältää 9 riviä, joista kompostoreita syntyy 3:
    # - PRT 134567890B (kimppa rivi 1)
    # - PRT 100456789B (kimppa rivi 2, eri prt → oma kompostori)
    # - PRT 123456789A (yksittäinen)
    # Loput rivit ovat hylättyjä, kohdentumattomia tai virheellisiä.
    assert session.query(func.count(Kompostori.id)).scalar() == 3

    # KompostorinKohteet-taulussa kolme kohdentunutta kohdetta (yksi per kompostori).
    assert session.query(func.count(KompostorinKohteet.kompostori_id)).scalar() == 3

    # Kohdentumattomat.xlsx sisältää neljä kohdentumatonta ilmoitusriviä.
    xlsx_file_path = os.path.join(datadir, "kohdentumattomat_ilmoitukset.xlsx")
    workbook = load_workbook(xlsx_file_path)
    sheet = workbook[workbook.sheetnames[0]]
    assert sheet.max_row == 6

def test_kompostori_osakkaan_lisays(engine, datadir):
    import_ilmoitukset(datadir + "/ilmoitukset_lisaa_komposti_osakas.xlsx")
    session = Session(engine)
    # Tuodaan kaksi riviä lisää:
    # - Rivi 1: PRT 200000000A, kimppa, eri osoite → uusi kompostori
    # - Rivi 2: PRT 123456789A, eri pvm (28.1.2022) → uusi kompostori
    # Yhteensä 3 (edellisestä testistä) + 2 = 5 kompostoria.
    assert session.query(func.count(Kompostori.id)).scalar() == 5

    # KompostorinKohteet-taulussa:
    # 3 (edellisestä testistä) + 1 (200000000A) + 1 (123456789A) = 5 kohdetta.
    assert session.query(func.count(KompostorinKohteet.kompostori_id)).scalar() == 5
