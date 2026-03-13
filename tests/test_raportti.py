"""
Testit raportti-toiminnolle.

Testaa:
- Suodatinparametrien muunnokset (tarkastelupvm, kunta, taajama, kohde_tyyppi, onko_viemari)
- Taajama-suodattimen laskentalogiikka (taajama_10000 ja taajama_200)
- CLI-komennon toiminta oikeilla syötteillä
- Virheenkäsittely virheellisillä päivämäärillä
- Excel-tiedoston luonti
"""

from datetime import date
from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest
from typer.testing import CliRunner

from jkrimporter.cli.jkr import app


def _make_mock_session(columns=None, rows=None):
    """Luo mock-tietokantaistunto, joka palauttaa halutut tulokset."""
    if columns is None:
        columns = ["id", "nimi"]
    if rows is None:
        rows = []

    mock_result = MagicMock()
    mock_result.keys.return_value = columns
    mock_result.fetchall.return_value = rows

    mock_session = MagicMock()
    mock_session.execute.return_value = mock_result
    mock_session.__enter__ = MagicMock(return_value=mock_session)
    mock_session.__exit__ = MagicMock(return_value=False)

    mock_session_factory = MagicMock()
    mock_session_factory.return_value = mock_session

    mock_scoped = MagicMock(return_value=mock_session_factory)

    return mock_scoped, mock_session


runner = CliRunner()


class TestRaporttiTarkastelupvm:
    """Testit tarkastelupäivämäärän parsinnalle."""

    def test_iso_date_parsed_correctly(self, tmp_path):
        """ISO-muotoinen päivämäärä parsitaan oikein."""
        mock_scoped, mock_session = _make_mock_session()
        output_file = tmp_path / "raportti.xlsx"

        with patch("jkrimporter.cli.jkr.scoped_session", mock_scoped), patch(
            "jkrimporter.cli.jkr.engine", MagicMock()
        ):
            result = runner.invoke(
                app, ["raportti", str(output_file), "2024-06-15", "0", "0", "0", "0", "0"]
            )

        call_args = mock_session.execute.call_args
        params = call_args[0][1]
        assert params["tarkastelupvm"] == date(2024, 6, 15)

    def test_finnish_date_parsed_correctly(self, tmp_path):
        """Suomalainen päivämäärämuoto parsitaan oikein."""
        mock_scoped, mock_session = _make_mock_session()
        output_file = tmp_path / "raportti.xlsx"

        with patch("jkrimporter.cli.jkr.scoped_session", mock_scoped), patch(
            "jkrimporter.cli.jkr.engine", MagicMock()
        ):
            result = runner.invoke(
                app, ["raportti", str(output_file), "15.6.2024", "0", "0", "0", "0", "0"]
            )

        call_args = mock_session.execute.call_args
        params = call_args[0][1]
        assert params["tarkastelupvm"] == date(2024, 6, 15)

    def test_zero_tarkastelupvm_gives_none(self, tmp_path):
        """Arvo '0' tarkastelupvm-parametrille antaa None."""
        mock_scoped, mock_session = _make_mock_session()
        output_file = tmp_path / "raportti.xlsx"

        with patch("jkrimporter.cli.jkr.scoped_session", mock_scoped), patch(
            "jkrimporter.cli.jkr.engine", MagicMock()
        ):
            result = runner.invoke(
                app, ["raportti", str(output_file), "0", "0", "0", "0", "0", "0"]
            )

        call_args = mock_session.execute.call_args
        params = call_args[0][1]
        assert params["tarkastelupvm"] is None

    def test_invalid_date_returns_error(self, tmp_path):
        """Virheellinen päivämäärä aiheuttaa virhepoistumisen."""
        mock_scoped, mock_session = _make_mock_session()
        output_file = tmp_path / "raportti.xlsx"

        with patch("jkrimporter.cli.jkr.scoped_session", mock_scoped), patch(
            "jkrimporter.cli.jkr.engine", MagicMock()
        ):
            result = runner.invoke(
                app, ["raportti", str(output_file), "ei-paivamaara", "0", "0", "0", "0", "0"]
            )

        assert result.exit_code == 1

    def test_invalid_date_format_returns_error(self, tmp_path):
        """Väärässä muodossa oleva päivämäärä aiheuttaa virhepoistumisen."""
        mock_scoped, mock_session = _make_mock_session()
        output_file = tmp_path / "raportti.xlsx"

        with patch("jkrimporter.cli.jkr.scoped_session", mock_scoped), patch(
            "jkrimporter.cli.jkr.engine", MagicMock()
        ):
            result = runner.invoke(
                app, ["raportti", str(output_file), "2024/06/15", "0", "0", "0", "0", "0"]
            )

        assert result.exit_code == 1


class TestRaporttiKuntaFilter:
    """Testit kunta-suodattimen muunnokselle."""

    def test_kunta_zero_string_gives_none(self, tmp_path):
        """Kunta-arvo '0' antaa None-suodattimen."""
        mock_scoped, mock_session = _make_mock_session()
        output_file = tmp_path / "raportti.xlsx"

        with patch("jkrimporter.cli.jkr.scoped_session", mock_scoped), patch(
            "jkrimporter.cli.jkr.engine", MagicMock()
        ):
            result = runner.invoke(
                app, ["raportti", str(output_file), "2024-01-01", "0", "0", "0", "0", "0"]
            )

        call_args = mock_session.execute.call_args
        params = call_args[0][1]
        assert params["kunta"] is None

    def test_kunta_name_passed_as_is(self, tmp_path):
        """Kunnan nimi välitetään muuttumattomana."""
        mock_scoped, mock_session = _make_mock_session()
        output_file = tmp_path / "raportti.xlsx"

        with patch("jkrimporter.cli.jkr.scoped_session", mock_scoped), patch(
            "jkrimporter.cli.jkr.engine", MagicMock()
        ):
            result = runner.invoke(
                app, ["raportti", str(output_file), "2024-01-01", "Lahti", "0", "0", "0", "0"]
            )

        call_args = mock_session.execute.call_args
        params = call_args[0][1]
        assert params["kunta"] == "Lahti"

    def test_kunta_other_municipality(self, tmp_path):
        """Muu kunnan nimi välitetään oikein."""
        mock_scoped, mock_session = _make_mock_session()
        output_file = tmp_path / "raportti.xlsx"

        with patch("jkrimporter.cli.jkr.scoped_session", mock_scoped), patch(
            "jkrimporter.cli.jkr.engine", MagicMock()
        ):
            result = runner.invoke(
                app, ["raportti", str(output_file), "2024-01-01", "Helsinki", "0", "0", "0", "0"]
            )

        call_args = mock_session.execute.call_args
        params = call_args[0][1]
        assert params["kunta"] == "Helsinki"


class TestRaporttiOnkoViemariFilter:
    """Testit viemäriliitossuodattimen muunnokselle."""

    def test_onko_viemari_zero_gives_none(self, tmp_path):
        """Arvo 0 antaa None (ei rajausta)."""
        mock_scoped, mock_session = _make_mock_session()
        output_file = tmp_path / "raportti.xlsx"

        with patch("jkrimporter.cli.jkr.scoped_session", mock_scoped), patch(
            "jkrimporter.cli.jkr.engine", MagicMock()
        ):
            result = runner.invoke(
                app, ["raportti", str(output_file), "2024-01-01", "0", "0", "0", "0", "0"]
            )

        call_args = mock_session.execute.call_args
        params = call_args[0][1]
        assert params["onko_viemari"] is None

    def test_onko_viemari_one_gives_true(self, tmp_path):
        """Arvo 1 antaa True (viemäriverkossa)."""
        mock_scoped, mock_session = _make_mock_session()
        output_file = tmp_path / "raportti.xlsx"

        with patch("jkrimporter.cli.jkr.scoped_session", mock_scoped), patch(
            "jkrimporter.cli.jkr.engine", MagicMock()
        ):
            result = runner.invoke(
                app, ["raportti", str(output_file), "2024-01-01", "0", "0", "0", "0", "1"]
            )

        call_args = mock_session.execute.call_args
        params = call_args[0][1]
        assert params["onko_viemari"] is True

    def test_onko_viemari_two_gives_false(self, tmp_path):
        """Arvo 2 antaa False (ei viemäriverkossa)."""
        mock_scoped, mock_session = _make_mock_session()
        output_file = tmp_path / "raportti.xlsx"

        with patch("jkrimporter.cli.jkr.scoped_session", mock_scoped), patch(
            "jkrimporter.cli.jkr.engine", MagicMock()
        ):
            result = runner.invoke(
                app, ["raportti", str(output_file), "2024-01-01", "0", "0", "0", "0", "2"]
            )

        call_args = mock_session.execute.call_args
        params = call_args[0][1]
        assert params["onko_viemari"] is False


class TestRaporttiKohdetyyppiFilter:
    """Testit kohdetyypin suodattimen muunnokselle."""

    def test_kohdetyyppi_zero_gives_none(self, tmp_path):
        """Arvo 0 antaa None (ei rajausta)."""
        mock_scoped, mock_session = _make_mock_session()
        output_file = tmp_path / "raportti.xlsx"

        with patch("jkrimporter.cli.jkr.scoped_session", mock_scoped), patch(
            "jkrimporter.cli.jkr.engine", MagicMock()
        ):
            result = runner.invoke(
                app, ["raportti", str(output_file), "2024-01-01", "0", "0", "0", "0", "0"]
            )

        call_args = mock_session.execute.call_args
        params = call_args[0][1]
        assert params["kohde_tyyppi_id"] is None

    def test_kohdetyyppi_hapa(self, tmp_path):
        """Kohdetyyppi 5 (hapa) välitetään oikein."""
        mock_scoped, mock_session = _make_mock_session()
        output_file = tmp_path / "raportti.xlsx"

        with patch("jkrimporter.cli.jkr.scoped_session", mock_scoped), patch(
            "jkrimporter.cli.jkr.engine", MagicMock()
        ):
            result = runner.invoke(
                app, ["raportti", str(output_file), "2024-01-01", "0", "0", "0", "5", "0"]
            )

        call_args = mock_session.execute.call_args
        params = call_args[0][1]
        assert params["kohde_tyyppi_id"] == 5

    def test_kohdetyyppi_biohapa(self, tmp_path):
        """Kohdetyyppi 6 (biohapa) välitetään oikein."""
        mock_scoped, mock_session = _make_mock_session()
        output_file = tmp_path / "raportti.xlsx"

        with patch("jkrimporter.cli.jkr.scoped_session", mock_scoped), patch(
            "jkrimporter.cli.jkr.engine", MagicMock()
        ):
            result = runner.invoke(
                app, ["raportti", str(output_file), "2024-01-01", "0", "0", "0", "6", "0"]
            )

        call_args = mock_session.execute.call_args
        params = call_args[0][1]
        assert params["kohde_tyyppi_id"] == 6

    def test_kohdetyyppi_asuinkiinteisto(self, tmp_path):
        """Kohdetyyppi 7 (asuinkiinteistö) välitetään oikein."""
        mock_scoped, mock_session = _make_mock_session()
        output_file = tmp_path / "raportti.xlsx"

        with patch("jkrimporter.cli.jkr.scoped_session", mock_scoped), patch(
            "jkrimporter.cli.jkr.engine", MagicMock()
        ):
            result = runner.invoke(
                app, ["raportti", str(output_file), "2024-01-01", "0", "0", "0", "7", "0"]
            )

        call_args = mock_session.execute.call_args
        params = call_args[0][1]
        assert params["kohde_tyyppi_id"] == 7

    def test_kohdetyyppi_muu(self, tmp_path):
        """Kohdetyyppi 8 (muu) välitetään oikein."""
        mock_scoped, mock_session = _make_mock_session()
        output_file = tmp_path / "raportti.xlsx"

        with patch("jkrimporter.cli.jkr.scoped_session", mock_scoped), patch(
            "jkrimporter.cli.jkr.engine", MagicMock()
        ):
            result = runner.invoke(
                app, ["raportti", str(output_file), "2024-01-01", "0", "0", "0", "8", "0"]
            )

        call_args = mock_session.execute.call_args
        params = call_args[0][1]
        assert params["kohde_tyyppi_id"] == 8


class TestRaporttiTaajamaFilter:
    """Testit taajama-suodattimen laskentalogiikalle.

    Taajama-parametrin arvot:
    - 0 / None = ei rajausta
    - 1 = yli 200 asukaan taajama
    - 2 = yli 10000 asukaan taajama
    - 3 = molemmat (yli 200 ja yli 10000)
    - 200 = yli 200 asukaan taajama
    - 10000 = yli 10000 asukaan taajama
    """

    def test_taajama_zero_gives_both_none(self, tmp_path):
        """Taajama-arvo 0 antaa molemmat None."""
        mock_scoped, mock_session = _make_mock_session()
        output_file = tmp_path / "raportti.xlsx"

        with patch("jkrimporter.cli.jkr.scoped_session", mock_scoped), patch(
            "jkrimporter.cli.jkr.engine", MagicMock()
        ):
            result = runner.invoke(
                app, ["raportti", str(output_file), "2024-01-01", "0", "0", "0", "0", "0"]
            )

        call_args = mock_session.execute.call_args
        params = call_args[0][1]
        assert params["taajama_10000"] is None
        assert params["taajama_200"] is None

    def test_taajama_2_gives_10000_true_200_none(self, tmp_path):
        """Taajama-arvo 2 antaa taajama_10000=True ja taajama_200=None."""
        mock_scoped, mock_session = _make_mock_session()
        output_file = tmp_path / "raportti.xlsx"

        with patch("jkrimporter.cli.jkr.scoped_session", mock_scoped), patch(
            "jkrimporter.cli.jkr.engine", MagicMock()
        ):
            result = runner.invoke(
                app, ["raportti", str(output_file), "2024-01-01", "0", "0", "2", "0", "0"]
            )

        call_args = mock_session.execute.call_args
        params = call_args[0][1]
        assert params["taajama_10000"] is True
        assert params["taajama_200"] is None

    def test_taajama_1_gives_10000_none_200_true(self, tmp_path):
        """Taajama-arvo 1 antaa taajama_10000=None ja taajama_200=True."""
        mock_scoped, mock_session = _make_mock_session()
        output_file = tmp_path / "raportti.xlsx"

        with patch("jkrimporter.cli.jkr.scoped_session", mock_scoped), patch(
            "jkrimporter.cli.jkr.engine", MagicMock()
        ):
            result = runner.invoke(
                app, ["raportti", str(output_file), "2024-01-01", "0", "0", "1", "0", "0"]
            )

        call_args = mock_session.execute.call_args
        params = call_args[0][1]
        assert params["taajama_10000"] is None
        assert params["taajama_200"] is True

    def test_taajama_3_gives_both_true(self, tmp_path):
        """Taajama-arvo 3 antaa molemmat True."""
        mock_scoped, mock_session = _make_mock_session()
        output_file = tmp_path / "raportti.xlsx"

        with patch("jkrimporter.cli.jkr.scoped_session", mock_scoped), patch(
            "jkrimporter.cli.jkr.engine", MagicMock()
        ):
            result = runner.invoke(
                app, ["raportti", str(output_file), "2024-01-01", "0", "0", "3", "0", "0"]
            )

        call_args = mock_session.execute.call_args
        params = call_args[0][1]
        assert params["taajama_10000"] is True
        assert params["taajama_200"] is True

    def test_taajama_10000_gives_10000_true_200_none(self, tmp_path):
        """Taajama-arvo 10000 antaa taajama_10000=True ja taajama_200=None."""
        mock_scoped, mock_session = _make_mock_session()
        output_file = tmp_path / "raportti.xlsx"

        with patch("jkrimporter.cli.jkr.scoped_session", mock_scoped), patch(
            "jkrimporter.cli.jkr.engine", MagicMock()
        ):
            result = runner.invoke(
                app, ["raportti", str(output_file), "2024-01-01", "0", "0", "10000", "0", "0"]
            )

        call_args = mock_session.execute.call_args
        params = call_args[0][1]
        assert params["taajama_10000"] is True
        assert params["taajama_200"] is None

    def test_taajama_200_gives_10000_none_200_true(self, tmp_path):
        """Taajama-arvo 200 antaa taajama_10000=None ja taajama_200=True."""
        mock_scoped, mock_session = _make_mock_session()
        output_file = tmp_path / "raportti.xlsx"

        with patch("jkrimporter.cli.jkr.scoped_session", mock_scoped), patch(
            "jkrimporter.cli.jkr.engine", MagicMock()
        ):
            result = runner.invoke(
                app, ["raportti", str(output_file), "2024-01-01", "0", "0", "200", "0", "0"]
            )

        call_args = mock_session.execute.call_args
        params = call_args[0][1]
        assert params["taajama_10000"] is None
        assert params["taajama_200"] is True


class TestRaporttiHuoneistomaara:
    """Testit huoneistomäärä-parametrin välitykselle."""

    def test_huoneistomaara_zero_passed_as_zero(self, tmp_path):
        """Huoneistomäärä 0 välitetään nollana (ei rajausta)."""
        mock_scoped, mock_session = _make_mock_session()
        output_file = tmp_path / "raportti.xlsx"

        with patch("jkrimporter.cli.jkr.scoped_session", mock_scoped), patch(
            "jkrimporter.cli.jkr.engine", MagicMock()
        ):
            result = runner.invoke(
                app, ["raportti", str(output_file), "2024-01-01", "0", "0", "0", "0", "0"]
            )

        call_args = mock_session.execute.call_args
        params = call_args[0][1]
        assert params["huoneistomaara"] == 0

    def test_huoneistomaara_four_passed_as_four(self, tmp_path):
        """Huoneistomäärä 4 välitetään nelosena (enintään 4 huoneistoa)."""
        mock_scoped, mock_session = _make_mock_session()
        output_file = tmp_path / "raportti.xlsx"

        with patch("jkrimporter.cli.jkr.scoped_session", mock_scoped), patch(
            "jkrimporter.cli.jkr.engine", MagicMock()
        ):
            result = runner.invoke(
                app, ["raportti", str(output_file), "2024-01-01", "0", "4", "0", "0", "0"]
            )

        call_args = mock_session.execute.call_args
        params = call_args[0][1]
        assert params["huoneistomaara"] == 4

    def test_huoneistomaara_five_passed_as_five(self, tmp_path):
        """Huoneistomäärä 5 välitetään viitosena (vähintään 5 huoneistoa)."""
        mock_scoped, mock_session = _make_mock_session()
        output_file = tmp_path / "raportti.xlsx"

        with patch("jkrimporter.cli.jkr.scoped_session", mock_scoped), patch(
            "jkrimporter.cli.jkr.engine", MagicMock()
        ):
            result = runner.invoke(
                app, ["raportti", str(output_file), "2024-01-01", "0", "5", "0", "0", "0"]
            )

        call_args = mock_session.execute.call_args
        params = call_args[0][1]
        assert params["huoneistomaara"] == 5


class TestRaporttiExcelOutput:
    """Testit Excel-tiedoston luonnille."""

    def test_excel_file_created(self, tmp_path):
        """Excel-tiedosto luodaan onnistuneesti."""
        columns = ["Kohde_id", "Nimi", "Osoite"]
        rows = [(1, "Testi Kohde", "Testikatu 1")]
        mock_scoped, mock_session = _make_mock_session(columns=columns, rows=rows)
        output_file = tmp_path / "raportti.xlsx"

        with patch("jkrimporter.cli.jkr.scoped_session", mock_scoped), patch(
            "jkrimporter.cli.jkr.engine", MagicMock()
        ):
            result = runner.invoke(
                app, ["raportti", str(output_file), "2024-01-01", "0", "0", "0", "0", "0"]
            )

        assert result.exit_code == 0
        assert output_file.exists()

    def test_excel_file_has_correct_columns(self, tmp_path):
        """Excel-tiedostossa on oikeat sarakkeet."""
        import openpyxl

        columns = ["Kohde_id", "Nimi", "Osoite"]
        rows = [(1, "Testi Kohde", "Testikatu 1")]
        mock_scoped, mock_session = _make_mock_session(columns=columns, rows=rows)
        output_file = tmp_path / "raportti.xlsx"

        with patch("jkrimporter.cli.jkr.scoped_session", mock_scoped), patch(
            "jkrimporter.cli.jkr.engine", MagicMock()
        ):
            result = runner.invoke(
                app, ["raportti", str(output_file), "2024-01-01", "0", "0", "0", "0", "0"]
            )

        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        header_row = [cell.value for cell in ws[1]]
        assert header_row == ["Kohde_id", "Nimi", "Osoite"]

    def test_excel_file_has_correct_data(self, tmp_path):
        """Excel-tiedostossa on oikeat tiedot."""
        import openpyxl

        columns = ["Kohde_id", "Nimi"]
        rows = [(42, "Testi"), (99, "Toinen")]
        mock_scoped, mock_session = _make_mock_session(columns=columns, rows=rows)
        output_file = tmp_path / "raportti.xlsx"

        with patch("jkrimporter.cli.jkr.scoped_session", mock_scoped), patch(
            "jkrimporter.cli.jkr.engine", MagicMock()
        ):
            result = runner.invoke(
                app, ["raportti", str(output_file), "2024-01-01", "0", "0", "0", "0", "0"]
            )

        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        # Rivi 1 = otsikko, rivi 2 = ensimmäinen datarivi
        assert ws.cell(row=2, column=1).value == 42
        assert ws.cell(row=2, column=2).value == "Testi"
        assert ws.cell(row=3, column=1).value == 99
        assert ws.cell(row=3, column=2).value == "Toinen"

    def test_excel_file_empty_result(self, tmp_path):
        """Excel-tiedosto luodaan myös tyhjällä tuloksella."""
        mock_scoped, mock_session = _make_mock_session(columns=["id"], rows=[])
        output_file = tmp_path / "raportti.xlsx"

        with patch("jkrimporter.cli.jkr.scoped_session", mock_scoped), patch(
            "jkrimporter.cli.jkr.engine", MagicMock()
        ):
            result = runner.invoke(
                app, ["raportti", str(output_file), "2024-01-01", "0", "0", "0", "0", "0"]
            )

        assert result.exit_code == 0
        assert output_file.exists()

    def test_success_message_printed(self, tmp_path):
        """Onnistumisilmoitus tulostetaan."""
        mock_scoped, mock_session = _make_mock_session()
        output_file = tmp_path / "raportti.xlsx"

        with patch("jkrimporter.cli.jkr.scoped_session", mock_scoped), patch(
            "jkrimporter.cli.jkr.engine", MagicMock()
        ):
            result = runner.invoke(
                app, ["raportti", str(output_file), "2024-01-01", "0", "0", "0", "0", "0"]
            )

        assert result.exit_code == 0
        assert "onnistuneesti" in result.output


class TestRaporttiKombinoidutSuodattimet:
    """Testit useampien suodattimien yhdistelmille."""

    def test_all_filters_set(self, tmp_path):
        """Kaikki suodattimet asetettu kerralla."""
        mock_scoped, mock_session = _make_mock_session()
        output_file = tmp_path / "raportti.xlsx"

        with patch("jkrimporter.cli.jkr.scoped_session", mock_scoped), patch(
            "jkrimporter.cli.jkr.engine", MagicMock()
        ):
            result = runner.invoke(
                app,
                [
                    "raportti",
                    str(output_file),
                    "2024-06-01",
                    "Lahti",
                    "4",
                    "3",
                    "7",
                    "1",
                ],
            )

        call_args = mock_session.execute.call_args
        params = call_args[0][1]
        assert params["tarkastelupvm"] == date(2024, 6, 1)
        assert params["kunta"] == "Lahti"
        assert params["huoneistomaara"] == 4
        assert params["taajama_10000"] is True
        assert params["taajama_200"] is True
        assert params["kohde_tyyppi_id"] == 7
        assert params["onko_viemari"] is True

    def test_no_filters_all_none(self, tmp_path):
        """Ilman suodattimia kaikki arvot None/0."""
        mock_scoped, mock_session = _make_mock_session()
        output_file = tmp_path / "raportti.xlsx"

        with patch("jkrimporter.cli.jkr.scoped_session", mock_scoped), patch(
            "jkrimporter.cli.jkr.engine", MagicMock()
        ):
            result = runner.invoke(
                app, ["raportti", str(output_file), "0", "0", "0", "0", "0", "0"]
            )

        call_args = mock_session.execute.call_args
        params = call_args[0][1]
        assert params["tarkastelupvm"] is None
        assert params["kunta"] is None
        assert params["huoneistomaara"] == 0
        assert params["taajama_10000"] is None
        assert params["taajama_200"] is None
        assert params["kohde_tyyppi_id"] is None
        assert params["onko_viemari"] is None

    def test_kunta_with_viemari_filter(self, tmp_path):
        """Kunta ja viemärisuodatin yhdistettynä."""
        mock_scoped, mock_session = _make_mock_session()
        output_file = tmp_path / "raportti.xlsx"

        with patch("jkrimporter.cli.jkr.scoped_session", mock_scoped), patch(
            "jkrimporter.cli.jkr.engine", MagicMock()
        ):
            result = runner.invoke(
                app,
                [
                    "raportti",
                    str(output_file),
                    "2024-01-01",
                    "Lahti",
                    "0",
                    "0",
                    "0",
                    "2",
                ],
            )

        call_args = mock_session.execute.call_args
        params = call_args[0][1]
        assert params["kunta"] == "Lahti"
        assert params["onko_viemari"] is False

    def test_kohdetyyppi_with_taajama_filter(self, tmp_path):
        """Kohdetyyppi ja taajama yhdistettynä."""
        mock_scoped, mock_session = _make_mock_session()
        output_file = tmp_path / "raportti.xlsx"

        with patch("jkrimporter.cli.jkr.scoped_session", mock_scoped), patch(
            "jkrimporter.cli.jkr.engine", MagicMock()
        ):
            result = runner.invoke(
                app,
                [
                    "raportti",
                    str(output_file),
                    "2024-01-01",
                    "0",
                    "0",
                    "2",
                    "5",
                    "0",
                ],
            )

        call_args = mock_session.execute.call_args
        params = call_args[0][1]
        assert params["kohde_tyyppi_id"] == 5
        assert params["taajama_10000"] is True
        assert params["taajama_200"] is None


class TestRaporttiSqlQuery:
    """Testit SQL-kyselyn rakenteelle."""

    def test_print_report_function_called(self, tmp_path):
        """print_report-funktio kutsutaan tietokannasta."""
        mock_scoped, mock_session = _make_mock_session()
        output_file = tmp_path / "raportti.xlsx"

        with patch("jkrimporter.cli.jkr.scoped_session", mock_scoped), patch(
            "jkrimporter.cli.jkr.engine", MagicMock()
        ):
            result = runner.invoke(
                app, ["raportti", str(output_file), "2024-01-01", "0", "0", "0", "0", "0"]
            )

        assert mock_session.execute.called
        call_args = mock_session.execute.call_args
        sql_text = str(call_args[0][0])
        assert "print_report" in sql_text

    def test_all_required_params_passed_to_sql(self, tmp_path):
        """Kaikki vaaditut parametrit välitetään SQL-kyselylle."""
        mock_scoped, mock_session = _make_mock_session()
        output_file = tmp_path / "raportti.xlsx"

        with patch("jkrimporter.cli.jkr.scoped_session", mock_scoped), patch(
            "jkrimporter.cli.jkr.engine", MagicMock()
        ):
            result = runner.invoke(
                app, ["raportti", str(output_file), "2024-01-01", "0", "0", "0", "0", "0"]
            )

        call_args = mock_session.execute.call_args
        params = call_args[0][1]
        expected_keys = {
            "tarkastelupvm",
            "kunta",
            "huoneistomaara",
            "taajama_10000",
            "taajama_200",
            "kohde_tyyppi_id",
            "onko_viemari",
        }
        assert expected_keys == set(params.keys())
